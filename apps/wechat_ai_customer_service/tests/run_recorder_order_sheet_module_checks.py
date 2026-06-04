"""Checks for order_sheet_lab_v1 module extraction and export."""

from __future__ import annotations

import json
import os
import shutil
import sys
import csv
from pathlib import Path
from typing import Any

os.environ.setdefault("WECHAT_STORAGE_BACKEND", "file")
os.environ.setdefault("WECHAT_CLOUD_REQUIRED", "0")
os.environ.setdefault("WECHAT_CLOUD_STRICT_ONLINE", "0")

from fastapi.testclient import TestClient
from openpyxl import load_workbook


APP_ROOT = Path(__file__).resolve().parents[1]
PROJECT_ROOT = APP_ROOT.parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from apps.wechat_ai_customer_service.admin_backend.app import create_app  # noqa: E402
from apps.wechat_ai_customer_service.admin_backend.services import recorder_export_run_service as export_run_module  # noqa: E402
from apps.wechat_ai_customer_service.admin_backend.services.background_handlers import handle_recorder_export_run_execute  # noqa: E402
from apps.wechat_ai_customer_service.admin_backend.services.recorder_export_run_service import RecorderExportRunService  # noqa: E402
from apps.wechat_ai_customer_service.admin_backend.services.work_queue import WorkQueueService  # noqa: E402
from apps.wechat_ai_customer_service.knowledge_paths import tenant_context, tenant_runtime_root  # noqa: E402


TEST_TENANT = "recorder_order_sheet_test"
ORDER_HEADERS = [
    None,
    "日期",
    "时间",
    "姓名",
    "责任人（老板）",
    "收货人",
    "校区",
    "订货单位",
    "货品名称",
    "数量",
    "单位",
    "规格",
    "品牌",
    "进价（单价）",
    "售价（单价）",
    "总进价",
    "总售价",
    "备注",
]


def main() -> int:
    try:
        with tenant_context(TEST_TENANT):
            cleanup_runtime()
            run_checks()
        print(json.dumps({"ok": True, "tenant": TEST_TENANT}, ensure_ascii=False, indent=2))
        return 0
    finally:
        with tenant_context(TEST_TENANT):
            cleanup_runtime()


def run_checks() -> None:
    original_call_deepseek_json = export_run_module.call_deepseek_json
    export_run_module.call_deepseek_json = lambda _prompt: {}
    try:
        client = TestClient(create_app())
        headers = {"X-Tenant-ID": TEST_TENANT}
        seed_order_messages(client, headers)
        bind_module(client, headers)
        run_id = create_run(client, headers)
        execute_run(run_id)
        validate_export(client, headers, run_id)
        validate_date_range_runs(client, headers)
        validate_name_context_inference_rules()
        validate_name_owner_price_prefix_cleanup()
        validate_ocr_broken_line_order_rules()
        validate_brand_single_use_rules()
        validate_close_duplicate_export_row_dedupe()
        validate_message_envelope_quote_timestamp_and_risk_flags()
    finally:
        export_run_module.call_deepseek_json = original_call_deepseek_json


def seed_order_messages(client: TestClient, headers: dict[str, str]) -> None:
    payload = {
        "conversation": {
            "target_name": "企点售后群",
            "display_name": "企点售后群",
            "conversation_type": "group",
            "selected_by_user": True,
            "learning_enabled": True,
            "notify_enabled": False,
            "status": "active",
        },
        "messages": [
            {
                "id": "order_sheet_001",
                "type": "text",
                "sender": "罗永志 企点",
                "content": "叶松霖-陈秋平老师\n订：甲醛试剂=甲醛测试盒 2盒 98元",
                "time": "2025-03-03 09:12:10",
            },
            {
                "id": "order_sheet_002",
                "type": "text",
                "sender": "罗永志 企点",
                "content": "王一博-李老师\n代付：水质试纸 3盒 45元",
                "time": "2025-03-03 10:20:33",
            },
            {
                "id": "order_sheet_003",
                "type": "quote",
                "sender": "罗永志 企点",
                "content": "李华-张老师\n订：PH试剂 1盒 120元",
                "time": "2025-03-04 08:00:00",
            },
            {
                "id": "order_sheet_004",
                "type": "text",
                "sender": "罗永志 企点",
                "content": "赵宁-王老师\nHEPES缓冲液 订 99元",
                "time": "2025-03-04 08:05:00",
            },
            {
                "id": "order_sheet_005",
                "type": "text",
                "sender": "罗永志 企点",
                "content": "是的2瓶",
                "time": "2025-03-04 08:05:20",
            },
            {
                "id": "order_sheet_006",
                "type": "text",
                "sender": "罗永志 企点",
                "content": "李华-张老师\n赠品：一次性手套 1盒",
                "time": "2025-03-04 08:06:00",
            },
            {
                "id": "order_sheet_007",
                "type": "text",
                "sender": "罗永志 企点",
                "content": "源叶\n羟基酪醇    S25716-1g   384*0.9=345.6元\n秦皮乙素    S31424-1g   100*0.9=90元\n赵伟睿老师",
                "time": "2025-03-04 08:10:00",
            },
            {
                "id": "order_sheet_008",
                "type": "text",
                "sender": "罗永志 企点",
                "content": "白鲨 50ml离心管 订2包 60元",
                "time": "2025-03-04 08:12:00",
            },
            {
                "id": "order_sheet_009",
                "type": "text",
                "sender": "罗永志 企点",
                "content": "杰乐普 透析袋MD77 8000-14000 1米 238元",
                "time": "2025-03-04 08:15:00",
            },
            {
                "id": "order_sheet_010",
                "type": "text",
                "sender": "罗永志 企点",
                "content": "周梦老师",
                "time": "2025-03-04 08:20:00",
            },
            {
                "id": "order_sheet_011",
                "type": "text",
                "sender": "罗永志 企点",
                "content": "白鲨 15ml离心管 2包 66元",
                "time": "2025-03-04 08:22:00",
            },
            {
                "id": "order_sheet_012",
                "type": "text",
                "sender": "罗永志 企点",
                "content": "@记录员 韩梅老师 请记录 枪头 1盒 20元",
                "time": "2025-03-04 08:23:00",
            },
            {
                "id": "order_sheet_013",
                "type": "text",
                "sender": "罗永志 企点",
                "content": "PBS缓冲液 1瓶 30元",
                "time": "2025-03-04 08:24:00",
            },
            {
                "id": "order_sheet_014",
                "type": "text",
                "sender": "罗永志 企点",
                "content": "葡萄糖 1瓶 10元",
                "time": "2025-03-04 08:28:10",
            },
            {
                "id": "order_sheet_015",
                "type": "text",
                "sender": "罗永志 企点",
                "content": "思科捷\nCY-09（NLRP3抑制剂）  SJ-MX0745\n10 mM * 1mL in DMSO  订1个  700*0.95=665元\n俞\n淑芳老师",
                "time": "2025-03-04 08:32:00",
            },
        ],
        "source_module": "order_sheet_test",
        "auto_learn": False,
    }
    response = client.post("/api/raw-messages/messages", headers=headers, json=payload)
    assert_equal(response.status_code, 200, "seed raw messages status")
    result = response.json()
    assert_true(result.get("ok") is True, "seed raw messages ok")
    assert_true(int(result.get("inserted_count", 0) or 0) >= 2, "seed raw messages inserted")


def bind_module(client: TestClient, headers: dict[str, str]) -> None:
    payload = {
        "binding_id": "order_sheet_local_admin_binding",
        "scope_type": "user",
        "scope_id": "local-admin",
        "user_id": "local-admin",
        "module_key": "order_sheet_lab_v1",
        "enabled": True,
    }
    response = client.post("/api/recorder/module-bindings", headers=headers, json=payload)
    assert_equal(response.status_code, 200, "bind module status")
    assert_true(response.json().get("ok") is True, "bind module ok")


def create_run(client: TestClient, headers: dict[str, str]) -> str:
    payload = {
        "target_names": ["企点售后群"],
        "limit": 500,
    }
    return create_run_with_payload(client, headers, payload)


def create_run_with_payload(client: TestClient, headers: dict[str, str], payload: dict[str, Any]) -> str:
    response = client.post("/api/recorder/exports/runs", headers=headers, json=payload)
    assert_equal(response.status_code, 200, "create run status")
    item = response.json().get("item", {})
    run_id = str(item.get("run_id") or "")
    assert_true(run_id.startswith("recorder_export_run_"), "run id generated")
    assert_equal(str(item.get("module_key") or ""), "order_sheet_lab_v1", "resolved module should be order_sheet_lab_v1")
    return run_id


def execute_run(run_id: str) -> None:
    queue = WorkQueueService(tenant_id=TEST_TENANT)
    target_job = None
    for _ in range(20):
        claimed = queue.claim(queue="recorder_exports", worker_id="order-sheet-test", limit=50, lock_seconds=120)
        target_job = next((job for job in claimed if str((job.get("payload") or {}).get("run_id") or "") == run_id), None)
        if target_job is not None:
            break
    assert_true(target_job is not None, "export run job should be claimable")
    payload = target_job.get("payload") if isinstance(target_job.get("payload"), dict) else {}
    result = handle_recorder_export_run_execute(payload)
    if result.get("ok"):
        queue.complete(str(target_job.get("job_id") or ""), result=result)
        return
    queue.fail(str(target_job.get("job_id") or ""), error=str(result.get("error") or "run failed"), retry=False)
    raise AssertionError(f"order sheet export failed: {result}")


def validate_export(client: TestClient, headers: dict[str, str], run_id: str) -> None:
    response = client.get(f"/api/recorder/exports/runs/{run_id}", headers=headers)
    assert_equal(response.status_code, 200, "run detail status")
    item = response.json().get("item", {})
    assert_equal(item.get("status"), "succeeded", "run status should be succeeded")
    assert_equal(str(item.get("stage") or ""), "completed", "run stage should be completed")
    stats = item.get("stats") if isinstance(item.get("stats"), dict) else {}
    assert_true(int(stats.get("export_row_count", 0) or 0) >= 2, "export row count should be >= 2")
    assert_true(int(stats.get("llm_calls", 0) or 0) >= 0, "llm call stats should exist")
    progress = item.get("progress") if isinstance(item.get("progress"), dict) else {}
    assert_true(float(progress.get("percent", 0) or 0) >= 1.0, "completed run progress should reach 100%")
    artifacts = item.get("artifacts") if isinstance(item.get("artifacts"), dict) else {}
    workbook_path = Path(str(artifacts.get("xlsx_path") or ""))
    csv_path = Path(str(artifacts.get("csv_path") or ""))
    report_path = Path(str(artifacts.get("report_path") or ""))
    assert_true(workbook_path.exists(), "xlsx exists")
    assert_true(csv_path.exists(), "csv exists")
    assert_true(report_path.exists(), "report exists")
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        csv_rows = list(csv.reader(handle))
    assert_true(len(csv_rows) >= 3, "csv should contain header and exported rows")
    assert_equal(csv_rows[0], ["" if item is None else item for item in ORDER_HEADERS], "csv header should match expected template")

    workbook = load_workbook(workbook_path)
    sheet = workbook["Sheet1"]
    headers_row = [sheet.cell(row=1, column=index + 1).value for index in range(len(ORDER_HEADERS))]
    assert_equal(headers_row, ORDER_HEADERS, "order sheet header should match expected template")
    first_date = str(sheet.cell(row=2, column=2).value or "")
    first_time = str(sheet.cell(row=2, column=3).value or "")
    first_name = str(sheet.cell(row=2, column=4).value or "")
    first_product = str(sheet.cell(row=2, column=9).value or "")
    assert_true(first_date in {"2025-03-03", "2025-03-04"}, "date should be YYYY-MM-DD text")
    assert_true(bool(first_time) and len(first_time.split(":")) == 3, f"time should be HH:MM:SS text, got {first_time!r}")
    assert_true(bool(first_name), "name should be extracted")
    assert_true(bool(first_product), "product name should be extracted")
    hep_es_rows = []
    gift_rows = []
    yuanye_rows = []
    baisha_rows = []
    jielepu_rows = []
    sikejie_rows = []
    context_name_rows = []
    late_rows = []
    mention_rows = []
    top_confidence_fill_type = ""
    has_colored_row = False
    for row_index in range(2, sheet.max_row + 1):
        product_name = str(sheet.cell(row=row_index, column=9).value or "")
        quantity = str(sheet.cell(row=row_index, column=10).value or "")
        brand = str(sheet.cell(row=row_index, column=13).value or "")
        name = str(sheet.cell(row=row_index, column=4).value or "")
        remark = str(sheet.cell(row=row_index, column=18).value or "")
        fill_type = str(sheet.cell(row=row_index, column=9).fill.fill_type or "")
        if "HEPES" in product_name.upper():
            hep_es_rows.append((product_name, quantity))
        if "手套" in product_name:
            gift_rows.append(product_name)
        if "羟基酪醇" in product_name or "秦皮乙素" in product_name:
            yuanye_rows.append((product_name, brand))
        if "离心管" in product_name and ("白鲨" in product_name or brand == "白鲨"):
            baisha_rows.append((product_name, brand))
        if "透析袋" in product_name and ("杰乐普" in product_name or brand == "杰乐普"):
            jielepu_rows.append((product_name, brand))
        if "CY-09" in product_name or "SJ-MX0745" in str(sheet.cell(row=row_index, column=12).value or ""):
            spec = str(sheet.cell(row=row_index, column=12).value or "")
            sikejie_rows.append((product_name, brand, spec, name))
        if "15ml离心管" in product_name or "PBS缓冲液" in product_name:
            context_name_rows.append((product_name, name, remark))
        if "葡萄糖" in product_name:
            late_rows.append((product_name, name, remark))
        if "枪头" in product_name:
            mention_rows.append((product_name, name, remark))
        if "甲醛测试盒" in product_name:
            top_confidence_fill_type = fill_type
        if fill_type == "solid":
            has_colored_row = True
    assert_true(any(item[1] in {"2", "2.0"} for item in hep_es_rows), f"follow-up confirmation should补全数量: {hep_es_rows}")
    assert_true(bool(gift_rows), "gift item should remain in exported sheet")
    assert_true(len(yuanye_rows) >= 2, f"yuanye multi-product should split into 2 rows: {yuanye_rows}")
    assert_true(
        all(item[1] == "源叶" for item in yuanye_rows),
        f"yuanye child product rows should inherit the standalone brand, not promote product names to brands: {yuanye_rows}",
    )
    assert_true(any("白鲨" in item[0] and item[1] == "白鲨" for item in baisha_rows), f"baisha row should carry brand and keep brand in product name: {baisha_rows}")
    assert_true(any("杰乐普" in item[0] and item[1] == "杰乐普" for item in jielepu_rows), f"jielepu row should carry brand and keep brand in product name: {jielepu_rows}")
    assert_true(
        any(item[1] == "思科捷" and "SJ-MX0745" in item[2] and "10 mM" in item[2] and item[3] == "俞淑芳" for item in sikejie_rows),
        f"sikejie multiline product should keep brand, product/spec, and whitespace-split teacher name: {sikejie_rows}",
    )
    assert_true(
        all(item[1] == "周梦" and "姓名来自近3句上下文" in item[2] for item in context_name_rows),
        f"context fallback should infer 周梦 with remark note: {context_name_rows}",
    )
    assert_true(
        all(item[1] != "韩梅" for item in mention_rows),
        f"@ mention name should not be used as buyer evidence: {mention_rows}",
    )
    assert_true(
        all(not item[1] for item in late_rows),
        f"rows out of 5-minute window should not inherit context name: {late_rows}",
    )
    assert_true(has_colored_row, "sheet should contain confidence-colored rows")
    if top_confidence_fill_type:
        assert_true(top_confidence_fill_type in {"", "none"}, f"high-confidence rows should not be colored: {top_confidence_fill_type!r}")
    cost_price = sheet.cell(row=2, column=14).value
    total_cost = sheet.cell(row=2, column=16).value
    assert_true(cost_price in ("", None), "cost price should be empty in V1")
    assert_true(total_cost in ("", None), "total cost should be empty in V1")

    # Deterministic confidence-band color checks.
    svc = RecorderExportRunService(tenant_id=TEST_TENANT)
    assert_true(svc._confidence_fill_for_row({"confidence": 0.96}) is None, ">=0.95 should have no row color")
    assert_true(str((svc._confidence_fill_for_row({"confidence": 0.85}) or {}).fill_type) == "solid", "green band should be solid fill")
    assert_true(str((svc._confidence_fill_for_row({"confidence": 0.70}) or {}).fill_type) == "solid", "yellow band should be solid fill")
    assert_true(str((svc._confidence_fill_for_row({"confidence": 0.40}) or {}).fill_type) == "solid", "red band should be solid fill")


def validate_message_envelope_quote_timestamp_and_risk_flags() -> None:
    svc = RecorderExportRunService(tenant_id=TEST_TENANT)
    captured_at = "2026-06-04T15:10:11"
    rows, _meta = svc._extract_order_rows(
        {"module_key": "order_sheet_lab_v1", "module_version": "test"},
        [
            {
                "raw_message_id": "raw_env_quote_001",
                "id": "env-quote-001",
                "type": "text",
                "sender": "unknown",
                "sender_role": "unknown",
                "source_adapter": "win32_ocr",
                "content": "许聪\n[引用 张老师：旧订单 试剂盒 9盒 1元]\n枪头 2盒 30元",
                "time": "昨天03:02",
                "captured_at": captured_at,
                "ocr_confidence": 0.98,
                "conversation_type": "group",
                "target_name": "实验订货群",
            }
        ],
    )
    assert_true(rows, "clean current OCR body should produce an order row")
    row = rows[0]
    assert_equal(row.get("date"), "2026-06-04", "export date should use captured_at")
    assert_equal(row.get("time"), "15:10:11", "export time should use captured_at seconds")
    assert_true("枪头" in str(row.get("product_name") or ""), f"current body product should be extracted: {row}")
    assert_true("试剂盒" not in str(row.get("product_name") or ""), f"quoted product must not be extracted: {row}")
    assert_true("quote_contamination" in set(row.get("risk_flags") or []), f"quote risk should force review: {row}")
    assert_true(row.get("needs_review") is True, f"quote risk should force review: {row}")
    fill = svc._confidence_fill_for_row(row)
    assert_true(fill is not None and str(fill.fill_type) == "solid", f"quote risk should color the row: {fill}")


def validate_date_range_runs(client: TestClient, headers: dict[str, str]) -> None:
    day_run = create_run_with_payload(
        client,
        headers,
        {
            "target_names": ["企点售后群"],
            "limit": 10000,
            "date_from": "2025-03-03",
            "date_to": "2025-03-03",
            "quick_range": "day",
            "module_key": "order_sheet_lab_v1",
        },
    )
    execute_run(day_run)
    day_item = fetch_run_item(client, headers, day_run)
    day_dates = read_export_dates(Path(str((day_item.get("artifacts") or {}).get("xlsx_path") or "")))
    assert_true(day_dates == {"2025-03-03"}, f"day export should only contain target day: {day_dates}")

    week_run = create_run_with_payload(
        client,
        headers,
        {
            "target_names": ["企点售后群"],
            "limit": 10000,
            "start_time": "2025-03-03 00:00:00",
            "end_time": "2025-03-09 23:59:59",
            "quick_range": "week",
            "module_key": "order_sheet_lab_v1",
        },
    )
    execute_run(week_run)
    week_item = fetch_run_item(client, headers, week_run)
    week_dates = read_export_dates(Path(str((week_item.get("artifacts") or {}).get("xlsx_path") or "")))
    assert_true({"2025-03-03", "2025-03-04"}.issubset(week_dates), f"week export should include both seeded days: {week_dates}")


def validate_name_context_inference_rules() -> None:
    svc = RecorderExportRunService(tenant_id=TEST_TENANT)
    messages = [
        {"content": "王磊老师", "message_time": "2025-03-10 10:00:00"},
        {"content": "@记录员 李娜老师 请处理", "message_time": "2025-03-10 10:01:00"},
        {"content": "枪头 2盒 30元", "message_time": "2025-03-10 10:03:00"},
        {"content": "PBS 1瓶 20元", "message_time": "2025-03-10 10:07:30"},
    ]
    name, owner, ok = svc._infer_name_owner_from_recent_messages(
        ordered_messages=messages,
        message_index=3,
        max_messages=3,
        max_minutes=5,
    )
    assert_true(ok, "context inference should find candidate name")
    assert_equal(name, "王磊", "context inference should prefer teacher suffix name")
    assert_equal(owner, "王磊", "owner should fallback to name from context")
    no_name, no_owner, no_ok = svc._infer_name_owner_from_recent_messages(
        ordered_messages=messages,
        message_index=4,
        max_messages=3,
        max_minutes=5,
    )
    assert_true(not no_ok and not no_name and not no_owner, "context inference should fail when outside 5-minute window")
    base_row = svc._empty_order_row()
    base_row.update({"product_name": "枪头", "quantity": "2", "unit": "盒", "sale_price": "30", "confidence": 0.9, "needs_review": False})
    adjusted = svc._apply_name_context_fallback_to_row(
        base_row,
        ordered_messages=messages,
        message_index=3,
        max_messages=3,
        max_minutes=5,
        confidence_penalty=0.06,
    )
    assert_equal(str(adjusted.get("name") or ""), "王磊", "fallback row should fill name from context")
    assert_true(float(adjusted.get("confidence") or 0) < 0.9, "context fallback should reduce confidence")
    assert_true("姓名来自近3句上下文" in str(adjusted.get("remark") or ""), "fallback row should include traceable remark")


def validate_name_owner_price_prefix_cleanup() -> None:
    svc = RecorderExportRunService(tenant_id=TEST_TENANT)
    examples = [
        ("加厚喷砂手套一双78元程建军-张明老师", "程建军", "张明"),
        ("普通5ml离心管1包25元卢南-戚向阳老师", "卢南", "戚向阳"),
        ("alrabbitpAb订一个20u360元朱宁伟-胡升老师", "朱宁伟", "胡升"),
        ("细胞培养瓶1箱640元崔明辉-姚晓敏老师", "崔明辉", "姚晓敏"),
        ("思科捷\nCY-09（NLRP3抑制剂）  SJ-MX0745\n10 mM * 1mL in DMSO  订1个  700*0.95=665元\n俞\n淑芳老师", "俞淑芳", "俞淑芳"),
    ]
    for text, expected_name, expected_owner in examples:
        name, owner = svc._extract_name_owner(text)
        assert_equal(name, expected_name, f"name cleanup failed for {text}")
        assert_equal(owner, expected_owner, f"owner cleanup failed for {text}")
    row = svc._empty_order_row()
    row.update(
        {
            "name": "G73537F1桶105元卢南",
            "owner": "戚向阳",
            "product_name": "普通5ml离心管",
            "quantity": "1",
            "unit": "包",
            "sale_price": "25",
            "confidence": 0.9,
            "needs_review": False,
        }
    )
    finalized = svc._finalize_order_row(row, {"content": "普通5ml离心管1包25元卢南-戚向阳老师", "message_time": "2025-03-04 08:30:00"})
    assert_equal(str(finalized.get("name") or ""), "卢南", "finalize should sanitize dirty name")
    assert_equal(str(finalized.get("owner") or ""), "戚向阳", "finalize should preserve clean owner")
    assert_true("元" not in str(finalized.get("name") or ""), "finalized name should not contain price text")
    glove_row = svc._empty_order_row()
    glove_row.update({"name": "78元程建军", "owner": "程建军", "product_name": "加厚喷砂手套一双", "sale_price": "78", "confidence": 0.8})
    glove_finalized = svc._finalize_order_row(glove_row, {"content": "加厚喷砂手套一双78元程建军-张明老师", "message_time": "2025-03-04 08:31:00"})
    assert_equal(str(glove_finalized.get("name") or ""), "程建军", "dirty price prefix should be removed from glove buyer")
    assert_equal(str(glove_finalized.get("quantity") or ""), "1", "一双/default quantity should be captured")
    assert_true(not svc._should_drop_order_row(glove_finalized), "valid glove order should stay")
    book_row = svc._empty_order_row()
    book_row.update({"product_name": "书", "quantity": "1", "sale_price": "55.66", "confidence": 0.9})
    assert_true(svc._should_drop_order_row(book_row), "generic book row should be dropped from lab order export")
    noise_row = svc._empty_order_row()
    noise_row.update({"product_name": "索莱宝油红O染色试剂盒(细胞专用) 1322 1T1△600*070", "confidence": 0.55})
    assert_true(svc._should_drop_order_row(noise_row), "product-only noise row without person/quantity/price should be dropped")
    assert_equal(
        svc._extract_product_name("雾化机代付89/0.85=105元黄敏伟-姚晓敏老师"),
        "雾化机",
        "short instrument product names should survive product extraction",
    )
    implied_one_row = svc._empty_order_row()
    implied_one_row.update(
        {
            "name": "朱宁伟",
            "owner": "朱宁伟",
            "product_name": "酶科 大鼠白细胞介素 1β(IL-1β)ELISA科研试剂盒",
            "unit": "盒",
            "sale_price": "900",
            "source_scope_text": "酶科 MK1588B大鼠白细胞介素 1β(IL-1β)ELISA科研试剂盒48T900元",
            "confidence": 0.62,
        }
    )
    implied_one_finalized = svc._finalize_order_row(implied_one_row, {"content": "酶科 MK1588B大鼠白细胞介素 1β(IL-1β)ELISA科研试剂盒48T900元 朱宁伟老师", "message_time": "2025-03-04 08:32:00"})
    assert_equal(str(implied_one_finalized.get("quantity") or ""), "1", "priced row without explicit quantity should default to 1 for structured export")
    assert_equal(str(implied_one_finalized.get("unit") or ""), "盒", "existing unit should be preserved when defaulting quantity")
    dash_one_row = svc._empty_order_row()
    dash_one_row.update(
        {
            "name": "贾姝",
            "owner": "任宇杰",
            "product_name": "Polyclonal antibody 14088-1-AP",
            "quantity": "1",
            "unit": "个",
            "sale_price": "1215",
            "source_scope_text": "Polyclonal antibody 14088-1-AP 订—个50u 1350*0.9=1215元 贾姝老师(任宇杰订)",
            "confidence": 0.62,
        }
    )
    dash_one_finalized = svc._finalize_order_row(dash_one_row, {"content": "Polyclonal antibody 14088-1-AP 订—个50u 1350*0.9=1215元 贾姝老师(任宇杰订)", "message_time": "2025-03-04 08:33:00"})
    assert_equal(str(dash_one_finalized.get("quantity") or ""), "1", "OCR dash in 订—个 should not clear supported quantity")
    assert_equal(str(dash_one_finalized.get("unit") or ""), "个", "OCR dash quantity row should preserve unit")
    assert_equal(str(dash_one_finalized.get("time") or ""), "08:33:00", "finalize should write local HH:MM:SS time")
    utc_row = svc._finalize_order_row(dash_one_row, {"content": "Polyclonal antibody 14088-1-AP 订—个50u 1350*0.9=1215元 贾姝老师(任宇杰订)", "message_time": "2025-03-04T00:33:00Z"})
    assert_equal(str(utc_row.get("time") or ""), "08:33:00", "timezone-aware message time should be converted to UTC+8")
    missing_product_row = svc._empty_order_row()
    missing_product_row.update({"name": "俞淑芳", "quantity": "1", "unit": "支", "sale_price": "140", "confidence": 0.79})
    assert_true(svc._should_drop_order_row(missing_product_row), "rows without product name should not enter the main order sheet")
    formulation_noise_row = svc._empty_order_row()
    formulation_noise_row.update({"product_name": "01C/nVIAI-rC mM *1mL in DMSO", "brand": "None", "quantity": "1", "sale_price": "665", "confidence": 0.55})
    assert_true(svc._should_drop_order_row(formulation_noise_row), "OCR formulation fragments should not enter the main order sheet")


def validate_ocr_broken_line_order_rules() -> None:
    svc = RecorderExportRunService(tenant_id=TEST_TENANT)
    examples = [
        (
            "源叶羟基酪醇S25716-1g 384*0.9=345.6元 秦皮\n乙素S31424-1g100*0.9=90元赵伟睿老师",
            [
                ("源叶", "源叶羟基酪醇", "S25716-1g", "赵伟睿"),
                ("源叶", "源叶秦皮乙素", "S31424-1g", "赵伟睿"),
            ],
        ),
        (
            "思科捷 CY-09（NLRP3抑制剂）SJ-MX0745 10\nmM*1mL in DMSO订1个700*0.95=665元俞\n淑芳老师",
            [
                ("思科捷", "思科捷 CY-09（NLRP3抑制剂）", "SJ-MX0745 10 mM*1mL in DMSO", "俞淑芳"),
            ],
        ),
        (
            "思科捷Deoxycholic acid（去氧胆酸）SJ\nMN0579 10 mM * 1mL in DMSO 订1个 380*0.95=361元俞淑芳老师",
            [
                ("思科捷", "思科捷Deoxycholic acid（去氧胆酸）", "SJ-MN0579 10 mM * 1mL in DMSO", "俞淑芳"),
            ],
        ),
        (
            "【验收标记 LIVE_OCR_FIX_20260604_1743_B】\n思科捷Deoxycholic acid（去氧胆酸）SJ-\nMN0579 10 mM * 1mL in DMSO 订1个 380*0.95=361元俞淑芳老师",
            [
                ("思科捷", "思科捷Deoxycholic acid（去氧胆酸）", "SJ-MN0579 10 mM * 1mL in DMSO", "俞淑芳"),
            ],
        ),
    ]
    for content, expected_rows in examples:
        rows = svc._rule_extract_rows_from_message(
            {
                "raw_message_id": "ocr_broken_line_case",
                "content": content,
                "message_time": "2026-06-04T16:40:30",
                "captured_at": "2026-06-04T16:40:30",
                "source_adapter": "win32_ocr",
                "type": "text",
            }
        )
        compact = [
            (
                str(row.get("brand") or ""),
                str(row.get("product_name") or ""),
                str(row.get("spec") or ""),
                str(row.get("name") or ""),
            )
            for row in rows
        ]
        for expected in expected_rows:
            assert_true(expected in compact, f"OCR broken-line extraction should include {expected}: {compact}")
        joined = " ".join(" ".join(item) for item in compact)
        assert_true("LIVE_OCR_FIX" not in joined, f"validation marker should not pollute extracted rows: {compact}")


def validate_brand_single_use_rules() -> None:
    svc = RecorderExportRunService(tenant_id=TEST_TENANT)
    rows = [
        {"product_name": "羟基酪醇", "remark": "羟基酪醇 订1个 100元", "brand": ""},
        {"product_name": "秦皮乙素", "remark": "秦皮乙素 订1个 90元", "brand": ""},
        {"product_name": "50ml离心管", "remark": "50ml离心管 订1包 20元", "brand": ""},
    ]
    source_text = "\n".join(
        [
            "源叶",
            "羟基酪醇 订1个 100元",
            "秦皮乙素 订1个 90元",
            "白鲨 50ml离心管 订1包 20元",
        ]
    )
    out = svc._apply_brand_context_to_rows(rows, source_text=source_text)
    assert_equal(str(out[0].get("brand") or ""), "源叶", "first item should consume nearest upstream brand once")
    assert_equal(str(out[1].get("brand") or ""), "源叶", "strong-anchor rows should allow limited consecutive brand inheritance")
    assert_equal(str(out[2].get("brand") or ""), "白鲨", "current-line explicit brand should win")
    assert_true(str(out[0].get("product_name") or "").startswith("源叶 "), "brand should be displayed in product name")
    assert_true(str(out[1].get("product_name") or "").startswith("源叶 "), "inherited brand should be displayed in product name")
    assert_true(str(out[2].get("product_name") or "").startswith("白鲨 "), "brand should be displayed in product name")


def validate_close_duplicate_export_row_dedupe() -> None:
    svc = RecorderExportRunService(tenant_id=TEST_TENANT)
    base = {
        "date": "2026-06-04",
        "name": "周梦",
        "owner": "周梦",
        "receiver": "周梦",
        "record_type": "order_item",
        "brand": "白鲨",
        "product_name": "白鲨15ml离心管",
        "spec": "15ml",
        "quantity": "2",
        "unit": "包",
        "sale_price": "66",
        "total_sale": "132",
        "confidence": 0.9,
        "needs_review": False,
        "evidence_message_ids": ["first"],
    }
    rows = [
        {**base, "time": "17:10:04"},
        {**base, "time": "17:11:20", "confidence": 0.91, "evidence_message_ids": ["second"]},
        {**base, "time": "17:25:05", "evidence_message_ids": ["late"]},
        {**base, "time": "17:11:30", "spec": "50ml", "product_name": "白鲨50ml离心管", "sale_price": "60", "total_sale": "120"},
    ]
    out = svc._dedupe_close_export_rows(rows, max_minutes=10)
    assert_equal(len(out), 3, "close duplicate rows should be collapsed without merging later/different-SKU orders")
    merged = [
        row
        for row in out
        if str(row.get("product_name") or "") == "白鲨15ml离心管"
        and str(row.get("time") or "") in {"17:10:04", "17:11:20"}
    ]
    assert_equal(len(merged), 1, "near duplicate white shark 15ml row should keep one representative")
    evidence_ids = set(merged[0].get("evidence_message_ids") or [])
    assert_true({"first", "second"}.issubset(evidence_ids), "dedupe should retain both source message ids")
    assert_true(
        any(str(row.get("time") or "") == "17:25:05" for row in out),
        "same order outside the close OCR repeat window should remain",
    )
    assert_true(
        any(str(row.get("product_name") or "") == "白鲨50ml离心管" for row in out),
        "different SKU should not be merged",
    )


def fetch_run_item(client: TestClient, headers: dict[str, str], run_id: str) -> dict[str, Any]:
    response = client.get(f"/api/recorder/exports/runs/{run_id}", headers=headers)
    assert_equal(response.status_code, 200, "run detail status")
    item = response.json().get("item", {})
    assert_equal(item.get("status"), "succeeded", "run status should be succeeded")
    return item


def read_export_dates(workbook_path: Path) -> set[str]:
    assert_true(workbook_path.exists(), f"workbook should exist: {workbook_path}")
    workbook = load_workbook(workbook_path, data_only=True)
    sheet = workbook["Sheet1"]
    values: set[str] = set()
    for row_index in range(2, sheet.max_row + 1):
        value = str(sheet.cell(row=row_index, column=2).value or "").strip()
        if value:
            values.add(value)
    return values


def cleanup_runtime() -> None:
    root = tenant_runtime_root(TEST_TENANT)
    if root.exists():
        shutil.rmtree(root)


def assert_true(value: bool, message: str) -> None:
    if not value:
        raise AssertionError(message)


def assert_equal(actual: Any, expected: Any, message: str) -> None:
    if actual != expected:
        raise AssertionError(f"{message}: expected {expected!r}, got {actual!r}")


if __name__ == "__main__":
    raise SystemExit(main())
