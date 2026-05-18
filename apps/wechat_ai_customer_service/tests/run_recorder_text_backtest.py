"""Run text-only backtests for recorder structured export against raw chat data."""

from __future__ import annotations

import argparse
import json
import os
import shutil
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

os.environ.setdefault("WECHAT_STORAGE_BACKEND", "file")
os.environ.setdefault("WECHAT_CLOUD_REQUIRED", "0")
os.environ.setdefault("WECHAT_CLOUD_STRICT_ONLINE", "0")

from fastapi.testclient import TestClient
from openpyxl import load_workbook


APP_ROOT = Path(__file__).resolve().parents[1]
PROJECT_ROOT = APP_ROOT.parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from apps.wechat_ai_customer_service.admin_backend.app import create_app  # noqa: E402
from apps.wechat_ai_customer_service.admin_backend.services.background_handlers import handle_recorder_export_run_execute  # noqa: E402
from apps.wechat_ai_customer_service.admin_backend.services.recorder_export_run_service import RecorderExportRunService  # noqa: E402
from apps.wechat_ai_customer_service.admin_backend.services.work_queue import WorkQueueService  # noqa: E402
from apps.wechat_ai_customer_service.knowledge_paths import tenant_context, tenant_runtime_root  # noqa: E402


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Recorder text-only backtest runner")
    parser.add_argument("--tenant", default="test02_eval_20260511_v2")
    parser.add_argument("--raw-xlsx", required=True, help="Path to raw WeChat export xlsx")
    parser.add_argument("--target-name", default="企点售后群")
    parser.add_argument("--module-key", default="order_sheet_lab_v1")
    parser.add_argument("--months", nargs="+", default=["2026-01", "2026-02"])
    parser.add_argument("--limit", type=int, default=10000)
    parser.add_argument("--reset-runtime", action="store_true")
    parser.add_argument("--output-json", default="")
    return parser.parse_args()


def normalize_time(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    text = str(value or "").strip()
    if not text:
        return ""
    for pattern in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M"):
        try:
            parsed = datetime.strptime(text, pattern)
            return parsed.strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue
    return text


def normalize_content_type(value: Any) -> str:
    text = str(value or "").strip()
    if "引用" in text:
        return "quote"
    if "图片" in text or "图像" in text:
        return "image"
    if "系统" in text:
        return "system"
    return "text"


def load_raw_messages_from_xlsx(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    messages: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=1, values_only=True):
        if not row:
            continue
        raw_id = row[0] if len(row) > 0 else ""
        observed_at = normalize_time(row[1] if len(row) > 1 else "")
        sender = str(row[2] if len(row) > 2 else "").strip()
        content_type = normalize_content_type(row[3] if len(row) > 3 else "")
        content = str(row[4] if len(row) > 4 else "").strip()
        if not observed_at or not content:
            continue
        if content_type not in {"text", "quote"}:
            continue
        message_id = f"raw_excel_{str(raw_id or '').strip()}" if str(raw_id or "").strip() else f"raw_excel_auto_{len(messages) + 1}"
        messages.append(
            {
                "id": message_id,
                "type": content_type,
                "sender": sender or "未知发送人",
                "content": content,
                "time": observed_at,
            }
        )
    workbook.close()
    return messages


def seed_messages(client: TestClient, headers: dict[str, str], *, target_name: str, messages: list[dict[str, Any]]) -> None:
    batch_size = 300
    for start in range(0, len(messages), batch_size):
        batch = messages[start : start + batch_size]
        payload = {
            "conversation": {
                "target_name": target_name,
                "display_name": target_name,
                "conversation_type": "group",
                "selected_by_user": True,
                "learning_enabled": True,
                "notify_enabled": False,
                "status": "active",
            },
            "messages": batch,
            "source_module": "backtest_loader",
            "auto_learn": False,
        }
        response = client.post("/api/raw-messages/messages", headers=headers, json=payload)
        assert response.status_code == 200, response.text


def bind_module(client: TestClient, headers: dict[str, str], *, module_key: str) -> None:
    payload = {
        "binding_id": "backtest_local_admin_binding",
        "scope_type": "user",
        "scope_id": "local-admin",
        "user_id": "local-admin",
        "module_key": module_key,
        "enabled": True,
    }
    response = client.post("/api/recorder/module-bindings", headers=headers, json=payload)
    assert response.status_code == 200, response.text


def month_range(month_text: str) -> tuple[str, str]:
    start = datetime.strptime(month_text, "%Y-%m")
    if start.month == 12:
        end = datetime(start.year + 1, 1, 1) - timedelta(days=1)
    else:
        end = datetime(start.year, start.month + 1, 1) - timedelta(days=1)
    return start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")


def execute_run(run_id: str, *, tenant_id: str) -> None:
    queue = WorkQueueService(tenant_id=tenant_id)
    target_job = None
    for _ in range(30):
        claimed = queue.claim(queue="recorder_exports", worker_id="text-backtest", limit=80, lock_seconds=120)
        target_job = next((job for job in claimed if str((job.get("payload") or {}).get("run_id") or "") == run_id), None)
        if target_job is not None:
            break
    assert target_job is not None, f"run job not found in queue: {run_id}"
    payload = target_job.get("payload") if isinstance(target_job.get("payload"), dict) else {}
    result = handle_recorder_export_run_execute(payload)
    if result.get("ok"):
        queue.complete(str(target_job.get("job_id") or ""), result=result)
        return
    queue.fail(str(target_job.get("job_id") or ""), error=str(result.get("error") or "run failed"), retry=False)
    raise RuntimeError(f"export failed: {result}")


def clip_text(text: str, max_len: int = 120) -> str:
    normalized = " ".join(str(text or "").split())
    if len(normalized) <= max_len:
        return normalized
    return normalized[: max_len - 1] + "…"


def evaluate_run(*, tenant_id: str, run_id: str) -> dict[str, Any]:
    service = RecorderExportRunService(tenant_id=tenant_id)
    run_item = service.get_run(run_id)
    assert run_item is not None, f"run not found: {run_id}"
    assert str(run_item.get("status") or "") == "succeeded", f"run failed: {run_item.get('status')}"

    messages = service._load_messages_for_run(run_item)  # noqa: SLF001
    extracted_rows, extraction_meta = service._extract_order_rows(run_item, messages, run_id="")  # noqa: SLF001
    candidates = []
    for msg in messages:
        content = str(msg.get("content") or "").strip()
        if not content:
            continue
        if service._looks_like_order_message(content) or service._looks_like_llm_order_candidate(content):  # noqa: SLF001
            candidates.append(msg)

    candidate_ids = {str(item.get("raw_message_id") or "") for item in candidates if str(item.get("raw_message_id") or "")}
    covered_ids: set[str] = set()
    for row in extracted_rows:
        for mid in row.get("evidence_message_ids", []):
            mid_text = str(mid or "").strip()
            if mid_text:
                covered_ids.add(mid_text)

    missed = [item for item in candidates if str(item.get("raw_message_id") or "") not in covered_ids]
    suspicious_rows = [
        row
        for row in extracted_rows
        if not str(row.get("product_name") or "").strip()
        or str(row.get("record_type") or "order_item") not in {"order_item", "gift_item"}
    ]
    needs_review_rows = [row for row in extracted_rows if bool(row.get("needs_review"))]

    return {
        "run_id": run_id,
        "stats": run_item.get("stats") if isinstance(run_item.get("stats"), dict) else {},
        "extraction_meta": extraction_meta,
        "message_count": len(messages),
        "candidate_message_count": len(candidates),
        "covered_candidate_count": len(candidate_ids & covered_ids),
        "coverage_rate": round((len(candidate_ids & covered_ids) / len(candidate_ids)) if candidate_ids else 1.0, 4),
        "export_row_count_recomputed": len(extracted_rows),
        "needs_review_count_recomputed": len(needs_review_rows),
        "suspicious_row_count": len(suspicious_rows),
        "missed_candidate_count": len(missed),
        "missed_candidate_samples": [
            {
                "raw_message_id": str(item.get("raw_message_id") or ""),
                "observed_at": str(item.get("observed_at") or ""),
                "sender": str(item.get("sender") or ""),
                "content": clip_text(str(item.get("content") or "")),
            }
            for item in missed[:20]
        ],
        "needs_review_samples": [
            {
                "date": str(row.get("date") or ""),
                "name": str(row.get("name") or ""),
                "product_name": str(row.get("product_name") or ""),
                "quantity": str(row.get("quantity") or ""),
                "remark": clip_text(str(row.get("remark") or "")),
                "record_type": str(row.get("record_type") or "order_item"),
            }
            for row in needs_review_rows[:20]
        ],
    }


def run_month_backtest(
    client: TestClient,
    headers: dict[str, str],
    *,
    tenant_id: str,
    target_name: str,
    module_key: str,
    limit: int,
    month_text: str,
) -> dict[str, Any]:
    date_from, date_to = month_range(month_text)
    create_payload = {
        "target_names": [target_name],
        "limit": limit,
        "date_from": date_from,
        "date_to": date_to,
        "module_key": module_key,
    }
    response = client.post("/api/recorder/exports/runs", headers=headers, json=create_payload)
    assert response.status_code == 200, response.text
    run_item = response.json().get("item", {})
    run_id = str(run_item.get("run_id") or "")
    assert run_id.startswith("recorder_export_run_"), f"invalid run id: {run_id}"
    execute_run(run_id, tenant_id=tenant_id)
    metrics = evaluate_run(tenant_id=tenant_id, run_id=run_id)
    metrics["month"] = month_text
    metrics["date_from"] = date_from
    metrics["date_to"] = date_to
    artifacts = (RecorderExportRunService(tenant_id=tenant_id).get_run(run_id) or {}).get("artifacts") or {}
    metrics["artifacts"] = {
        "xlsx_path": str(artifacts.get("xlsx_path") or ""),
        "report_path": str(artifacts.get("report_path") or ""),
    }
    return metrics


def cleanup_runtime(tenant_id: str) -> None:
    root = tenant_runtime_root(tenant_id)
    if root.exists():
        shutil.rmtree(root)


def main() -> int:
    args = parse_args()
    raw_path = Path(args.raw_xlsx).expanduser().resolve()
    if not raw_path.exists():
        raise FileNotFoundError(f"raw xlsx not found: {raw_path}")
    raw_messages = load_raw_messages_from_xlsx(raw_path)
    if not raw_messages:
        raise RuntimeError("no valid text/quote messages loaded from raw xlsx")

    with tenant_context(args.tenant):
        if args.reset_runtime:
            cleanup_runtime(args.tenant)
        client = TestClient(create_app())
        headers = {"X-Tenant-ID": args.tenant}
        seed_messages(client, headers, target_name=args.target_name, messages=raw_messages)
        bind_module(client, headers, module_key=args.module_key)

        month_reports: list[dict[str, Any]] = []
        for month_text in args.months:
            month_reports.append(
                run_month_backtest(
                    client,
                    headers,
                    tenant_id=args.tenant,
                    target_name=args.target_name,
                    module_key=args.module_key,
                    limit=max(1, int(args.limit or 10000)),
                    month_text=month_text,
                )
            )

    summary = {
        "ok": True,
        "tenant": args.tenant,
        "module_key": args.module_key,
        "raw_xlsx": str(raw_path),
        "raw_message_count": len(raw_messages),
        "months": month_reports,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
    }
    output_path = Path(args.output_json).resolve() if str(args.output_json or "").strip() else None
    if output_path:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
