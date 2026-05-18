"""Raw upload persistence."""

from __future__ import annotations

import hashlib
import json
import re
import sys
from io import BytesIO
from datetime import datetime
from pathlib import Path
from typing import Any

from .audit_log import append_audit
from apps.wechat_ai_customer_service.knowledge_paths import active_tenant_id, tenant_admin_upload_index_path, tenant_raw_inbox_root
from apps.wechat_ai_customer_service.storage import get_postgres_store, load_storage_config


APP_ROOT = Path(__file__).resolve().parents[2]
WORKFLOWS_ROOT = APP_ROOT / "workflows"
if str(WORKFLOWS_ROOT) not in sys.path:
    sys.path.insert(0, str(WORKFLOWS_ROOT))

from rag_layer import RagService  # noqa: E402
ALLOWED_KINDS = {"products", "chats", "policies", "erp_exports"}
ALLOWED_KIND_REQUESTS = set(ALLOWED_KINDS) | {"auto"}
ALLOWED_SUFFIXES = {".txt", ".md", ".json", ".jsonl", ".csv", ".xlsx"}
SPREADSHEET_SUFFIXES = {".xlsx"}


class UploadStore:
    def __init__(self, *, tenant_id: str | None = None) -> None:
        self.tenant_id = active_tenant_id(tenant_id)
        self.raw_inbox_root = tenant_raw_inbox_root(self.tenant_id)
        self.index_path = tenant_admin_upload_index_path(self.tenant_id)

    def save_upload(self, filename: str, content: bytes, kind: str) -> dict[str, Any]:
        requested_kind = str(kind or "").strip() or "auto"
        if requested_kind not in ALLOWED_KIND_REQUESTS:
            return {"ok": False, "message": f"unsupported kind: {requested_kind}"}
        resolved_kind, detect_reason = resolve_upload_kind(requested_kind, filename, content)
        if resolved_kind not in ALLOWED_KINDS:
            return {"ok": False, "message": f"unsupported resolved kind: {resolved_kind}"}
        suffix = Path(filename).suffix.lower()
        if suffix not in ALLOWED_SUFFIXES:
            return {"ok": False, "message": f"unsupported suffix: {suffix}"}
        if not content.strip():
            return {"ok": False, "message": "empty upload"}

        stored_content = content
        stored_suffix = suffix
        normalized_from = ""
        if suffix in SPREADSHEET_SUFFIXES:
            converted = spreadsheet_to_text(content)
            if not converted.strip():
                return {"ok": False, "message": "spreadsheet has no readable cells"}
            stored_content = converted.encode("utf-8")
            stored_suffix = ".txt"
            normalized_from = suffix

        digest = hashlib.sha256(content).hexdigest()
        upload_id = "upload_" + digest[:16]
        safe_name = re.sub(r"[^A-Za-z0-9_.\-\u4e00-\u9fff]", "_", Path(filename).name)
        if stored_suffix != suffix:
            safe_name = safe_name + stored_suffix
        target_dir = self.raw_inbox_root / resolved_kind
        target_dir.mkdir(parents=True, exist_ok=True)
        target_path = target_dir / f"{upload_id}_{safe_name}"
        target_path.write_bytes(stored_content)
        record = {
            "upload_id": upload_id,
            "filename": filename,
            "kind": resolved_kind,
            "requested_kind": requested_kind,
            "kind_detect_reason": detect_reason,
            "path": str(target_path),
            "original_suffix": suffix,
            "stored_suffix": stored_suffix,
            "normalized_from": normalized_from,
            "sha256": digest,
            "size": len(stored_content),
            "uploaded_at": datetime.now().isoformat(timespec="seconds"),
            "learned": False,
        }
        records = [item for item in self.list_uploads() if item.get("upload_id") != upload_id]
        records.append(record)
        db = postgres_store()
        config = load_storage_config()
        if db:
            db.upsert_upload(active_tenant_id(self.tenant_id), record)
            if not config.mirror_files:
                append_audit("upload_created", {"upload_id": upload_id, "kind": resolved_kind, "requested_kind": requested_kind, "path": str(target_path)})
                return {"ok": True, "item": record}
        self.write_index(records)
        append_audit("upload_created", {"upload_id": upload_id, "kind": resolved_kind, "requested_kind": requested_kind, "path": str(target_path)})
        return {"ok": True, "item": record}

    def list_uploads(self) -> list[dict[str, Any]]:
        db = postgres_store()
        if db:
            records = db.list_uploads(active_tenant_id(self.tenant_id))
            if records:
                return records
        if not self.index_path.exists():
            return []
        return json.loads(self.index_path.read_text(encoding="utf-8"))

    def get_upload(self, upload_id: str) -> dict[str, Any] | None:
        for item in self.list_uploads():
            if item.get("upload_id") == upload_id:
                return item
        return None

    def delete_upload(self, upload_id: str) -> dict[str, Any]:
        records = self.list_uploads()
        target = next((item for item in records if item.get("upload_id") == upload_id), None)
        if not target:
            return {"ok": False, "message": f"upload not found: {upload_id}"}
        remaining = [item for item in records if item.get("upload_id") != upload_id]
        deleted_file = False
        file_path = Path(str(target.get("path") or ""))
        try:
            resolved_file = file_path.resolve()
            raw_root = self.raw_inbox_root.resolve()
            if raw_root in resolved_file.parents and resolved_file.exists():
                RagService(tenant_id=self.tenant_id).delete_source_by_path(file_path)
                resolved_file.unlink()
                deleted_file = True
        except OSError:
            deleted_file = False
        db = postgres_store()
        config = load_storage_config()
        if db:
            db.delete_upload(active_tenant_id(self.tenant_id), upload_id)
            if not config.mirror_files:
                append_audit("upload_deleted", {"upload_id": upload_id, "path": str(file_path), "deleted_file": deleted_file})
                return {"ok": True, "item": target, "deleted_file": deleted_file}
        self.write_index(remaining)
        append_audit("upload_deleted", {"upload_id": upload_id, "path": str(file_path), "deleted_file": deleted_file})
        return {"ok": True, "item": target, "deleted_file": deleted_file}

    def mark_learned(self, upload_ids: list[str], candidate_ids: list[str]) -> None:
        records = self.list_uploads()
        for item in records:
            if item.get("upload_id") in upload_ids:
                item["learned"] = True
                item["candidate_ids"] = sorted(set([*item.get("candidate_ids", []), *candidate_ids]))
                item["learned_at"] = datetime.now().isoformat(timespec="seconds")
                db = postgres_store()
                if db:
                    db.upsert_upload(active_tenant_id(self.tenant_id), item)
        if postgres_store() and not load_storage_config().mirror_files:
            return
        self.write_index(records)

    def write_index(self, records: list[dict[str, Any]]) -> None:
        self.index_path.parent.mkdir(parents=True, exist_ok=True)
        self.index_path.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")


def spreadsheet_to_text(content: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("openpyxl is required for xlsx uploads") from exc
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    lines: list[str] = []
    for sheet in workbook.worksheets:
        lines.append(f"# Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            values = [cell_to_text(value) for value in row]
            if any(values):
                lines.append(",".join(values))
    return "\n".join(lines).strip() + "\n"


def cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def resolve_upload_kind(requested_kind: str, filename: str, content: bytes) -> tuple[str, str]:
    if requested_kind in ALLOWED_KINDS:
        return requested_kind, "手动指定"

    name = str(filename or "").lower()
    preview = content_preview(content).lower()
    if any(token in name for token in ("chat_templates", "chat_template", "knowledge_chats", "dialogue", "dialogues", "话术模板", "客服话术")):
        return "chats", "文件名命中聊天模板关键词"
    score = {
        "products": 0,
        "chats": 0,
        "policies": 0,
        "erp_exports": 0,
    }

    for token in ("聊天", "会话", "对话", "客服", "群聊", "私聊", "chat", "conversation", "message", "wechat"):
        if token in name or token in preview:
            score["chats"] += 2
    for token in ("政策", "规则", "条款", "售后", "发票", "付款", "合同", "承诺", "赔付", "policy", "rule"):
        if token in name or token in preview:
            score["policies"] += 2
    for token in ("erp", "导出", "报表", "库存表", "订单", "sku", "台账", "excel", "sheet", "csv"):
        if token in name or token in preview:
            score["erp_exports"] += 2
    for token in ("商品", "车源", "车型", "库存", "报价", "价格", "配置", "product", "catalog", "inventory"):
        if token in name or token in preview:
            score["products"] += 2

    if any(token in preview for token in ("客户:", "用户:", "买家:", "客服:", "销售:", "顾问:")):
        score["chats"] += 3
    if any(token in preview for token in ("customer_message", "service_reply", "intent_tags", "tone_tags", "template_id", "scenario")):
        score["chats"] += 4
    if any(token in preview for token in ("商品名称", "车型", "VIN", "指导价", "库存", "上牌")):
        score["products"] += 3
    if any(token in preview for token in ("退款", "质保", "合同", "违约", "付款方式", "开票")):
        score["policies"] += 3

    suffix = Path(filename).suffix.lower()
    if suffix in {".csv", ".xlsx"}:
        score["erp_exports"] += 1

    ordered = sorted(score.items(), key=lambda item: (-item[1], item[0]))
    top_kind, top_score = ordered[0]
    if top_score <= 0:
        return "chats", "未命中关键词，默认按聊天记录"
    return top_kind, f"关键词命中分数 {top_score}"


def content_preview(content: bytes, *, limit: int = 6000) -> str:
    if not content:
        return ""
    try:
        return content[:limit].decode("utf-8", errors="ignore")
    except Exception:
        return ""


def postgres_store():
    config = load_storage_config()
    if not config.use_postgres or not config.postgres_configured:
        return None
    store = get_postgres_store(config=config)
    return store if store.available() else None
