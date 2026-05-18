"""Export tracked raw WeChat messages into chat-log style Excel workbooks."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from .raw_message_store import RawMessageStore
from apps.wechat_ai_customer_service.knowledge_paths import tenant_runtime_root


HEADER_FILL = PatternFill("solid", fgColor="E8EEF3")
MESSAGE_TYPE_LABELS = {
    "text": "文本消息",
    "quote": "引用消息",
    "image": "图片消息",
    "file": "文件消息",
    "voice": "语音消息",
    "video": "视频消息",
    "system": "系统消息",
}


class RawChatExportService:
    def __init__(self, *, tenant_id: str | None = None) -> None:
        self.store = RawMessageStore(tenant_id=tenant_id)
        self.export_root = tenant_runtime_root(self.store.tenant_id) / "exports" / "raw_chat_exports"

    def build_export(self, *, mode: str = "session") -> dict[str, Any]:
        normalized_mode = "time" if str(mode or "").strip().lower() == "time" else "session"
        messages = self._list_all_messages()
        conversation_map = self._conversation_map()

        self.export_root.mkdir(parents=True, exist_ok=True)
        filename = f"wechat_chat_records_{normalized_mode}_{self.store.tenant_id}_{timestamp_id()}.xlsx"
        path = self.export_root / filename

        workbook = Workbook()
        if normalized_mode == "time":
            self._build_time_workbook(workbook, messages=messages, conversation_map=conversation_map)
        else:
            self._build_session_workbook(workbook, messages=messages, conversation_map=conversation_map)
        workbook.save(path)

        conversation_ids = {str(item.get("conversation_id") or "") for item in messages if str(item.get("conversation_id") or "")}
        return {
            "ok": True,
            "mode": normalized_mode,
            "tenant_id": self.store.tenant_id,
            "path": str(path),
            "filename": filename,
            "message_count": len(messages),
            "conversation_count": len(conversation_ids),
        }

    def _build_session_workbook(
        self,
        workbook: Workbook,
        *,
        messages: list[dict[str, Any]],
        conversation_map: dict[str, dict[str, Any]],
    ) -> None:
        grouped: dict[tuple[str, str, str], list[dict[str, Any]]] = {}
        for item in messages:
            conversation_id = str(item.get("conversation_id") or "")
            conversation = conversation_map.get(conversation_id, {})
            name = conversation_name(item, conversation)
            conversation_type = str(
                conversation.get("conversation_type")
                or item.get("conversation_type")
                or "unknown"
            ).strip()
            key = (conversation_id, name, conversation_type)
            grouped.setdefault(key, []).append(item)

        default = workbook.active
        workbook.remove(default)

        used_titles: set[str] = set()
        if not grouped:
            sheet = workbook.create_sheet("聊天记录")
            self._write_sheet_metadata(
                sheet,
                session_id="",
                session_name="暂无追踪记录",
                conversation_type="unknown",
            )
            write_message_header(sheet, include_session_column=False)
            autosize_columns(sheet, 5)
            return

        sorted_groups = sorted(
            grouped.items(),
            key=lambda item: (
                normalize_sort_time(first_message_time(item[1])),
                item[0][1],
            ),
        )
        for (conversation_id, name, conversation_type), rows in sorted_groups:
            title = unique_sheet_title(name or conversation_id or "聊天记录", used_titles)
            used_titles.add(title)
            sheet = workbook.create_sheet(title)
            self._write_sheet_metadata(
                sheet,
                session_id=conversation_id,
                session_name=name,
                conversation_type=conversation_type,
            )
            write_message_header(sheet, include_session_column=False)
            for index, message in enumerate(sorted(rows, key=message_sort_key), start=1):
                sheet.cell(row=4 + index, column=1, value=index)
                sheet.cell(row=4 + index, column=2, value=display_message_time(message))
                sheet.cell(row=4 + index, column=3, value=display_sender(message))
                sheet.cell(row=4 + index, column=4, value=display_message_type(message))
                sheet.cell(row=4 + index, column=5, value=str(message.get("content") or ""))
            autosize_columns(sheet, 5)

    def _build_time_workbook(
        self,
        workbook: Workbook,
        *,
        messages: list[dict[str, Any]],
        conversation_map: dict[str, dict[str, Any]],
    ) -> None:
        sheet = workbook.active
        sheet.title = "聊天记录"
        conversation_names = {
            conversation_name(item, conversation_map.get(str(item.get("conversation_id") or ""), {}))
            for item in messages
            if conversation_name(item, conversation_map.get(str(item.get("conversation_id") or ""), {}))
        }
        self._write_sheet_metadata(
            sheet,
            session_id="ALL",
            session_name=f"全部会话（{len(conversation_names)}）",
            conversation_type="mixed",
        )
        write_message_header(sheet, include_session_column=True)
        for index, message in enumerate(sorted(messages, key=message_sort_key), start=1):
            conversation = conversation_map.get(str(message.get("conversation_id") or ""), {})
            sheet.cell(row=4 + index, column=1, value=index)
            sheet.cell(row=4 + index, column=2, value=display_message_time(message))
            sheet.cell(row=4 + index, column=3, value=conversation_name(message, conversation))
            sheet.cell(row=4 + index, column=4, value=display_sender(message))
            sheet.cell(row=4 + index, column=5, value=display_message_type(message))
            sheet.cell(row=4 + index, column=6, value=str(message.get("content") or ""))
        autosize_columns(sheet, 6)

    def _write_sheet_metadata(
        self,
        sheet: Worksheet,
        *,
        session_id: str,
        session_name: str,
        conversation_type: str,
    ) -> None:
        sheet.cell(row=1, column=1, value="会话信息")
        sheet.cell(row=2, column=1, value="微信ID")
        sheet.cell(row=2, column=2, value=session_id)
        sheet.cell(row=2, column=4, value="昵称")
        sheet.cell(row=2, column=5, value=session_name)
        sheet.cell(row=2, column=6, value="会话类型")
        sheet.cell(row=2, column=7, value=conversation_type_label(conversation_type))
        sheet.cell(row=3, column=1, value="导出工具")
        sheet.cell(row=3, column=2, value="OmniAuto")
        sheet.cell(row=3, column=3, value="导出版本")
        sheet.cell(row=3, column=4, value="1.0.0")
        sheet.cell(row=3, column=5, value="平台")
        sheet.cell(row=3, column=6, value="wechat")
        sheet.cell(row=3, column=7, value="导出时间")
        sheet.cell(row=3, column=8, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    def _list_all_messages(self) -> list[dict[str, Any]]:
        all_items: list[dict[str, Any]] = []
        offset = 0
        page_size = 1000
        while True:
            items = self.store.list_messages_advanced(limit=page_size, offset=offset)
            if not items:
                break
            all_items.extend(items)
            if len(items) < page_size:
                break
            offset += len(items)
            if offset >= 200000:
                break
        return all_items

    def _conversation_map(self) -> dict[str, dict[str, Any]]:
        items = self.store.list_conversations(status="all", limit=500)
        return {
            str(item.get("conversation_id") or ""): item
            for item in items
            if isinstance(item, dict) and str(item.get("conversation_id") or "")
        }


def write_message_header(sheet: Worksheet, *, include_session_column: bool) -> None:
    headers = ["序号", "时间"]
    if include_session_column:
        headers.append("会话")
    headers.extend(["发送者身份", "消息类型", "内容"])
    for column, title in enumerate(headers, start=1):
        cell = sheet.cell(row=4, column=column, value=title)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL


def autosize_columns(sheet: Worksheet, column_count: int) -> None:
    widths = {
        1: 9,
        2: 21,
        3: 24,
        4: 18,
        5: 14,
        6: 88,
    }
    for column in range(1, column_count + 1):
        sheet.column_dimensions[column_letter(column)].width = widths.get(column, 24)


def column_letter(index: int) -> str:
    current = max(1, index)
    result = ""
    while current > 0:
        current, remainder = divmod(current - 1, 26)
        result = chr(65 + remainder) + result
    return result


def message_sort_key(message: dict[str, Any]) -> tuple[str, str]:
    primary = normalize_sort_time(display_message_time(message))
    secondary = str(message.get("raw_message_id") or message.get("message_id") or "")
    return (primary, secondary)


def first_message_time(messages: list[dict[str, Any]]) -> str:
    if not messages:
        return ""
    ordered = sorted(messages, key=message_sort_key)
    return display_message_time(ordered[0])


def display_message_time(message: dict[str, Any]) -> str:
    raw = str(message.get("message_time") or message.get("observed_at") or "").strip()
    if not raw:
        return ""
    value = raw.replace("T", " ")
    if value.endswith("Z"):
        value = value[:-1]
    if "+" in value and len(value) >= 20:
        value = value.split("+", 1)[0]
    if "." in value:
        value = value.split(".", 1)[0]
    return value.strip()


def normalize_sort_time(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return "0000-00-00 00:00:00"
    if len(text) >= 19:
        return text[:19]
    return text


def display_sender(message: dict[str, Any]) -> str:
    sender = str(message.get("sender") or "").strip()
    role = str(message.get("sender_role") or "").strip().lower()
    if sender.lower() == "self" or role == "self":
        return "我"
    if sender:
        return sender
    if role == "system":
        return "系统"
    if role in {"assistant", "ai", "bot"}:
        return "AI"
    return "未知"


def display_message_type(message: dict[str, Any]) -> str:
    key = str(message.get("content_type") or "text").strip().lower()
    return MESSAGE_TYPE_LABELS.get(key, "文本消息")


def conversation_name(message: dict[str, Any], conversation: dict[str, Any]) -> str:
    return (
        str(conversation.get("display_name") or "").strip()
        or str(conversation.get("target_name") or "").strip()
        or str(message.get("target_name") or "").strip()
        or str(message.get("group_name") or "").strip()
        or str(message.get("conversation_id") or "").strip()
    )


def conversation_type_label(value: str) -> str:
    key = str(value or "").strip().lower()
    if key == "group":
        return "群聊"
    if key == "private":
        return "私聊"
    if key == "file_transfer":
        return "文件传输助手"
    if key == "mixed":
        return "全部会话"
    return "未知"


def unique_sheet_title(name: str, used: set[str]) -> str:
    cleaned = sanitize_sheet_title(name or "聊天记录")
    if cleaned not in used:
        return cleaned
    counter = 2
    while True:
        suffix = f"_{counter}"
        base = cleaned[: max(1, 31 - len(suffix))]
        candidate = f"{base}{suffix}"
        if candidate not in used:
            return candidate
        counter += 1


def sanitize_sheet_title(name: str) -> str:
    text = str(name or "").strip()
    for char in ("\\", "/", "*", "[", "]", ":", "?"):
        text = text.replace(char, "_")
    text = text or "聊天记录"
    return text[:31]


def timestamp_id() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")
