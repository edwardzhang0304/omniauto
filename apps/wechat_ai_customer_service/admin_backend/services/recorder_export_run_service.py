"""Async recorder export run service with module-based execution."""

from __future__ import annotations

import hashlib
import csv
import json
import os
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from apps.wechat_ai_customer_service.admin_backend.services.raw_message_store import RawMessageStore
from apps.wechat_ai_customer_service.admin_backend.services.recorder_module_registry import RecorderModuleRegistryService
from apps.wechat_ai_customer_service.admin_backend.services.work_queue import WorkQueueService
from apps.wechat_ai_customer_service.knowledge_paths import active_tenant_id, tenant_runtime_root
from apps.wechat_ai_customer_service.wechat_message_envelope import OCR_RPA_ADAPTERS, recorder_view_from_message
from apps.wechat_ai_customer_service.workflows.generate_review_candidates import call_deepseek_json


MAX_RUN_RECORDS = 2000
LLM_CACHE_VERSION = "recorder_export_llm_cache_v1"
RUN_STAGE_LABELS = {
    "queued": "排队中",
    "preprocessing": "预处理中",
    "scanning": "筛选候选消息",
    "extracting": "结构化抽取中",
    "llm_extracting": "LLM语义抽取中",
    "llm_branding": "品牌推断中",
    "finalizing": "整理导出行",
    "reviewing": "质量复核中",
    "exporting": "生成导出文件",
    "completed": "已完成",
    "failed": "失败",
}
ORDER_HEADERS = [
    "",
    "日期",
    "时间",
    "姓名",
    "责任人（老板）",
    "收货人",
    "校区",
    "订货单位",
    "货品名称",
    "数量",
    "单位",
    "规格",
    "品牌",
    "进价（单价）",
    "售价（单价）",
    "总进价",
    "总售价",
    "备注",
]
NAME_OWNER_RE = re.compile(r"(?P<name>[\u4e00-\u9fa5A-Za-z0-9]{1,24})[-—](?P<owner>[\u4e00-\u9fa5A-Za-z0-9]{1,24})老师")
ORDER_UNIT_PATTERN = r"盒|箱|袋|瓶|套|个|包|桶|台|件|支|根|片|株|卷|张|双|箱装|包/盒|公斤|斤|份|块|提|板|把|组|升"
QTY_UNIT_RE = re.compile(rf"(?P<qty>\d+(?:\.\d+)?)\s*(?P<unit>{ORDER_UNIT_PATTERN})")
QTY_CN_UNIT_RE = re.compile(rf"(?:各)?(?:订|买|代付)?\s*(?P<qty_cn>[一二两俩三四五六七八九十百半])\s*(?P<unit>{ORDER_UNIT_PATTERN})")
PRICE_RE = re.compile(r"(?P<price>\d+(?:\.\d+)?)\s*元")
DATE_RE = re.compile(r"(?P<year>\d{4})[-/](?P<month>\d{1,2})[-/](?P<day>\d{1,2})")
MONTH_DAY_RE = re.compile(r"(?P<month>\d{1,2})月(?P<day>\d{1,2})日")
YEAR_MONTH_DAY_CN_RE = re.compile(r"(?P<year>\d{4})年(?P<month>\d{1,2})月(?P<day>\d{1,2})日")
MONTH_DAY_SLASH_RE = re.compile(r"(?P<month>\d{1,2})[./-](?P<day>\d{1,2})")
MMDD_RE = re.compile(r"^(?P<month>0[1-9]|1[0-2])(?P<day>[0-2]\d|3[0-1])$")
ORDER_HINT_TERMS = ("订", "代付", "老师", "元", "收货", "发货", "下单")
ORDER_ACTION_TERMS = ("订", "代付", "下单", "买", "各订", "已订", "补订", "重订", "再订")
CANCEL_OR_STATUS_TERMS = ("不下单", "取消", "退掉", "不用了", "暂时不用", "已带走", "库存有现货", "按表格订货", "照表格订货")
SUMMARY_OR_PROMO_TERMS = ("合计", "共计", "满", "赠品")
NON_ORDER_PRODUCT_HINTS = {"代购", "抗体", "满", "共计", "合计", "文件", "备注"}
LLM_ORDER_INTENT_TERMS = ("订", "下单", "代付", "补订", "再订", "买", "发货", "货号", "规格", "元", "盒", "瓶", "套")
DEFAULT_INCLUDE_RECORD_TYPES = ("order_item", "gift_item")
FOLLOWUP_CONFIRM_RE = re.compile(r"^(?:是的|对|好的|嗯|行|确认|收到|ok|OK|嗯嗯)\s*[,，:：]?\s*(?P<body>.+)?$")
FOLLOWUP_SETTLEMENT_HINTS = ("合计", "共计", "满减", "优惠", "抹零", "返款", "抵扣", "欠款", "预存")
VALIDATION_MARKER_RE = re.compile(r"【\s*验收标记[^】]*】|验收标记\s*[:：]?\s*[A-Za-z0-9_.-]{6,}", re.IGNORECASE)
INVENTORY_STATUS_RE = re.compile(
    r"(?:^|[\s，,。；;])(?:现货|库存)(?:[^，,。；;\n\r]{0,24})?(?:差|缺)\s*\d+(?:\.\d+)?\s*(?:瓶|盒|箱|包|支|套|个|件|板|提|组)(?:[^，,。；;\n\r]{0,24})?(?:满|凑)\s*\d+\s*(?:箱|盒|包|件|套)",
    re.IGNORECASE,
)
SHORT_OFFICE_PRODUCT_HINT_RE = re.compile(r"(硒鼓|墨盒|碳粉|打印纸|a4纸|复印纸|标签纸|胶带|电池|文件夹|订书钉|打印机)", re.IGNORECASE)
BRAND_INHERIT_BLOCK_RE = re.compile(r"(娃哈哈|矿泉水|饮用水|会员|打印纸|复印纸|喷壶|办公|快递|运费)", re.IGNORECASE)
UTC8 = timezone(timedelta(hours=8))
PRODUCT_BRAND_CONTEXT_HINT_RE = re.compile(
    r"(离心管|透析袋|试剂|培养基|蛋白|抗体|细胞|缓冲液|滤膜|枪头|吸头|培养皿|孔板|货号|型号|md\d+|cat\.?\s*no\.?)",
    re.IGNORECASE,
)
NON_BRAND_STOPWORDS = {
    "PBS",
    "DMEM",
    "DMSO",
    "FBS",
    "MEM",
    "RPMI",
    "HEP",
    "HEPG2",
    "ANNEXIN",
    "CELL",
    "UN",
    "ML",
}
SPEC_SIZE_RE = re.compile(
    r"(?:\d+\s*[Tt]\s*/\s*\d+)|(?:\d+\s*[*xX×]\s*)?\d+(?:\.\d+)?\s*(?:μl|ul|ml|l|g|kg|mg|cm|mm|kda|da|bp|t)(?![A-Za-z0-9])",
    re.IGNORECASE,
)
SPEC_MARKER_RE = re.compile(
    r"(?:货号|型号|cat\s*no\.?|catno|no\.)\s*[:：]?\s*(?P<spec>[A-Za-z0-9][A-Za-z0-9._/\-*]{1,64}(?:\s+[A-Za-z0-9μµ._/\-*]{1,24})?)",
    re.IGNORECASE,
)
SPEC_TOKEN_RE = re.compile(r"[A-Za-z0-9][A-Za-z0-9._/\-*]{1,64}")
SPEC_CODE_TOKEN_RE = re.compile(
    r"(?:[A-Za-z]{1,10}[A-Za-z0-9]{0,8}[-_/][A-Za-z0-9][A-Za-z0-9._/\-*]{1,32})|(?:[A-Za-z]{1,8}\d{2,}[A-Za-z0-9._/\-*]{0,24})",
    re.IGNORECASE,
)
FORMULATION_SPEC_RE = re.compile(
    r"\d+(?:\.\d+)?\s*mM\s*[*xX×]\s*\d+(?:\.\d+)?\s*mL(?:\s*in\s*DMSO)?",
    re.IGNORECASE,
)
DEFAULT_LAB_BRAND_ALIASES = (
    "津腾",
    "康为",
    "源叶",
    "麦克林",
    "施睿康",
    "甄选",
    "赛宁",
    "建成",
    "索莱宝",
    "毕得医药",
    "白鲨",
    "杰乐普",
    "思科捷",
)
CONFIDENCE_COLOR_GREEN = PatternFill("solid", fgColor="E8F7EE")
CONFIDENCE_COLOR_YELLOW = PatternFill("solid", fgColor="FFF6DB")
CONFIDENCE_COLOR_RED = PatternFill("solid", fgColor="FDE8E8")
EXPORT_RED_RISK_FLAGS = {
    "bubble_boundary_ambiguous",
    "multi_bubble_possible_merge",
    "quote_contamination",
    "long_press_overlay_detected",
    "missing_product",
    "person_as_brand_candidate",
    "sender_name_in_content",
}
EXPORT_YELLOW_RISK_FLAGS = {
    "ocr_low_confidence",
    "speaker_prefix_split_from_ocr_text",
    "missing_quantity",
    "missing_price",
    "name_from_recent_context",
    "quantity_defaulted_to_one",
    "brand_from_weak_context",
    "spec_from_weak_context",
}


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def progress_between(start: float, end: float, current: int, total: int) -> float:
    total_safe = max(1, int(total or 0))
    ratio = max(0.0, min(float(current or 0) / float(total_safe), 1.0))
    return max(0.0, min(start + (end - start) * ratio, 0.99))


def stable_digest(value: str, length: int = 16) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()[:length]


def recorder_message_view(message: dict[str, Any]) -> dict[str, Any]:
    return recorder_view_from_message(message if isinstance(message, dict) else {})


def recorder_content_text(message: dict[str, Any]) -> str:
    view = recorder_message_view(message)
    return str(view.get("content_for_export") or message.get("content_clean") or message.get("content_body") or message.get("content") or "").strip()


def recorder_time_text(message: dict[str, Any]) -> str:
    view = recorder_message_view(message)
    source_adapter = str(message.get("source_adapter") or "").strip().lower()
    if source_adapter in OCR_RPA_ADAPTERS:
        return str(view.get("captured_at") or message.get("captured_at") or message.get("observed_at") or message.get("message_time") or "")
    return str(message.get("message_time") or message.get("time") or message.get("observed_at") or view.get("captured_at") or message.get("captured_at") or "")


def recorder_message_for_export(message: dict[str, Any]) -> dict[str, Any]:
    item = dict(message or {})
    view = recorder_message_view(item)
    content = str(view.get("content_for_export") or "").strip()
    captured_at = str(view.get("captured_at") or "").strip()
    if content:
        item["content"] = content
        item["content_body"] = content
        item["content_clean"] = content
    if captured_at:
        item["captured_at"] = captured_at
        source_adapter = str(item.get("source_adapter") or "").strip().lower()
        if source_adapter in OCR_RPA_ADAPTERS:
            item["message_time"] = captured_at
            item.setdefault("observed_at", captured_at)
        elif not str(item.get("message_time") or "").strip():
            item["message_time"] = str(item.get("time") or item.get("observed_at") or captured_at)
    item["risk_flags"] = list(view.get("risk_flags") or [])
    item["quality_flags"] = list(view.get("quality_flags") or item.get("quality_flags") or [])
    item["quoted_fragments"] = list(view.get("quoted_fragments") or item.get("quoted_fragments") or [])
    item["excluded_fragments"] = list(view.get("excluded_fragments") or item.get("excluded_fragments") or [])
    item["bubble_id"] = str(view.get("bubble_id") or item.get("bubble_id") or "")
    item["speaker_name"] = str(view.get("speaker_name") or item.get("speaker_name") or "")
    return item


class RecorderExportRunService:
    """Create, execute, track and download recorder export runs."""

    def __init__(self, *, tenant_id: str | None = None) -> None:
        self.tenant_id = active_tenant_id(tenant_id)
        self.raw_store = RawMessageStore(tenant_id=self.tenant_id)
        self.modules = RecorderModuleRegistryService()
        self.queue = WorkQueueService(tenant_id=self.tenant_id)
        self._active_brand_aliases: list[str] = list(DEFAULT_LAB_BRAND_ALIASES)

    @property
    def root(self) -> Path:
        return tenant_runtime_root(self.tenant_id) / "recorder" / "exports"

    @property
    def runs_path(self) -> Path:
        return self.root / "runs.json"

    @property
    def files_root(self) -> Path:
        return self.root / "files"

    @property
    def reports_root(self) -> Path:
        return self.root / "reports"

    @property
    def llm_cache_root(self) -> Path:
        return tenant_runtime_root(self.tenant_id) / "recorder" / "llm_cache" / LLM_CACHE_VERSION

    def _call_deepseek_json_cached(self, prompt: dict[str, Any], *, namespace: str) -> dict[str, Any]:
        cache_key = stable_digest(
            json.dumps(
                {
                    "version": LLM_CACHE_VERSION,
                    "namespace": str(namespace or "default"),
                    "prompt": prompt,
                },
                ensure_ascii=False,
                sort_keys=True,
            ),
            length=32,
        )
        cache_path = self.llm_cache_root / f"{cache_key}.json"
        if cache_path.exists():
            try:
                cached = json.loads(cache_path.read_text(encoding="utf-8"))
                payload = cached.get("payload") if isinstance(cached, dict) else None
                if isinstance(payload, dict):
                    return payload
            except (OSError, json.JSONDecodeError):
                pass
        try:
            payload = call_deepseek_json(prompt)
        except Exception:
            return {}
        if not isinstance(payload, dict):
            return {}
        if payload:
            try:
                self.llm_cache_root.mkdir(parents=True, exist_ok=True)
                temp_path = cache_path.with_suffix(".json.tmp")
                temp_path.write_text(
                    json.dumps(
                        {
                            "namespace": str(namespace or "default"),
                            "created_at": now_iso(),
                            "payload": payload,
                        },
                        ensure_ascii=False,
                    ),
                    encoding="utf-8",
                )
                os.replace(temp_path, cache_path)
            except OSError:
                pass
        return payload

    def list_runs(self, *, status: str = "all", limit: int = 100) -> list[dict[str, Any]]:
        records = [item for item in self._read_json(self.runs_path, default=[]) if isinstance(item, dict)]
        if status and status != "all":
            records = [item for item in records if str(item.get("status") or "") == status]
        records.sort(key=lambda item: str(item.get("created_at") or ""), reverse=True)
        return records[: max(1, min(int(limit or 100), 500))]

    def get_run(self, run_id: str) -> dict[str, Any] | None:
        for item in self.list_runs(status="all", limit=MAX_RUN_RECORDS):
            if str(item.get("run_id") or "") == run_id:
                return item
        return None

    def delete_run(self, run_id: str) -> dict[str, Any]:
        target_run_id = str(run_id or "").strip()
        if not target_run_id:
            raise ValueError("run_id is required")
        records = [item for item in self._read_json(self.runs_path, default=[]) if isinstance(item, dict)]
        target = next((item for item in records if str(item.get("run_id") or "") == target_run_id), None)
        if not target:
            return {"deleted": False, "run_id": target_run_id}

        cancelled_jobs = self._cancel_run_jobs(target_run_id)
        artifacts = target.get("artifacts") if isinstance(target.get("artifacts"), dict) else {}
        deleted_files = 0
        for key in ("xlsx_path", "csv_path", "report_path"):
            if self._safe_delete_artifact(str(artifacts.get(key) or "")):
                deleted_files += 1

        remaining = [item for item in records if str(item.get("run_id") or "") != target_run_id]
        self._write_json(self.runs_path, remaining)
        return {
            "deleted": True,
            "run_id": target_run_id,
            "deleted_files": deleted_files,
            "cancelled_jobs": len(cancelled_jobs),
        }

    def create_run(
        self,
        payload: dict[str, Any],
        *,
        requested_by_user_id: str,
        requested_by_username: str = "",
        requested_module_key: str = "",
    ) -> dict[str, Any]:
        module = self.modules.resolve_module(
            tenant_id=self.tenant_id,
            user_id=requested_by_user_id,
            requested_module_key=requested_module_key or str(payload.get("module_key") or ""),
        )
        filters = normalize_filters(payload)
        now_text = now_iso()
        run_id = "recorder_export_run_" + stable_digest(
            json.dumps(
                {
                    "tenant_id": self.tenant_id,
                    "module_key": module.get("module_key"),
                    "filters": filters,
                    "requested_by_user_id": requested_by_user_id,
                    "created_at": now_text,
                },
                ensure_ascii=False,
                sort_keys=True,
            ),
            20,
        )
        run = {
            "run_id": run_id,
            "tenant_id": self.tenant_id,
            "requested_by_user_id": requested_by_user_id,
            "requested_by_username": requested_by_username,
            "module_key": str(module.get("module_key") or ""),
            "module_version": str(module.get("version") or ""),
            "module_name": str(module.get("module_name") or module.get("module_key") or ""),
            "status": "queued",
            "stage": "queued",
            "queue_name": "recorder_exports",
            "filters": filters,
            "progress": {
                "stage": "queued",
                "stage_label": RUN_STAGE_LABELS["queued"],
                "processed_messages": 0,
                "total_messages": 0,
                "percent": 0.0,
                "updated_at": now_text,
            },
            "stats": {
                "input_message_count": 0,
                "export_row_count": 0,
                "needs_review_count": 0,
                "skipped_count": 0,
            },
            "artifacts": {
                "xlsx_path": "",
                "csv_path": "",
                "report_path": "",
            },
            "error": "",
            "created_at": now_text,
            "updated_at": now_text,
            "started_at": "",
            "finished_at": "",
        }
        self._upsert_run(run)
        self.queue.enqueue(
            kind="recorder_export_run_execute",
            payload={"tenant_id": self.tenant_id, "run_id": run_id},
            queue="recorder_exports",
            dedupe_key=f"recorder_export_run_execute:{run_id}",
            priority=5,
        )
        return run

    def process_run(self, run_id: str) -> dict[str, Any]:
        run = self.get_run(run_id)
        if not run:
            raise FileNotFoundError(run_id)
        if str(run.get("status") or "") == "succeeded":
            return {"ok": True, "already_done": True, "run": run}

        started_at = now_iso()
        running = {
            **run,
            "status": "running",
            "stage": "preprocessing",
            "started_at": started_at,
            "updated_at": started_at,
            "error": "",
            "progress": {
                "stage": "preprocessing",
                "stage_label": RUN_STAGE_LABELS["preprocessing"],
                "processed_messages": 0,
                "total_messages": 0,
                "percent": 0.0,
                "updated_at": started_at,
            },
        }
        self._upsert_run(running)
        try:
            self._update_run_progress(
                run_id,
                stage="preprocessing",
                processed=0,
                total=0,
                force_percent=0.02,
                stage_detail="读取筛选条件与原始消息",
            )
            messages = self._load_messages_for_run(running)
            self._update_run_progress(
                run_id,
                stage="scanning",
                processed=0,
                total=len(messages),
                force_percent=0.05,
                stage_detail="正在筛选候选订单消息",
                extra={"unit_label": "消息"},
            )
            workbook_result = self._build_workbook(running, messages, run_id=run_id)
            self._update_run_progress(run_id, stage="exporting", processed=len(messages), total=len(messages), force_percent=0.95)
            report_path = self._write_report(running, messages, workbook_result)
            completed = {
                **running,
                "status": "succeeded",
                "stage": "completed",
                "stats": {
                    "input_message_count": len(messages),
                    "export_row_count": int(workbook_result.get("export_row_count", 0) or 0),
                    "needs_review_count": int(workbook_result.get("needs_review_count", 0) or 0),
                    "skipped_count": int(workbook_result.get("skipped_count", 0) or 0),
                    "llm_calls": int(workbook_result.get("llm_calls", 0) or 0),
                    "llm_segment_calls": int(workbook_result.get("llm_segment_calls", 0) or 0),
                    "llm_repair_calls": int(workbook_result.get("llm_repair_calls", 0) or 0),
                },
                "artifacts": {
                    "xlsx_path": str(workbook_result.get("xlsx_path") or ""),
                    "csv_path": str(workbook_result.get("csv_path") or ""),
                    "report_path": str(report_path),
                },
                "progress": {
                    "stage": "completed",
                    "stage_label": RUN_STAGE_LABELS["completed"],
                    "processed_messages": len(messages),
                    "total_messages": len(messages),
                    "percent": 1.0,
                    "updated_at": now_iso(),
                },
                "updated_at": now_iso(),
                "finished_at": now_iso(),
            }
            self._upsert_run(completed)
            return {"ok": True, "run": completed}
        except Exception as exc:
            failed = {
                **running,
                "status": "failed",
                "stage": "failed",
                "error": repr(exc),
                "progress": {
                    "stage": "failed",
                    "stage_label": RUN_STAGE_LABELS["failed"],
                    "processed_messages": int(((running.get("progress") or {}) if isinstance(running.get("progress"), dict) else {}).get("processed_messages", 0) or 0),
                    "total_messages": int(((running.get("progress") or {}) if isinstance(running.get("progress"), dict) else {}).get("total_messages", 0) or 0),
                    "percent": float(((running.get("progress") or {}) if isinstance(running.get("progress"), dict) else {}).get("percent", 0.0) or 0.0),
                    "updated_at": now_iso(),
                },
                "updated_at": now_iso(),
                "finished_at": now_iso(),
            }
            self._upsert_run(failed)
            return {"ok": False, "error": repr(exc), "run": failed}

    def ensure_run_job(self, run_id: str) -> dict[str, Any]:
        run_id = str(run_id or "").strip()
        if not run_id:
            raise ValueError("run_id is required")
        run = self.get_run(run_id)
        if not run:
            raise FileNotFoundError(run_id)
        status = str(run.get("status") or "")
        if status not in {"queued", "running"}:
            return {"ok": True, "skipped": True, "reason": f"run_status_{status or 'unknown'}"}
        if status == "running" and self._run_has_recent_progress(run, stale_minutes=20):
            return {"ok": True, "skipped": True, "reason": "run_still_active_recent_progress"}
        dedupe_key = f"recorder_export_run_execute:{run_id}"
        existing = self.queue.find_active_dedupe(queue="recorder_exports", dedupe_key=dedupe_key)
        if existing:
            return {"ok": True, "existing_job": existing}
        job = self.queue.enqueue(
            kind="recorder_export_run_execute",
            payload={"tenant_id": self.tenant_id, "run_id": run_id},
            queue="recorder_exports",
            dedupe_key=dedupe_key,
            priority=5,
        )
        return {"ok": True, "job": job, "requeued": True}

    @staticmethod
    def _run_has_recent_progress(run: dict[str, Any], *, stale_minutes: int) -> bool:
        progress = run.get("progress") if isinstance(run.get("progress"), dict) else {}
        timestamp = str(progress.get("updated_at") or run.get("updated_at") or run.get("started_at") or "").strip()
        if not timestamp:
            return False
        try:
            updated_at = datetime.fromisoformat(timestamp.replace("Z", ""))
        except ValueError:
            return False
        age = datetime.now() - updated_at
        return age.total_seconds() < max(60, int(stale_minutes or 20) * 60)

    def _cancel_run_jobs(self, run_id: str) -> list[dict[str, Any]]:
        cancelled: list[dict[str, Any]] = []
        dedupe_key = f"recorder_export_run_execute:{run_id}"
        for job in self.queue.list_jobs(status="all", limit=500):
            if str(job.get("queue") or "") != "recorder_exports":
                continue
            if str(job.get("dedupe_key") or "") != dedupe_key:
                continue
            if str(job.get("status") or "") not in {"pending", "running"}:
                continue
            job_id = str(job.get("job_id") or "").strip()
            if not job_id:
                continue
            cancelled_job = self.queue.cancel(job_id, reason=f"recorder_export_run_deleted:{run_id}")
            if isinstance(cancelled_job, dict):
                cancelled.append(cancelled_job)
        return cancelled

    def _safe_delete_artifact(self, path_text: str) -> bool:
        path_text = str(path_text or "").strip()
        if not path_text:
            return False
        try:
            path = Path(path_text).resolve()
            files_root = self.files_root.resolve()
            reports_root = self.reports_root.resolve()
            inside_allowed_root = path.is_relative_to(files_root) or path.is_relative_to(reports_root)
            if not inside_allowed_root:
                return False
            if not path.exists() or not path.is_file():
                return False
            path.unlink()
            return True
        except (OSError, ValueError):
            return False

    def _update_run_progress(
        self,
        run_id: str,
        *,
        stage: str,
        processed: int | None = None,
        total: int | None = None,
        force_percent: float | None = None,
        stage_detail: str | None = None,
        extra: dict[str, Any] | None = None,
    ) -> None:
        run = self.get_run(run_id)
        if not run:
            return
        progress = run.get("progress") if isinstance(run.get("progress"), dict) else {}
        next_total = int(total if total is not None else progress.get("total_messages", 0) or 0)
        next_processed = int(processed if processed is not None else progress.get("processed_messages", 0) or 0)
        percent = float(progress.get("percent", 0.0) or 0.0)
        if force_percent is not None:
            percent = max(0.0, min(float(force_percent), 1.0))
        elif next_total > 0:
            percent = max(0.0, min(next_processed / float(next_total), 0.99))
        next_progress = {
            **progress,
            "stage": stage,
            "stage_label": RUN_STAGE_LABELS.get(stage, stage),
            "processed_messages": max(0, next_processed),
            "total_messages": max(0, next_total),
            "percent": percent,
            "updated_at": now_iso(),
        }
        if stage_detail is not None:
            next_progress["stage_detail"] = str(stage_detail)
        elif str(progress.get("stage") or "") != stage:
            next_progress["stage_detail"] = ""
        if extra:
            next_progress.update(extra)
        updated = {
            **run,
            "stage": stage,
            "progress": next_progress,
            "updated_at": now_iso(),
        }
        self._upsert_run(updated)

    def _load_messages_for_run(self, run: dict[str, Any]) -> list[dict[str, Any]]:
        filters = run.get("filters") if isinstance(run.get("filters"), dict) else {}
        conversation_ids = [str(item) for item in filters.get("conversation_ids", []) if str(item)]
        target_names = [str(item) for item in filters.get("target_names", []) if str(item)]
        if target_names:
            conversations = self.raw_store.list_conversations(status="all", limit=500)
            lookup = {str(item.get("target_name") or ""): str(item.get("conversation_id") or "") for item in conversations}
            for target_name in target_names:
                conv_id = lookup.get(target_name, "")
                if conv_id:
                    conversation_ids.append(conv_id)
        conversation_ids = sorted(set(item for item in conversation_ids if item))
        if not conversation_ids:
            return self.raw_store.list_messages_advanced(
                query=str(filters.get("query") or ""),
                limit=int(filters.get("limit", 10000) or 10000),
                offset=int(filters.get("offset", 0) or 0),
                start_time=str(filters.get("start_time") or ""),
                end_time=str(filters.get("end_time") or ""),
                sender=str(filters.get("sender") or ""),
                content_type=str(filters.get("content_type") or ""),
                conversation_type=str(filters.get("conversation_type") or ""),
                keywords=[str(item) for item in filters.get("keywords", []) if str(item)],
            )
        merged: dict[str, dict[str, Any]] = {}
        for conversation_id in conversation_ids:
            items = self.raw_store.list_messages_advanced(
                conversation_id=conversation_id,
                query=str(filters.get("query") or ""),
                limit=int(filters.get("limit", 10000) or 10000),
                offset=int(filters.get("offset", 0) or 0),
                start_time=str(filters.get("start_time") or ""),
                end_time=str(filters.get("end_time") or ""),
                sender=str(filters.get("sender") or ""),
                content_type=str(filters.get("content_type") or ""),
                conversation_type=str(filters.get("conversation_type") or ""),
                keywords=[str(item) for item in filters.get("keywords", []) if str(item)],
            )
            for item in items:
                message_id = str(item.get("raw_message_id") or item.get("dedupe_key") or "")
                if message_id:
                    merged[message_id] = item
        return sorted(merged.values(), key=recorder_time_text, reverse=True)

    def _build_workbook(self, run: dict[str, Any], messages: list[dict[str, Any]], *, run_id: str = "") -> dict[str, Any]:
        module_key = str(run.get("module_key") or "")
        if module_key == "raw_message_log_v1":
            path = self._build_raw_message_workbook(run, messages)
            csv_path = self._build_raw_message_csv(run, messages)
            return {
                "xlsx_path": str(path),
                "csv_path": str(csv_path),
                "export_row_count": len(messages),
                "needs_review_count": 0,
                "skipped_count": 0,
                "rows_preview": [],
            }
        if module_key == "order_sheet_lab_v1":
            extracted, extraction_meta = self._extract_order_rows(run, messages, run_id=run_id)
            path = self._build_order_sheet_workbook(run, extracted)
            csv_path = self._build_order_sheet_csv(run, extracted)
            needs_review = sum(1 for item in extracted if item.get("needs_review"))
            return {
                "xlsx_path": str(path),
                "csv_path": str(csv_path),
                "export_row_count": len(extracted),
                "needs_review_count": needs_review,
                "skipped_count": max(0, len(messages) - len(extracted)),
                "rows_preview": extracted,
                "llm_calls": int(extraction_meta.get("llm_calls", 0) or 0),
                "llm_segment_calls": int(extraction_meta.get("llm_segment_calls", 0) or 0),
                "llm_repair_calls": int(extraction_meta.get("llm_repair_calls", 0) or 0),
                "llm_brand_calls": int(extraction_meta.get("llm_brand_calls", 0) or 0),
            }
        raise ValueError(f"unsupported module for current generic runtime: {module_key}")

    def _build_raw_message_workbook(self, run: dict[str, Any], messages: list[dict[str, Any]]) -> Path:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "记录导出"
        headers = ["会话ID", "会话类型", "发送人", "消息类型", "程序读取时间", "消息内容", "原始消息ID", "群成员名", "OCR质量标志", "被排除引用"]
        header_fill = PatternFill("solid", fgColor="EAF2F8")
        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=column, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for row_index, item in enumerate(messages, start=2):
            values = [
                item.get("conversation_id") or "",
                item.get("conversation_type") or "",
                item.get("sender") or item.get("sender_role") or "",
                item.get("content_type") or "",
                recorder_time_text(item),
                recorder_content_text(item),
                item.get("raw_message_id") or "",
                item.get("speaker_name") or item.get("group_member_name") or "",
                ",".join(str(flag) for flag in item.get("quality_flags", []) if str(flag)),
                "; ".join(str(fragment.get("text") or "") for fragment in item.get("quoted_fragments", []) if isinstance(fragment, dict)),
            ]
            for column, value in enumerate(values, start=1):
                sheet.cell(row=row_index, column=column, value=value)
        widths = [24, 14, 16, 12, 20, 80, 28, 16, 28, 42]
        for idx, width in enumerate(widths, start=1):
            sheet.column_dimensions[sheet.cell(row=1, column=idx).column_letter].width = width

        self.files_root.mkdir(parents=True, exist_ok=True)
        output_path = self.files_root / f"{run['run_id']}.xlsx"
        workbook.save(output_path)
        return output_path

    def _build_raw_message_csv(self, run: dict[str, Any], messages: list[dict[str, Any]]) -> Path:
        self.files_root.mkdir(parents=True, exist_ok=True)
        output_path = self.files_root / f"{run['run_id']}.csv"
        headers = ["会话ID", "会话类型", "发送人", "消息类型", "程序读取时间", "消息内容", "原始消息ID", "群成员名", "OCR质量标志", "被排除引用"]
        with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(headers)
            for item in messages:
                writer.writerow(
                    [
                        item.get("conversation_id") or "",
                        item.get("conversation_type") or "",
                        item.get("sender") or item.get("sender_role") or "",
                        item.get("content_type") or "",
                        recorder_time_text(item),
                        recorder_content_text(item),
                        item.get("raw_message_id") or "",
                        item.get("speaker_name") or item.get("group_member_name") or "",
                        ",".join(str(flag) for flag in item.get("quality_flags", []) if str(flag)),
                        "; ".join(str(fragment.get("text") or "") for fragment in item.get("quoted_fragments", []) if isinstance(fragment, dict)),
                    ]
                )
        return output_path

    def _build_order_sheet_workbook(self, run: dict[str, Any], rows: list[dict[str, Any]]) -> Path:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        header_fill = PatternFill("solid", fgColor="EAF2F8")
        for column, header in enumerate(ORDER_HEADERS, start=1):
            cell = sheet.cell(row=1, column=column, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill

        for row_index, item in enumerate(rows, start=2):
            values = [
                "",
                item.get("date") or "",
                item.get("time") or "",
                item.get("name") or "",
                item.get("owner") or "",
                item.get("receiver") or "",
                item.get("campus") or "",
                item.get("order_unit") or "",
                item.get("product_name") or "",
                item.get("quantity") or "",
                item.get("unit") or "",
                item.get("spec") or "",
                item.get("brand") or "",
                item.get("cost_price") or "",
                item.get("sale_price") or "",
                item.get("total_cost") or "",
                item.get("total_sale") or "",
                item.get("remark") or "",
            ]
            row_fill = self._confidence_fill_for_row(item)
            for column, value in enumerate(values, start=1):
                cell = sheet.cell(row=row_index, column=column, value=value)
                if row_fill is not None:
                    cell.fill = row_fill
        widths = [8, 10, 10, 14, 16, 14, 12, 12, 32, 10, 10, 18, 12, 12, 12, 12, 12, 28]
        for idx, width in enumerate(widths, start=1):
            sheet.column_dimensions[sheet.cell(row=1, column=idx).column_letter].width = width

        meta = workbook.create_sheet(title="抽取报告")
        meta_rows = [
            ("模块", str(run.get("module_key") or "")),
            ("版本", str(run.get("module_version") or "")),
            ("导出时间", now_iso()),
            ("记录行数", len(rows)),
            ("需复核行数", sum(1 for item in rows if item.get("needs_review"))),
        ]
        for index, (key, value) in enumerate(meta_rows, start=1):
            meta.cell(row=index, column=1, value=key).font = Font(bold=True)
            meta.cell(row=index, column=2, value=value)
        meta.cell(row=8, column=1, value="抽取结果（完整）").font = Font(bold=True)
        report_headers = ["日期", "时间", "姓名", "责任人", "货品", "数量", "单位", "售价", "总售价", "置信度", "需复核", "记录类型", "风险标志", "证据消息ID", "证据片段"]
        for column, header in enumerate(report_headers, start=1):
            cell = meta.cell(row=9, column=column, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for row_index, item in enumerate(rows[:100], start=10):
            values = [
                item.get("date") or "",
                item.get("time") or "",
                item.get("name") or "",
                item.get("owner") or "",
                item.get("product_name") or "",
                item.get("quantity") or "",
                item.get("unit") or "",
                item.get("sale_price") or "",
                item.get("total_sale") or "",
                float(item.get("confidence") or 0),
                "是" if item.get("needs_review") else "",
                item.get("record_type") or "order_item",
                ",".join(str(x) for x in item.get("risk_flags", []) if str(x)),
                ",".join(str(x) for x in item.get("evidence_message_ids", []) if str(x)),
                item.get("evidence_text") or "",
            ]
            for column, value in enumerate(values, start=1):
                meta.cell(row=row_index, column=column, value=value)
        for idx, width in enumerate([10, 10, 14, 14, 24, 10, 10, 10, 10, 10, 10, 12, 22, 24, 42], start=1):
            meta.column_dimensions[meta.cell(row=9, column=idx).column_letter].width = width

        self.files_root.mkdir(parents=True, exist_ok=True)
        output_path = self.files_root / f"{run['run_id']}.xlsx"
        workbook.save(output_path)
        return output_path

    def _build_order_sheet_csv(self, run: dict[str, Any], rows: list[dict[str, Any]]) -> Path:
        self.files_root.mkdir(parents=True, exist_ok=True)
        output_path = self.files_root / f"{run['run_id']}.csv"
        with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(ORDER_HEADERS)
            for item in rows:
                writer.writerow(
                    [
                        "",
                        item.get("date") or "",
                        item.get("time") or "",
                        item.get("name") or "",
                        item.get("owner") or "",
                        item.get("receiver") or "",
                        item.get("campus") or "",
                        item.get("order_unit") or "",
                        item.get("product_name") or "",
                        item.get("quantity") or "",
                        item.get("unit") or "",
                        item.get("spec") or "",
                        item.get("brand") or "",
                        item.get("cost_price") or "",
                        item.get("sale_price") or "",
                        item.get("total_cost") or "",
                        item.get("total_sale") or "",
                        item.get("remark") or "",
                    ]
                )
        return output_path

    def _write_report(self, run: dict[str, Any], messages: list[dict[str, Any]], workbook_result: dict[str, Any]) -> Path:
        self.reports_root.mkdir(parents=True, exist_ok=True)
        report = {
            "run_id": run.get("run_id"),
            "tenant_id": run.get("tenant_id"),
            "module_key": run.get("module_key"),
            "module_version": run.get("module_version"),
            "message_count": len(messages),
            "artifact": str(workbook_result.get("xlsx_path") or ""),
            "csv_artifact": str(workbook_result.get("csv_path") or ""),
            "artifacts": {
                "xlsx_path": str(workbook_result.get("xlsx_path") or ""),
                "csv_path": str(workbook_result.get("csv_path") or ""),
            },
            "stats": {
                "export_row_count": int(workbook_result.get("export_row_count", 0) or 0),
                "needs_review_count": int(workbook_result.get("needs_review_count", 0) or 0),
                "skipped_count": int(workbook_result.get("skipped_count", 0) or 0),
                "llm_calls": int(workbook_result.get("llm_calls", 0) or 0),
                "llm_segment_calls": int(workbook_result.get("llm_segment_calls", 0) or 0),
                "llm_repair_calls": int(workbook_result.get("llm_repair_calls", 0) or 0),
                "llm_brand_calls": int(workbook_result.get("llm_brand_calls", 0) or 0),
            },
            "generated_at": now_iso(),
            "filters": run.get("filters") if isinstance(run.get("filters"), dict) else {},
            "rows_preview": workbook_result.get("rows_preview", []),
        }
        report_path = self.reports_root / f"{run['run_id']}.json"
        report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        return report_path

    def _extract_order_rows(self, run: dict[str, Any], messages: list[dict[str, Any]], *, run_id: str = "") -> tuple[list[dict[str, Any]], dict[str, int]]:
        module = self.modules.get_module(str(run.get("module_key") or "")) or {}
        config = module.get("config") if isinstance(module.get("config"), dict) else {}
        llm_enabled = bool(config.get("llm_enabled", True))
        llm_budget = max(0, int(config.get("llm_max_rows_per_run", 12) or 12))
        llm_repair_budget = max(0, int(config.get("llm_repair_max_rows_per_run", llm_budget) or llm_budget))
        llm_dynamic_budget_enabled = bool(config.get("llm_dynamic_budget_enabled", True))
        llm_dynamic_budget_ratio = max(0.0, min(float(config.get("llm_dynamic_budget_ratio", 0.35) or 0.35), 2.0))
        llm_dynamic_budget_max = max(llm_budget, int(config.get("llm_dynamic_budget_max", 64) or 64))
        llm_dynamic_repair_ratio = max(0.0, min(float(config.get("llm_dynamic_repair_ratio", 0.25) or 0.25), 2.0))
        llm_dynamic_repair_max = max(llm_repair_budget, int(config.get("llm_dynamic_repair_max", 32) or 32))
        llm_skip_strong_rule_rows = bool(config.get("llm_skip_strong_rule_rows", True))
        llm_segmentation_enabled = bool(config.get("llm_segmentation_enabled", True))
        llm_segmentation_max_segments = max(1, min(int(config.get("llm_segmentation_max_segments", 6) or 6), 12))
        llm_supplement_mode = str(config.get("llm_supplement_mode") or "missing_core_fields_only")
        llm_parallel_workers = max(1, min(int(config.get("llm_parallel_workers", 4) or 4), 8))
        date_output_mode = str(config.get("date_output_mode") or "YYYY-MM-DD")
        extract_mode = str(config.get("extract_mode") or "llm_first").strip().lower()
        missing_quantity_strategy = str(config.get("missing_quantity_strategy") or "strict").strip().lower()
        if missing_quantity_strategy not in {"strict", "assume_one", "llm_guess"}:
            missing_quantity_strategy = "strict"
        include_record_types_raw = config.get("include_record_types_in_main_sheet")
        if isinstance(include_record_types_raw, list):
            include_record_types = {
                str(item).strip().lower()
                for item in include_record_types_raw
                if str(item).strip()
            }
        else:
            include_record_types = set(DEFAULT_INCLUDE_RECORD_TYPES)
        if not include_record_types:
            include_record_types = set(DEFAULT_INCLUDE_RECORD_TYPES)
        force_multi_sku_split_enabled = bool(config.get("force_multi_sku_split_enabled", True))
        force_multi_sku_min_skus = max(2, int(config.get("force_multi_sku_min_skus", 2) or 2))
        force_multi_order_signal_threshold = max(1, int(config.get("force_multi_order_signal_threshold", 2) or 2))
        context_followup_enabled = bool(config.get("context_followup_enabled", True))
        name_context_fallback_enabled = bool(config.get("name_context_fallback_enabled", True))
        name_context_max_messages = max(1, min(int(config.get("name_context_max_messages", 3) or 3), 3))
        name_context_max_minutes = max(1, int(config.get("name_context_max_minutes", 5) or 5))
        name_context_confidence_penalty = max(
            0.0,
            min(float(config.get("name_context_confidence_penalty", 0.06) or 0.06), 0.3),
        )
        brand_llm_inference_enabled = bool(config.get("brand_llm_inference_enabled", True))
        brand_llm_inference_max_calls = max(0, int(config.get("brand_llm_inference_max_calls_per_run", 20) or 20))
        brand_llm_inference_min_confidence = max(0.0, min(float(config.get("brand_llm_inference_min_confidence", 0.62) or 0.62), 1.0))
        content_types = {str(item).lower() for item in (config.get("supported_content_types") or ["text", "quote"])}
        self._active_brand_aliases = self._normalize_brand_aliases(config.get("brand_aliases"))

        ordered_messages = sorted(
            [recorder_message_for_export(item) for item in messages if isinstance(item, dict)],
            key=recorder_time_text,
        )
        total_messages = len(ordered_messages)
        candidates: list[dict[str, Any]] = []
        rows: list[dict[str, Any]] = []
        recent_rows: list[dict[str, Any]] = []
        llm_extract_calls = 0
        llm_segment_calls = 0
        llm_repair_calls = 0
        llm_brand_calls = 0

        # Pass-1: rule baseline
        for index, message in enumerate(ordered_messages, start=1):
            content_type = str(message.get("content_type") or "text").lower()
            if content_types and content_type not in content_types:
                continue
            content = str(message.get("content") or "").strip()
            if not content:
                continue
            looks_like_rule_order = self._looks_like_order_message(content)
            looks_like_llm_candidate = self._looks_like_llm_order_candidate(content)
            is_followup = context_followup_enabled and self._looks_like_followup_confirmation(content)
            if not looks_like_rule_order and not looks_like_llm_candidate and not is_followup:
                continue
            baseline_rows = (
                self._rule_extract_rows_from_message(message, date_output_mode=date_output_mode)
                if extract_mode != "llm_only" and not is_followup
                else []
            )
            candidates.append(
                {
                    "index": index,
                    "message": message,
                    "content": content,
                    "looks_like_rule_order": looks_like_rule_order,
                    "looks_like_llm_candidate": looks_like_llm_candidate,
                    "is_followup": is_followup,
                    "baseline_rows": baseline_rows,
                    "complexity": self._message_complexity_score(content),
                    "force_multi_split": bool(
                        force_multi_sku_split_enabled
                        and self._should_force_multi_sku_split(
                            content,
                            min_skus=force_multi_sku_min_skus,
                            action_threshold=force_multi_order_signal_threshold,
                        )
                    ),
                }
            )
            if run_id and (index % 25 == 0 or index == total_messages):
                self._update_run_progress(
                    run_id,
                    stage="scanning",
                    processed=index,
                    total=total_messages,
                    force_percent=progress_between(0.05, 0.20, index, total_messages),
                    stage_detail=f"正在筛选候选订单消息 {index}/{total_messages}",
                    extra={"unit_label": "消息"},
                )

        if llm_dynamic_budget_enabled and candidates:
            llm_candidate_count = sum(1 for item in candidates if item.get("looks_like_llm_candidate"))
            if llm_candidate_count > 0:
                scaled_budget = max(llm_budget, int(round(llm_candidate_count * llm_dynamic_budget_ratio)))
                scaled_repair = max(llm_repair_budget, int(round(llm_candidate_count * llm_dynamic_repair_ratio)))
                llm_budget = min(llm_dynamic_budget_max, scaled_budget)
                llm_repair_budget = min(llm_dynamic_repair_max, scaled_repair)

        # Pass-2: LLM upgrade for complex / low-confidence candidates
        if llm_enabled and llm_budget > 0:
            if extract_mode in {"llm_only", "llm_first", "hybrid_llm_first"}:
                llm_pool = [item for item in candidates if item.get("looks_like_llm_candidate") and not item.get("is_followup")]
                if llm_skip_strong_rule_rows and extract_mode != "llm_only":
                    llm_pool = [item for item in llm_pool if self._candidate_needs_llm_upgrade(item)]
                llm_targets = sorted(
                    llm_pool,
                    key=lambda item: (int(item.get("complexity", 0) or 0), -int(item.get("index", 0) or 0)),
                    reverse=True,
                )
            else:
                llm_targets = sorted(
                    [
                        item
                        for item in candidates
                        if not item.get("is_followup")
                        if not item.get("baseline_rows")
                        or any(bool(row.get("needs_review")) for row in item.get("baseline_rows", []))
                        or int(item.get("complexity", 0) or 0) >= 2
                    ],
                    key=lambda item: (int(item.get("complexity", 0) or 0), -int(item.get("index", 0) or 0)),
                    reverse=True,
                )
            llm_target_total = min(llm_budget, len(llm_targets))
            if run_id:
                self._update_run_progress(
                    run_id,
                    stage="llm_extracting",
                    processed=0,
                    total=llm_target_total,
                    force_percent=progress_between(0.20, 0.68, 0, llm_target_total),
                    stage_detail=f"准备进行 LLM 语义抽取，共 {llm_target_total} 次调用，并发 {min(llm_parallel_workers, max(1, llm_target_total))}",
                    extra={"unit_label": "LLM调用"},
                )

            def apply_llm_result(target: dict[str, Any], content: str, llm_rows: list[dict[str, Any]]) -> None:
                if not llm_rows:
                    return
                baseline_rows = target.get("baseline_rows") if isinstance(target.get("baseline_rows"), list) else []
                if extract_mode == "llm_only":
                    target["selected_rows"] = llm_rows
                    return
                baseline_quality = self._rows_quality_score(baseline_rows, source_text=content)
                llm_quality = self._rows_quality_score(llm_rows, source_text=content)
                complexity = int(target.get("complexity", 0) or 0)
                baseline_count = len(baseline_rows)
                llm_count = len(llm_rows)
                group_hint = bool(re.search(r"(这\s*\d+\s*(?:个|株|盒|瓶|支|包)|各订|分别|买二送一|赠品选择)", content))
                baseline_strong = any(
                    bool(str(item.get("product_name") or "").strip())
                    and bool(str(item.get("quantity") or "").strip())
                    and bool(item.get("needs_review") is False)
                    for item in baseline_rows
                    if isinstance(item, dict)
                )
                llm_strong = any(
                    bool(str(item.get("product_name") or "").strip())
                    and bool(str(item.get("quantity") or "").strip())
                    for item in llm_rows
                    if isinstance(item, dict)
                )
                prefer_llm = False
                # Prefer LLM when baseline empty, low quality, or LLM quality is better.
                if not baseline_rows or llm_quality >= baseline_quality + 0.5:
                    prefer_llm = True
                elif complexity >= 2 and llm_quality >= baseline_quality:
                    prefer_llm = True
                elif complexity >= 4 and llm_count > baseline_count:
                    prefer_llm = True
                elif group_hint and llm_count >= baseline_count and llm_quality + 0.2 >= baseline_quality:
                    prefer_llm = True
                # Guardrail: when baseline is already strong, do not let unstable LLM output replace it.
                if baseline_rows and baseline_strong and not bool(target.get("force_multi_split")):
                    if llm_count < baseline_count or not llm_strong or llm_quality < baseline_quality + 0.8:
                        prefer_llm = False
                if prefer_llm:
                    target["selected_rows"] = llm_rows

            def extract_target(target: dict[str, Any]) -> tuple[dict[str, Any], str, list[dict[str, Any]], str]:
                message = target.get("message") if isinstance(target.get("message"), dict) else {}
                content = str(target.get("content") or "")
                try:
                    return (
                        target,
                        content,
                        self._llm_extract_rows(
                            message,
                            date_output_mode=date_output_mode,
                            use_segmentation=llm_segmentation_enabled,
                            max_segments=llm_segmentation_max_segments,
                            force_multi_split=bool(target.get("force_multi_split")),
                        ),
                        "",
                    )
                except Exception as exc:  # pragma: no cover - keeps one flaky LLM call from killing the export.
                    return target, content, [], repr(exc)

            limited_llm_targets = llm_targets[:llm_target_total]
            worker_count = min(llm_parallel_workers, max(1, llm_target_total))
            if worker_count > 1 and llm_target_total > 1:
                with ThreadPoolExecutor(max_workers=worker_count) as executor:
                    future_to_target = {executor.submit(extract_target, target): target for target in limited_llm_targets}
                    for future in as_completed(future_to_target):
                        target, content, llm_rows, llm_error = future.result()
                        llm_extract_calls += 1
                        if llm_segmentation_enabled:
                            llm_segment_calls += 1
                        if llm_error:
                            target["llm_error"] = llm_error
                        apply_llm_result(target, content, llm_rows)
                        if run_id:
                            self._update_run_progress(
                                run_id,
                                stage="llm_extracting",
                                processed=llm_extract_calls,
                                total=llm_target_total,
                                force_percent=progress_between(0.20, 0.68, llm_extract_calls, llm_target_total),
                                stage_detail=f"并发 LLM 语义抽取已完成 {llm_extract_calls}/{llm_target_total}",
                                extra={"unit_label": "LLM调用"},
                            )
            else:
                for target in limited_llm_targets:
                    if run_id:
                        next_call = llm_extract_calls + 1
                        self._update_run_progress(
                            run_id,
                            stage="llm_extracting",
                            processed=llm_extract_calls,
                            total=llm_target_total,
                            force_percent=progress_between(0.20, 0.68, llm_extract_calls, llm_target_total),
                            stage_detail=f"正在调用 LLM 语义抽取 {next_call}/{llm_target_total}",
                            extra={"unit_label": "LLM调用"},
                        )
                    target, content, llm_rows, llm_error = extract_target(target)
                    llm_extract_calls += 1
                    if llm_segmentation_enabled:
                        llm_segment_calls += 1
                    if llm_error:
                        target["llm_error"] = llm_error
                    apply_llm_result(target, content, llm_rows)
                    if run_id:
                        self._update_run_progress(
                            run_id,
                            stage="llm_extracting",
                            processed=llm_extract_calls,
                            total=llm_target_total,
                            force_percent=progress_between(0.20, 0.68, llm_extract_calls, llm_target_total),
                            stage_detail=f"LLM 语义抽取已完成 {llm_extract_calls}/{llm_target_total}",
                            extra={"unit_label": "LLM调用"},
                        )

        # Pass-3: finalize and row-level supplement
        finalize_targets = sorted(candidates, key=lambda item: int(item.get("index", 0) or 0))
        finalize_total = len(finalize_targets)
        for finalize_index, target in enumerate(finalize_targets, start=1):
            if run_id and (finalize_index % 25 == 1 or finalize_index == finalize_total):
                self._update_run_progress(
                    run_id,
                    stage="finalizing",
                    processed=finalize_index,
                    total=finalize_total,
                    force_percent=progress_between(0.68, 0.86, finalize_index, finalize_total),
                    stage_detail=f"正在整理导出行 {finalize_index}/{finalize_total}",
                    extra={"unit_label": "候选消息"},
                )
            message = target.get("message") if isinstance(target.get("message"), dict) else {}
            content = str(target.get("content") or "")
            message_index = int(target.get("index", 0) or 0)
            if target.get("is_followup"):
                if context_followup_enabled:
                    self._apply_followup_confirmation(content, recent_rows=recent_rows)
                continue
            selected_rows = target.get("selected_rows") if isinstance(target.get("selected_rows"), list) else target.get("baseline_rows", [])
            selected_rows = self._apply_brand_context_to_rows(selected_rows, source_text=content)
            if (
                llm_enabled
                and brand_llm_inference_enabled
                and llm_brand_calls < brand_llm_inference_max_calls
                and selected_rows
                and any(not str((row or {}).get("brand") or "").strip() for row in selected_rows if isinstance(row, dict))
            ):
                next_brand_call = llm_brand_calls + 1
                if run_id:
                    self._update_run_progress(
                        run_id,
                        stage="llm_branding",
                        processed=llm_brand_calls,
                        total=brand_llm_inference_max_calls,
                        force_percent=progress_between(0.86, 0.92, llm_brand_calls, brand_llm_inference_max_calls),
                        stage_detail=f"正在进行品牌推断 {next_brand_call}/{brand_llm_inference_max_calls}",
                        extra={"unit_label": "品牌LLM调用"},
                    )
                inferred_rows = self._llm_infer_brands_for_rows(
                    source_text=content,
                    rows=selected_rows,
                    min_confidence=brand_llm_inference_min_confidence,
                )
                llm_brand_calls += 1
                if run_id:
                    self._update_run_progress(
                        run_id,
                        stage="llm_branding",
                        processed=llm_brand_calls,
                        total=brand_llm_inference_max_calls,
                        force_percent=progress_between(0.86, 0.92, llm_brand_calls, brand_llm_inference_max_calls),
                        stage_detail=f"品牌推断已完成 {llm_brand_calls}/{brand_llm_inference_max_calls}",
                        extra={"unit_label": "品牌LLM调用"},
                    )
                if inferred_rows:
                    selected_rows = inferred_rows
            if not selected_rows:
                continue
            used_spec_owners: dict[str, str] = {}
            message_scope_candidates = self._split_message_by_price_terminator(content, max_segments=max(4, llm_segmentation_max_segments))
            if not message_scope_candidates:
                message_scope_candidates = [self._normalize_order_text(content)]
            for row in selected_rows:
                row_input = dict(row) if isinstance(row, dict) else {}
                if not self._normalize_order_text(str(row_input.get("source_scope_text") or "")):
                    row_input["source_scope_text"] = self._guess_scope_text_for_row(
                        row_input,
                        candidate_segments=message_scope_candidates,
                        fallback_text=content,
                    )
                finalized = self._finalize_order_row(row_input, message, date_output_mode=date_output_mode)
                finalized["record_type"] = self._classify_record_type(finalized, source_text=content)
                finalized = self._apply_missing_quantity_strategy(
                    finalized,
                    source_text=content,
                    strategy=missing_quantity_strategy,
                )
                if self._should_drop_order_row(finalized):
                    continue
                if llm_enabled and finalized.get("needs_review") and llm_repair_calls < llm_repair_budget and llm_extract_calls < llm_budget:
                    if run_id:
                        next_repair_call = llm_repair_calls + 1
                        self._update_run_progress(
                            run_id,
                            stage="reviewing",
                            processed=llm_repair_calls,
                            total=llm_repair_budget,
                            force_percent=progress_between(0.92, 0.94, llm_repair_calls, llm_repair_budget),
                            stage_detail=f"正在进行行级 LLM 修复 {next_repair_call}/{llm_repair_budget}",
                            extra={"unit_label": "修复LLM调用"},
                        )
                    llm_repair_calls += 1
                    llm_extract_calls += 1
                    improved = self._llm_supplement_row(
                        message,
                        finalized,
                        date_output_mode=date_output_mode,
                        mode=llm_supplement_mode,
                    )
                    if run_id:
                        self._update_run_progress(
                            run_id,
                            stage="reviewing",
                            processed=llm_repair_calls,
                            total=llm_repair_budget,
                            force_percent=progress_between(0.92, 0.94, llm_repair_calls, llm_repair_budget),
                            stage_detail=f"行级 LLM 修复已完成 {llm_repair_calls}/{llm_repair_budget}",
                            extra={"unit_label": "修复LLM调用"},
                        )
                    if improved:
                        finalized = improved
                finalized = self._finalize_order_row(finalized, message, date_output_mode=date_output_mode)
                finalized = self._enforce_single_use_spec_for_row(
                    finalized,
                    used_spec_owners=used_spec_owners,
                    source_text=content,
                )
                if name_context_fallback_enabled:
                    finalized = self._apply_name_context_fallback_to_row(
                        finalized,
                        ordered_messages=ordered_messages,
                        message_index=message_index,
                        max_messages=name_context_max_messages,
                        max_minutes=name_context_max_minutes,
                        confidence_penalty=name_context_confidence_penalty,
                    )
                finalized["record_type"] = self._classify_record_type(finalized, source_text=content)
                finalized = self._apply_missing_quantity_strategy(
                    finalized,
                    source_text=content,
                    strategy=missing_quantity_strategy,
                )
                record_type = str(finalized.get("record_type") or "order_item").strip().lower() or "order_item"
                if record_type not in include_record_types:
                    continue
                rows.append(finalized)
                recent_rows.append(finalized)
                if len(recent_rows) > 80:
                    recent_rows = recent_rows[-80:]

        if run_id:
            self._update_run_progress(
                run_id,
                stage="reviewing",
                processed=total_messages,
                total=total_messages,
                force_percent=0.94,
                stage_detail="正在汇总抽取结果并准备生成 Excel",
                extra={"unit_label": "消息"},
            )
        rows = self._dedupe_close_export_rows(rows)
        return rows, {
            "llm_calls": llm_extract_calls,
            "llm_segment_calls": llm_segment_calls,
            "llm_repair_calls": llm_repair_calls,
            "llm_brand_calls": llm_brand_calls,
        }

    def _dedupe_close_export_rows(self, rows: list[dict[str, Any]], *, max_minutes: int = 10) -> list[dict[str, Any]]:
        if not rows:
            return []
        max_seconds = max(1, int(max_minutes or 10)) * 60
        kept: list[dict[str, Any]] = []
        index_by_key: dict[tuple[str, ...], list[int]] = {}
        for row in rows:
            candidate = dict(row) if isinstance(row, dict) else {}
            if not candidate:
                continue
            key = self._close_export_duplicate_key(candidate)
            if not key:
                kept.append(candidate)
                continue
            candidate_seconds = self._export_row_time_seconds(candidate)
            duplicate_index: int | None = None
            for kept_index in reversed(index_by_key.get(key, [])):
                existing = kept[kept_index]
                existing_seconds = self._export_row_time_seconds(existing)
                if candidate_seconds is None or existing_seconds is None:
                    continue
                if abs(candidate_seconds - existing_seconds) <= max_seconds:
                    duplicate_index = kept_index
                    break
            if duplicate_index is None:
                index_by_key.setdefault(key, []).append(len(kept))
                kept.append(candidate)
                continue
            kept[duplicate_index] = self._prefer_close_duplicate_export_row(kept[duplicate_index], candidate)
        return kept

    def _close_export_duplicate_key(self, row: dict[str, Any]) -> tuple[str, ...]:
        product = self._normalize_export_duplicate_text(row.get("product_name"), remove_spaces=True)
        sale_price = self._normalize_export_duplicate_number(row.get("sale_price"))
        total_sale = self._normalize_export_duplicate_number(row.get("total_sale"))
        if not product or not (sale_price or total_sale):
            return ()
        return (
            self._normalize_export_duplicate_text(row.get("date")),
            self._normalize_export_duplicate_text(row.get("name"), remove_spaces=True),
            self._normalize_export_duplicate_text(row.get("owner"), remove_spaces=True),
            self._normalize_export_duplicate_text(row.get("receiver"), remove_spaces=True),
            self._normalize_export_duplicate_text(row.get("record_type")),
            self._normalize_export_duplicate_text(row.get("brand"), remove_spaces=True),
            product,
            self._normalize_spec_key(str(row.get("spec") or "")),
            self._normalize_export_duplicate_number(row.get("quantity")),
            self._normalize_export_duplicate_text(row.get("unit"), remove_spaces=True),
            sale_price,
            total_sale,
        )

    @staticmethod
    def _normalize_export_duplicate_text(value: Any, *, remove_spaces: bool = False) -> str:
        text = re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip().lower()
        if remove_spaces:
            text = re.sub(r"\s+", "", text)
        return text

    def _normalize_export_duplicate_number(self, value: Any) -> str:
        text = str(value or "").strip()
        if not text:
            return ""
        number = self._to_number(text)
        if number:
            return self._format_number(number)
        return self._normalize_export_duplicate_text(text, remove_spaces=True)

    @staticmethod
    def _export_row_time_seconds(row: dict[str, Any]) -> int | None:
        text = str(row.get("time") or "").strip()
        match = re.fullmatch(r"(?P<hour>\d{1,2}):(?P<minute>\d{2})(?::(?P<second>\d{2}))?", text)
        if not match:
            return None
        hour = int(match.group("hour"))
        minute = int(match.group("minute"))
        second = int(match.group("second") or 0)
        if hour > 23 or minute > 59 or second > 59:
            return None
        return hour * 3600 + minute * 60 + second

    def _prefer_close_duplicate_export_row(self, existing: dict[str, Any], candidate: dict[str, Any]) -> dict[str, Any]:
        existing_score = self._duplicate_row_quality_score(existing)
        candidate_score = self._duplicate_row_quality_score(candidate)
        candidate_wins = candidate_score > existing_score + 0.02
        winner = dict(candidate if candidate_wins else existing)
        loser = existing if candidate_wins else candidate
        winner["evidence_message_ids"] = self._merge_unique_list_values(
            winner.get("evidence_message_ids"),
            loser.get("evidence_message_ids"),
        )
        winner["risk_flags"] = self._merge_unique_list_values(winner.get("risk_flags"), loser.get("risk_flags"))
        for key in ("remark", "evidence_text"):
            if not str(winner.get(key) or "").strip() and str(loser.get(key) or "").strip():
                winner[key] = loser.get(key)
        return winner

    def _duplicate_row_quality_score(self, row: dict[str, Any]) -> float:
        score = self._coerce_confidence(row.get("confidence"))
        if not row.get("needs_review"):
            score += 0.04
        for key in ("brand", "product_name", "spec", "quantity", "unit", "sale_price", "total_sale", "name"):
            if str(row.get(key) or "").strip():
                score += 0.01
        return score

    @staticmethod
    def _merge_unique_list_values(*values: Any) -> list[str]:
        merged: list[str] = []
        seen: set[str] = set()
        for value in values:
            items = value if isinstance(value, list) else [value]
            for item in items:
                text = str(item or "").strip()
                if not text or text in seen:
                    continue
                seen.add(text)
                merged.append(text)
        return merged

    def _candidate_needs_llm_upgrade(self, candidate: dict[str, Any]) -> bool:
        if bool(candidate.get("force_multi_split")):
            return True
        complexity = int(candidate.get("complexity", 0) or 0)
        if complexity >= 2:
            return True
        content = self._normalize_order_text(str(candidate.get("content") or ""))
        if re.search(r"(这\s*\d+\s*(?:个|株|盒|瓶|支|包)|各订|分别|买二送一|赠品选择)", content):
            return True
        baseline_rows = candidate.get("baseline_rows") if isinstance(candidate.get("baseline_rows"), list) else []
        if not baseline_rows:
            return True
        if any(bool(row.get("needs_review")) for row in baseline_rows if isinstance(row, dict)):
            return True
        valid_rows = [row for row in baseline_rows if isinstance(row, dict)]
        if not valid_rows:
            return True
        strong_rows = 0
        for row in valid_rows:
            if (
                str(row.get("product_name") or "").strip()
                and str(row.get("quantity") or "").strip()
                and (str(row.get("sale_price") or "").strip() or str(row.get("total_sale") or "").strip())
            ):
                strong_rows += 1
        if strong_rows >= len(valid_rows):
            return False
        return True

    def _normalize_brand_aliases(self, raw_aliases: Any) -> list[str]:
        aliases: list[str] = []
        if isinstance(raw_aliases, list):
            for item in raw_aliases:
                token = self._normalize_order_text(str(item or ""))
                if token:
                    aliases.append(token)
        aliases.extend(list(DEFAULT_LAB_BRAND_ALIASES))
        seen: set[str] = set()
        deduped: list[str] = []
        for alias in aliases:
            key = alias.lower()
            if key in seen:
                continue
            seen.add(key)
            deduped.append(alias)
        deduped.sort(key=len, reverse=True)
        return deduped

    def _extract_brand_name(self, text: str, *, fallback: str = "") -> str:
        normalized = self._normalize_order_text(text)
        fallback_token = self._normalize_order_text(fallback)
        if not normalized:
            return fallback_token
        aliases = self._active_brand_aliases or list(DEFAULT_LAB_BRAND_ALIASES)
        alias_keys = {self._normalize_order_text(item).lower() for item in aliases}
        for alias in aliases:
            escaped = re.escape(alias)
            if re.search(rf"(^|[\s，,；;:：/]){escaped}($|[\s，,；;:：/])", normalized):
                return alias
            if normalized.startswith(alias):
                rest = self._normalize_order_text(normalized[len(alias) :].lstrip("：: "))
                if self._looks_like_product_body(rest):
                    return alias
        english_head = self._extract_english_brand_head(normalized)
        if english_head:
            return english_head
        inferred = self._infer_brand_from_leading_phrase(normalized)
        if inferred:
            if fallback_token and inferred != fallback_token:
                inferred_key = inferred.lower()
                is_explicit_known = inferred_key in alias_keys or bool(re.fullmatch(r"[A-Za-z][A-Za-z-]{2,24}", inferred))
                if not is_explicit_known:
                    # In a message-level brand context, generic Chinese leading tokens are
                    # more often product names (e.g. 秦皮乙素) than a safe brand override.
                    return fallback_token
            return inferred
        return fallback_token

    def _looks_like_brand_token(self, text: str) -> bool:
        token = self._normalize_order_text(text).strip("：:")
        if not token:
            return False
        if len(token) < 2 or len(token) > 12:
            return False
        if token in {"普通", "普通型"}:
            return False
        upper_token = token.upper().replace(" ", "")
        if upper_token in NON_BRAND_STOPWORDS:
            return False
        if re.search(r"\d", token):
            return False
        if re.search(r"(试剂盒|试剂|封片剂|培养基|培养液|缓冲液|细胞|蛋白|抗体|离心管|枪头|吸头|滤器|滤膜|培养皿|孔板|培养瓶|培养板|血清|会员|打印纸)", token):
            return False
        if re.search(r"(酸|醇|酮|酶|胺|碱|盐|钠|钾|钙|镁|锌|铜|铁|水|液|剂|粉|盒|管|头|膜|皿|瓶)$", token):
            return False
        if re.search(r"(老师|收货|订|代付|下单|买|元|货号|型号|规格)", token):
            return False
        return bool(re.fullmatch(r"[\u4e00-\u9fa5A-Za-z]+", token))

    def _looks_like_product_body(self, text: str) -> bool:
        body = self._normalize_order_text(text)
        if not body:
            return False
        if PRODUCT_BRAND_CONTEXT_HINT_RE.search(body):
            return True
        if QTY_UNIT_RE.search(body) or QTY_CN_UNIT_RE.search(body):
            return True
        if PRICE_RE.search(body) or self._extract_formula_total(body):
            return True
        return False

    def _extract_english_brand_head(self, text: str) -> str:
        normalized = self._normalize_order_text(text)
        if not normalized:
            return ""
        match = re.match(r"^(?P<brand>[A-Za-z][A-Za-z\-]{2,24})\s+(?P<body>.+)$", normalized)
        if not match:
            return ""
        brand = self._normalize_order_text(str(match.group("brand") or ""))
        body = self._normalize_order_text(str(match.group("body") or ""))
        if not brand or not body:
            return ""
        if not self._looks_like_brand_token(brand):
            return ""
        if re.match(r"^(?:订|下单|代付|买|各订|补订|已订|再订)\s*", body):
            return ""
        if re.search(rf"{re.escape(brand)}\s*老师", normalized):
            return ""
        if not (
            self._looks_like_product_body(body)
            or bool(re.search(r"[A-Za-z]{1,8}[-_]?[A-Za-z0-9]{2,24}", body))
            or bool(SPEC_MARKER_RE.search(body))
            or bool(FORMULATION_SPEC_RE.search(body))
        ):
            return ""
        return brand

    def _infer_brand_from_leading_phrase(self, text: str) -> str:
        normalized = self._normalize_order_text(text)
        if not normalized:
            return ""
        alias_keys = {self._normalize_order_text(item).lower() for item in (self._active_brand_aliases or list(DEFAULT_LAB_BRAND_ALIASES))}
        clauses = [self._normalize_order_text(item) for item in re.split(r"[\r\n]+", normalized) if self._normalize_order_text(item)]
        for clause in clauses[:4]:
            match = re.match(r"^(?P<brand>[\u4e00-\u9fa5A-Za-z]{2,10})\s*(?:[:：])?\s+(?P<body>.+)$", clause)
            if not match:
                continue
            brand = self._normalize_order_text(str(match.group("brand") or ""))
            body = self._normalize_order_text(str(match.group("body") or ""))
            if not self._looks_like_brand_token(brand):
                continue
            if not self._looks_like_product_body(body):
                continue
            if re.match(r"^(?:订|下单|代付|买|各订|补订|已订|再订)\s*", body):
                continue
            if brand.lower() not in alias_keys:
                if re.fullmatch(r"[A-Za-z]{2,18}", brand):
                    if self._extract_english_brand_head(f"{brand} {body}") != brand:
                        continue
                else:
                    has_model_like = bool(re.search(r"[A-Za-z]{1,8}[-_]?[A-Za-z0-9]{1,16}", body))
                    has_lab_hint = bool(PRODUCT_BRAND_CONTEXT_HINT_RE.search(body))
                    if not has_model_like and not has_lab_hint:
                        continue
            if re.search(rf"{re.escape(brand)}\s*老师", clause):
                continue
            return brand
        return ""

    def _brand_token_looks_like_person_in_source(self, brand: str, source_text: str) -> bool:
        token = self._normalize_order_text(brand)
        source = self._normalize_order_text(source_text)
        if not token or not source:
            return False
        if re.search(rf"{re.escape(token)}\s*老师", source):
            return True
        if re.search(rf"(?:收货人|联系人|老师)[：:\s]*{re.escape(token)}", source):
            return True
        return False

    def _prefix_brand_to_product_name(self, product_name: str, brand: str, *, source_text: str) -> str:
        product = self._normalize_order_text(product_name)
        brand_token = self._normalize_order_text(brand)
        source = self._normalize_order_text(source_text)
        if not product or not brand_token:
            return product_name
        if brand_token in product:
            return product_name
        if source and self._brand_token_looks_like_person_in_source(brand_token, source):
            return product_name
        return self._normalize_order_text(f"{brand_token} {product}")

    def _extract_message_brand_context(self, content: str) -> str:
        lines = [self._normalize_order_text(line) for line in re.split(r"[\r\n]+", str(content or "")) if self._normalize_order_text(line)]
        if not lines:
            return ""
        alias_keys = {self._normalize_order_text(item).lower() for item in (self._active_brand_aliases or list(DEFAULT_LAB_BRAND_ALIASES))}
        # Priority-0: first line as a standalone brand with product-price lines following.
        first = lines[0].strip("：:")
        has_follow_price_line = any(bool(PRICE_RE.search(line) or self._extract_formula_total(line)) for line in lines[1:4])
        if has_follow_price_line and self._looks_like_brand_token(first):
            return first
        # Priority-1: explicit alias / leading-brand pattern in lines.
        for line in lines[:6]:
            head = self._normalize_order_text(re.split(r"[，,；;：:\s]", line, maxsplit=1)[0])
            alias = self._extract_brand_name(line) or self._extract_brand_name(head)
            if alias:
                return alias
            if self._looks_like_brand_token(head):
                rest = self._normalize_order_text(line[len(head):].lstrip("：: "))
                if (
                    self._looks_like_product_body(rest)
                    and not re.match(r"^(?:订|下单|代付|买|各订|补订|已订|再订)\s*", rest)
                ):
                    if head.lower() not in alias_keys:
                        if re.fullmatch(r"[A-Za-z]{2,18}", head):
                            if self._extract_english_brand_head(f"{head} {rest}") != head:
                                continue
                        else:
                            has_model_like = bool(re.search(r"[A-Za-z]{1,8}[-_]?[A-Za-z0-9]{1,16}", rest))
                            has_lab_hint = bool(PRODUCT_BRAND_CONTEXT_HINT_RE.search(rest))
                            if not has_model_like and not has_lab_hint:
                                continue
                    return head
        return ""

    def _extract_brand_mentions_with_line(self, content: str) -> tuple[list[str], list[dict[str, Any]]]:
        lines = [self._normalize_order_text(line) for line in re.split(r"[\r\n]+", str(content or "")) if self._normalize_order_text(line)]
        mentions: list[dict[str, Any]] = []
        alias_keys = {self._normalize_order_text(item).lower() for item in (self._active_brand_aliases or list(DEFAULT_LAB_BRAND_ALIASES))}
        message_brand_context = self._extract_message_brand_context(content)

        def suppress_generic_context_shadow(brand_value: str, source_kind_value: str, line_index: int) -> bool:
            brand_token = self._normalize_order_text(brand_value)
            if not brand_token or not message_brand_context or line_index <= 0:
                return False
            if brand_token == message_brand_context:
                return False
            if brand_token.lower() in alias_keys:
                return False
            if re.fullmatch(r"[A-Za-z][A-Za-z-]{2,24}", brand_token):
                return False
            return source_kind_value in {"line_extract", "head_line"}

        for idx, line in enumerate(lines):
            source_kind = ""
            brand = self._extract_brand_name(line)
            if brand:
                source_kind = "line_extract"
            if not brand:
                head = self._normalize_order_text(re.split(r"[，,；;：:\s]", line, maxsplit=1)[0])
                rest = self._normalize_order_text(line[len(head):].lstrip("：: ")) if head else ""
                candidate_head = ""
                if (
                    self._looks_like_brand_token(head)
                    and self._looks_like_product_body(rest)
                    and not re.match(r"^(?:订|下单|代付|买|各订|补订|已订|再订)\s*", rest)
                ):
                    if head.lower() not in alias_keys:
                        if re.fullmatch(r"[A-Za-z]{2,18}", head):
                            if self._extract_english_brand_head(f"{head} {rest}") != head:
                                head = ""
                        else:
                            has_model_like = bool(re.search(r"[A-Za-z]{1,8}[-_]?[A-Za-z0-9]{1,16}", rest))
                            has_lab_hint = bool(PRODUCT_BRAND_CONTEXT_HINT_RE.search(rest))
                            if not has_model_like and not has_lab_hint:
                                head = ""
                    if head:
                        candidate_head = head
                if candidate_head:
                    brand = candidate_head
                    source_kind = "head_line"
            if not brand and self._looks_like_brand_token(line.strip("：:")):
                lookahead = lines[idx + 1 : min(len(lines), idx + 4)]
                if any(self._looks_like_product_body(item) for item in lookahead):
                    brand = line.strip("：:")
                    source_kind = "standalone_line"
            if not brand:
                tail_match = re.search(r"(?P<brand>[\u4e00-\u9fa5A-Za-z]{2,18})\s*(?:订|下单|代付|买)\s*[一二两俩三四五六七八九十百半\d]", line)
                if tail_match:
                    brand_start = int(tail_match.start("brand"))
                    if brand_start > 0 and str(line[:brand_start]).strip():
                        tail_brand = self._normalize_order_text(str(tail_match.group("brand") or ""))
                        if tail_brand and self._looks_like_brand_token(tail_brand):
                            if tail_brand.lower() in alias_keys or bool(re.search(r"[A-Za-z]", tail_brand)):
                                brand = tail_brand
                                source_kind = "tail_hint"
            brand = self._normalize_order_text(brand)
            if not brand:
                continue
            if suppress_generic_context_shadow(brand, source_kind, idx):
                continue
            if self._brand_token_looks_like_person_in_source(brand, line):
                continue
            strong_anchor = bool(
                source_kind in {"line_extract", "head_line", "standalone_line"}
                and (
                    brand.lower() in alias_keys
                    or bool(re.search(rf"(^|[\s，,；;:：/]){re.escape(brand)}($|[\s，,；;:：/])", line))
                )
            )
            mentions.append(
                {
                    "brand": brand,
                    "line_index": idx,
                    "used": False,
                    "source_kind": source_kind,
                    "strong_anchor": strong_anchor,
                }
            )
        return lines, mentions

    def _resolve_row_line_index(
        self,
        row: dict[str, Any],
        *,
        lines: list[str],
        fallback_index: int,
    ) -> int:
        if not lines:
            return 0
        product = self._normalize_order_text(str(row.get("product_name") or ""))
        remark = self._normalize_order_text(str(row.get("remark") or ""))
        evidence = self._normalize_order_text(str(row.get("evidence_text") or ""))
        candidates = [product, remark, evidence]
        for candidate in candidates:
            if not candidate:
                continue
            for idx, line in enumerate(lines):
                if candidate and (candidate in line or line in candidate):
                    return idx
            fragments = [frag for frag in re.split(r"[\s，,；;：:()（）]+", candidate) if len(frag) >= 4]
            for frag in fragments[:6]:
                for idx, line in enumerate(lines):
                    if frag in line:
                        return idx
        return max(0, min(fallback_index, len(lines) - 1))

    def _consume_brand_mention(
        self,
        mentions: list[dict[str, Any]],
        *,
        line_index: int,
        expect_brand: str = "",
        max_context_back: int = 3,
    ) -> tuple[str, dict[str, Any] | None]:
        if not mentions:
            return "", None
        expected = self._normalize_order_text(expect_brand)
        for distance in range(0, max(0, int(max_context_back)) + 1):
            target_line = line_index - distance
            for item in mentions:
                if bool(item.get("used")):
                    continue
                if int(item.get("line_index", -9999)) != target_line:
                    continue
                brand = self._normalize_order_text(str(item.get("brand") or ""))
                if not brand:
                    continue
                if expected and brand != expected:
                    continue
                item["used"] = True
                return brand, item
        return "", None

    def _apply_brand_context_to_rows(self, rows: list[dict[str, Any]], *, source_text: str) -> list[dict[str, Any]]:
        if not rows:
            return rows
        lines, mentions = self._extract_brand_mentions_with_line(source_text)
        alias_keys = {self._normalize_order_text(item).lower() for item in (self._active_brand_aliases or list(DEFAULT_LAB_BRAND_ALIASES))}
        normalized_rows: list[dict[str, Any]] = []
        last_brand_anchor: dict[str, Any] | None = None
        max_inherit_rows = 2
        for row_index, row in enumerate(rows):
            if not isinstance(row, dict):
                continue
            item = dict(row)
            product_text = self._normalize_order_text(str(item.get("product_name") or ""))
            remark_text = self._normalize_order_text(str(item.get("remark") or ""))
            evidence_text = self._normalize_order_text(str(item.get("evidence_text") or ""))
            scope_text = self._normalize_order_text(str(item.get("source_scope_text") or ""))
            explicit_brand = self._extract_brand_name(f"{product_text} {remark_text}")
            row_brand = self._normalize_order_text(str(item.get("brand") or ""))
            leading_brand = self._infer_brand_from_leading_phrase(product_text) or self._infer_brand_from_leading_phrase(remark_text)
            line_index = self._resolve_row_line_index(item, lines=lines, fallback_index=row_index)
            resolved_brand = ""
            mention_meta: dict[str, Any] | None = None
            resolved_source = ""
            if explicit_brand:
                resolved_brand, mention_meta = self._consume_brand_mention(
                    mentions,
                    line_index=line_index,
                    expect_brand=explicit_brand,
                    max_context_back=3,
                )
                if resolved_brand:
                    resolved_source = "explicit_match_mention"
                if (
                    not resolved_brand
                    and explicit_brand.lower() in alias_keys
                    and explicit_brand in f"{product_text} {remark_text}"
                ):
                    # Brand appears in this row text but mention anchor may be missed by parser;
                    # keep explicit known-alias brand as a safe fallback.
                    resolved_brand = explicit_brand
                    resolved_source = "explicit_alias_fallback"
            if not resolved_brand and row_brand:
                resolved_brand, mention_meta = self._consume_brand_mention(
                    mentions,
                    line_index=line_index,
                    expect_brand=row_brand,
                    max_context_back=3,
                )
                if resolved_brand:
                    resolved_source = "row_brand_match_mention"
            if not resolved_brand and leading_brand:
                resolved_brand, mention_meta = self._consume_brand_mention(
                    mentions,
                    line_index=line_index,
                    expect_brand=leading_brand,
                    max_context_back=3,
                )
                if resolved_brand:
                    resolved_source = "leading_brand_match_mention"
            if not resolved_brand:
                resolved_brand, mention_meta = self._consume_brand_mention(
                    mentions,
                    line_index=line_index,
                    expect_brand="",
                    max_context_back=3,
                )
                if resolved_brand:
                    resolved_source = "nearby_mention_fallback"
            if not resolved_brand and last_brand_anchor and bool(last_brand_anchor.get("strong_anchor")):
                previous_row_index = int(last_brand_anchor.get("row_index", -999))
                previous_line_index = int(last_brand_anchor.get("line_index", -999))
                previous_depth = int(last_brand_anchor.get("inherit_depth", 0))
                previous_brand = self._normalize_order_text(str(last_brand_anchor.get("brand") or ""))
                if (
                    previous_brand
                    and row_index == previous_row_index + 1
                    and line_index >= previous_line_index
                    and (line_index - previous_line_index) <= 2
                    and previous_depth < max_inherit_rows
                    and not BRAND_INHERIT_BLOCK_RE.search(f"{product_text} {remark_text} {evidence_text}")
                ):
                    row_local_brand = self._extract_brand_name(f"{product_text} {remark_text} {evidence_text}")
                    if not row_local_brand or row_local_brand == previous_brand:
                        resolved_brand = previous_brand
                        resolved_source = "inherit_from_previous_strong"
            if not resolved_brand:
                local_fallback = (
                    row_brand
                    or explicit_brand
                    or self._extract_brand_name(f"{scope_text} {evidence_text} {remark_text}")
                )
                if local_fallback and not self._brand_token_looks_like_person_in_source(local_fallback, source_text):
                    resolved_brand = local_fallback
                    resolved_source = "row_local_fallback"
            resolved_brand = self._normalize_order_text(resolved_brand)
            if resolved_brand:
                item["brand"] = resolved_brand
                item["product_name"] = self._prefix_brand_to_product_name(
                    str(item.get("product_name") or ""),
                    resolved_brand,
                    source_text=source_text,
                )
                inherited_depth = 0
                strong_anchor = bool((mention_meta or {}).get("strong_anchor"))
                if resolved_source == "inherit_from_previous_strong" and last_brand_anchor:
                    strong_anchor = True
                    inherited_depth = int(last_brand_anchor.get("inherit_depth", 0)) + 1
                elif resolved_source == "explicit_alias_fallback":
                    strong_anchor = True
                    inherited_depth = 0
                last_brand_anchor = {
                    "brand": resolved_brand,
                    "line_index": line_index,
                    "row_index": row_index,
                    "strong_anchor": strong_anchor,
                    "inherit_depth": inherited_depth,
                }
            else:
                item["brand"] = ""
                last_brand_anchor = None
            normalized_rows.append(item)
        return normalized_rows

    def _llm_infer_brands_for_rows(
        self,
        *,
        source_text: str,
        rows: list[dict[str, Any]],
        min_confidence: float = 0.62,
    ) -> list[dict[str, Any]]:
        if not rows:
            return rows
        candidates: list[dict[str, Any]] = []
        for index, row in enumerate(rows):
            if not isinstance(row, dict):
                continue
            if str(row.get("brand") or "").strip():
                continue
            product_name = self._normalize_order_text(str(row.get("product_name") or ""))
            remark = self._normalize_order_text(str(row.get("remark") or ""))
            if not (product_name or remark):
                continue
            candidates.append(
                {
                    "index": index,
                    "product_name": product_name,
                    "remark": remark,
                }
            )
        if not candidates:
            return rows
        prompt = {
            "task": "从微信订货原文中，为每个候选行推断品牌 brand。要举一反三识别未见过的新品牌词（例如品牌常出现在句首短词）。",
            "source_text": self._normalize_order_text(source_text),
            "rows": candidates,
            "rules": [
                "仅依据 source_text，不可虚构。",
                "品牌通常是产品前的短词，可跨行继承；若后续未出现新品牌，允许继承前一个品牌。",
                "人名/老师名/收货人不是品牌。",
                "如果无法判断，brand 返回空字符串。",
            ],
            "response_contract": {
                "rows": [
                    {
                        "index": "number",
                        "brand": "string",
                        "confidence": "0-1",
                        "reason": "string",
                    }
                ],
            },
        }
        payload: dict[str, Any] = {}
        try:
            raw_payload = self._call_deepseek_json_cached(prompt, namespace="brand_inference")
            if isinstance(raw_payload, dict):
                payload = raw_payload
        except Exception:
            return rows
        inferred_rows = payload.get("rows")
        if not isinstance(inferred_rows, list):
            return rows
        updated = [dict(item) if isinstance(item, dict) else item for item in rows]
        for entry in inferred_rows:
            if not isinstance(entry, dict):
                continue
            try:
                idx = int(entry.get("index"))
            except (TypeError, ValueError):
                continue
            if idx < 0 or idx >= len(updated):
                continue
            row = updated[idx] if isinstance(updated[idx], dict) else {}
            if not isinstance(row, dict):
                continue
            if str(row.get("brand") or "").strip():
                continue
            brand = self._normalize_order_text(str(entry.get("brand") or ""))
            conf = self._coerce_confidence(entry.get("confidence"))
            if not brand or conf < float(min_confidence):
                continue
            if not self._looks_like_brand_token(brand):
                continue
            if self._brand_token_looks_like_person_in_source(brand, source_text):
                continue
            row["brand"] = brand
            row["product_name"] = self._prefix_brand_to_product_name(
                str(row.get("product_name") or ""),
                brand,
                source_text=source_text,
            )
            reason = self._normalize_order_text(str(entry.get("reason") or ""))
            if reason:
                remark = str(row.get("remark") or "")
                note = f"LLM品牌推断: {brand}"
                if note not in remark:
                    row["remark"] = (remark + f" | {note}").strip(" |")
            updated[idx] = row
        return self._apply_brand_context_to_rows(updated, source_text=source_text)

    def _rule_extract_rows_from_message(self, message: dict[str, Any], *, date_output_mode: str = "YYYY-MM-DD") -> list[dict[str, Any]]:
        content = str(message.get("content") or "")
        if not content.strip():
            return []
        name, owner = self._extract_name_owner(content)
        message_brand = self._extract_message_brand_context(content)
        date_code = self._extract_date_code(str(message.get("message_time") or message.get("observed_at") or ""))
        deterministic_segments = self._split_message_by_price_terminator(content, max_segments=12)
        normalized_content = self._normalize_order_text(content)
        preseg_mode = len(deterministic_segments) >= 2 or (
            len(deterministic_segments) == 1
            and ("\n" in content or "\r" in content)
            and deterministic_segments[0] != normalized_content
        )
        if preseg_mode:
            lines = [self._normalize_order_text(item) for item in deterministic_segments if self._normalize_order_text(item)]
        else:
            lines = [line.strip() for line in re.split(r"[\r\n]+", content) if line.strip()]
            if not lines:
                lines = [content.strip()]
        strong_rows: list[dict[str, Any]] = []
        weak_rows: list[dict[str, Any]] = []
        for line in lines:
            segments = [self._normalize_order_text(line)] if preseg_mode else self._split_order_segments(line)
            for segment in segments:
                if not self._line_has_order_signal(segment):
                    continue
                parsed = self._parse_line_item(segment, brand_context=message_brand)
                if not parsed:
                    continue
                if (
                    not str(parsed.get("product_name") or "").strip()
                    and not str(parsed.get("sale_price") or "").strip()
                    and not str(parsed.get("total_sale") or "").strip()
                ):
                    continue
                row = self._empty_order_row()
                row["date"] = date_code
                row["name"] = name
                row["owner"] = owner
                row["receiver"] = parsed.get("receiver") or ""
                row["order_unit"] = parsed.get("order_unit") or parsed.get("unit") or ""
                row["product_name"] = parsed.get("product_name") or ""
                row["quantity"] = parsed.get("quantity") or ""
                row["unit"] = parsed.get("unit") or ""
                row["brand"] = parsed.get("brand") or message_brand
                row["sale_price"] = parsed.get("sale_price") or ""
                row["total_sale"] = parsed.get("total_sale") or ""
                row["remark"] = parsed.get("remark") or ""
                row["source_scope_text"] = self._normalize_order_text(segment)
                row["confidence"] = self._estimate_confidence(row)
                row["needs_review"] = bool(row["confidence"] < 0.75 or not row["product_name"] or not row["quantity"])
                row["evidence_message_ids"] = [str(message.get("raw_message_id") or "")]
                finalized = self._finalize_order_row(row, message, date_output_mode=date_output_mode)
                if self._should_drop_order_row(finalized):
                    continue
                if str(finalized.get("product_name") or "").strip():
                    strong_rows.append(finalized)
                else:
                    weak_rows.append(finalized)
        if strong_rows:
            return self._apply_brand_context_to_rows(strong_rows, source_text=content)
        multiline = self._parse_multiline_message_item(content, brand_context=message_brand)
        if multiline and str(multiline.get("product_name") or "").strip():
            row = self._empty_order_row()
            row["date"] = date_code
            row["name"] = name
            row["owner"] = owner
            row["product_name"] = multiline.get("product_name") or ""
            row["quantity"] = multiline.get("quantity") or ""
            row["unit"] = multiline.get("unit") or ""
            row["brand"] = multiline.get("brand") or message_brand
            row["sale_price"] = multiline.get("sale_price") or ""
            row["total_sale"] = multiline.get("total_sale") or ""
            row["order_unit"] = multiline.get("order_unit") or multiline.get("unit") or ""
            row["remark"] = multiline.get("remark") or ""
            row["source_scope_text"] = self._normalize_order_text(str(multiline.get("remark") or content))
            row["confidence"] = self._estimate_confidence(row)
            row["needs_review"] = bool(row["confidence"] < 0.75 or not row["product_name"])
            row["evidence_message_ids"] = [str(message.get("raw_message_id") or "")]
            finalized = self._finalize_order_row(row, message, date_output_mode=date_output_mode)
            if not self._should_drop_order_row(finalized):
                return self._apply_brand_context_to_rows([finalized], source_text=content)
        fallback = self._parse_line_item(content, brand_context=message_brand)
        if not fallback:
            return self._apply_brand_context_to_rows(weak_rows[:1], source_text=content) if weak_rows else []
        row = self._empty_order_row()
        row["date"] = date_code
        row["name"] = name
        row["owner"] = owner
        row["product_name"] = fallback.get("product_name") or ""
        row["quantity"] = fallback.get("quantity") or ""
        row["unit"] = fallback.get("unit") or ""
        row["brand"] = fallback.get("brand") or message_brand
        row["sale_price"] = fallback.get("sale_price") or ""
        row["total_sale"] = fallback.get("total_sale") or ""
        row["remark"] = fallback.get("remark") or ""
        row["source_scope_text"] = self._normalize_order_text(content)
        row["confidence"] = self._estimate_confidence(row)
        row["needs_review"] = bool(row["confidence"] < 0.75 or not row["product_name"])
        row["evidence_message_ids"] = [str(message.get("raw_message_id") or "")]
        if not str(row.get("product_name") or "").strip():
            return self._apply_brand_context_to_rows(weak_rows[:1], source_text=content) if weak_rows else []
        finalized = self._finalize_order_row(row, message, date_output_mode=date_output_mode)
        if self._should_drop_order_row(finalized):
            return []
        return self._apply_brand_context_to_rows([finalized], source_text=content)

    def _split_order_segments(self, line: str) -> list[str]:
        text = self._normalize_order_text(line)
        if not text:
            return []
        yuan_segments = self._split_text_by_yuan_terminator(text)
        if len(yuan_segments) >= 2:
            return yuan_segments
        marker_count = len(re.findall(r"(?:各订|订|下单|代付|买)\s*[一二两俩三四五六七八九十百半\d]", text))
        if marker_count <= 1:
            return [text]
        parts = [part.strip() for part in re.split(r"[，,；;、]+", text) if part.strip()]
        if len(parts) <= 1:
            return [text]
        segments: list[str] = []
        current = ""
        for part in parts:
            if not current:
                current = part
                continue
            part_has_marker = bool(re.search(r"(?:各订|订|下单|代付|买)\s*[一二两俩三四五六七八九十百半\d]", part))
            current_has_price = bool(PRICE_RE.search(current) or self._extract_formula_total(current))
            if part_has_marker and current_has_price:
                segments.append(current)
                current = part
            else:
                current = self._normalize_order_text(f"{current} {part}")
        if current:
            segments.append(current)
        return [item for item in segments if item]

    def _looks_like_leading_spec_tail(self, text: str) -> bool:
        normalized = self._normalize_order_text(text)
        if not normalized:
            return False
        if PRICE_RE.search(normalized) or self._extract_formula_total(normalized):
            return False
        if re.search(r"(运费|快递费|合计|共计|满减|优惠|抹零|备注|赠品)", normalized):
            return False
        if re.search(r"(老师|收货人|联系人|发货人)", normalized):
            return False
        if re.search(r"^(?:货号|型号|规格|cat\s*no\.?|catno|no\.)\s*[:：]?", normalized, flags=re.IGNORECASE):
            return True
        if re.search(r"[\u4e00-\u9fa5]{2,}", normalized):
            return False
        if re.search(r"(?:订|下单|买|各订|代付|补订|再订)\s*[一二两俩三四五六七八九十百半\d]", normalized):
            return False
        spec = self._extract_product_spec(normalized)
        if spec:
            spec_key = self._normalize_spec_key(spec)
            normalized_key = self._normalize_spec_key(normalized)
            if spec_key and normalized_key and (spec_key == normalized_key or len(spec_key) >= max(4, int(len(normalized_key) * 0.6))):
                return True
        has_alnum_model = bool(re.search(r"(?:[A-Za-z].*\d|\d.*[A-Za-z])", normalized))
        has_numeric_code = bool(re.search(r"(?<!\d)\d{5,12}(?!\d)", normalized))
        if (has_alnum_model or has_numeric_code) and len(normalized) <= 56:
            return True
        return False

    def _find_next_product_start_index_after_yuan(self, text: str) -> int:
        normalized = self._normalize_order_text(text)
        if not normalized:
            return -1
        if not (PRICE_RE.search(normalized) or self._extract_formula_total(normalized)):
            return -1
        indexes: list[int] = []
        aliases = sorted((self._active_brand_aliases or list(DEFAULT_LAB_BRAND_ALIASES)), key=len, reverse=True)
        for alias in aliases:
            escaped = re.escape(self._normalize_order_text(alias))
            if not escaped:
                continue
            for match in re.finditer(escaped, normalized):
                idx = int(match.start())
                if idx <= 0:
                    continue
                prefix = self._normalize_order_text(normalized[:idx]).strip()
                suffix = self._normalize_order_text(normalized[idx:]).strip()
                if not prefix or not suffix:
                    continue
                if not (PRICE_RE.search(suffix) or self._extract_formula_total(suffix)):
                    continue
                if not self._looks_like_leading_spec_tail(prefix):
                    continue
                if not (self._extract_brand_name(suffix) or self._extract_product_name(suffix)):
                    continue
                indexes.append(idx)
        for separator in re.finditer(r"[，,；;、。\s]+", normalized):
            idx = int(separator.end())
            if idx <= 0 or idx >= len(normalized):
                continue
            prefix = self._normalize_order_text(normalized[:idx]).strip()
            suffix = self._normalize_order_text(normalized[idx:]).strip()
            if not prefix or not suffix:
                continue
            if not self._looks_like_leading_spec_tail(prefix):
                continue
            if not (PRICE_RE.search(suffix) or self._extract_formula_total(suffix)):
                continue
            if self._looks_like_cancel_or_status_line(suffix):
                continue
            suffix_brand = self._extract_brand_name(suffix)
            suffix_product = self._extract_product_name(suffix)
            if suffix_product and self._looks_like_non_product_line(suffix_product):
                suffix_product = ""
            if not suffix_brand and not suffix_product:
                continue
            if not self._line_has_order_signal(suffix) and not (suffix_product and PRICE_RE.search(suffix)):
                continue
            indexes.append(idx)
        if not indexes:
            return -1
        return min(indexes)

    def _split_price_part_with_start_anchor(self, part: str) -> tuple[str, str]:
        normalized = self._normalize_order_text(part)
        if not normalized:
            return "", ""
        if not (PRICE_RE.search(normalized) or self._extract_formula_total(normalized)):
            return "", normalized
        split_index = self._find_next_product_start_index_after_yuan(normalized)
        if split_index <= 0:
            return "", normalized
        prefix = self._normalize_order_text(normalized[:split_index]).strip()
        suffix = self._normalize_order_text(normalized[split_index:]).strip()
        if not prefix or not suffix:
            return "", normalized
        if not self._looks_like_leading_spec_tail(prefix):
            return "", normalized
        return prefix, suffix

    def _split_text_by_yuan_terminator(self, text: str, *, max_segments: int = 12) -> list[str]:
        normalized = self._normalize_order_text(text)
        if not normalized:
            return []
        if len(re.findall(r"元", normalized)) <= 1:
            return [normalized]
        raw_parts = [self._normalize_order_text(part) for part in re.split(r"(?<=元)", normalized) if self._normalize_order_text(part)]
        if len(raw_parts) <= 1:
            return [normalized]
        segments: list[str] = []
        buffer = ""
        for part in raw_parts:
            working_part = self._normalize_order_text(part)
            if segments and (PRICE_RE.search(working_part) or self._extract_formula_total(working_part)):
                carry_tail, anchored = self._split_price_part_with_start_anchor(working_part)
                if carry_tail:
                    segments[-1] = self._normalize_order_text(f"{segments[-1]} {carry_tail}")
                working_part = anchored or working_part
            candidate = self._normalize_order_text(f"{buffer} {working_part}") if buffer else working_part
            if PRICE_RE.search(working_part) or self._extract_formula_total(working_part):
                segments.append(candidate)
                buffer = ""
            else:
                buffer = candidate
        if buffer:
            if segments:
                segments[-1] = self._normalize_order_text(f"{segments[-1]} {buffer}")
            else:
                segments.append(buffer)
        cleaned = [item for item in segments if item]
        if not cleaned:
            return [normalized]
        return cleaned[:max_segments]

    def _parse_line_item(self, line: str, *, brand_context: str = "") -> dict[str, Any]:
        text = self._normalize_order_text(str(line or ""))
        if not text:
            return {}
        if self._looks_like_cancel_or_status_line(text):
            return {}
        if any(term in text for term in ("按表格订货", "照表格订货")):
            return {}
        quantity, unit = self._extract_quantity_unit(text)
        price_match = PRICE_RE.search(text)
        sale_price = price_match.group("price") if price_match else ""
        total_sale = ""
        formula_total = self._extract_formula_total(text)
        if formula_total:
            total_sale = self._format_number(self._to_number(formula_total))
        receiver = ""
        if "收货" in text and "老师" in text:
            receiver_match = re.search(r"收货[人者]?[：:\s]*(?P<receiver>[\u4e00-\u9fa5A-Za-z0-9]{1,20})老师", text)
            if receiver_match:
                receiver = str(receiver_match.group("receiver") or "")
        product_name = self._extract_product_name(text)
        brand_name = self._extract_brand_name(text, fallback=brand_context)
        spec_text = self._extract_product_spec(text, product_name=product_name, brand_hint=brand_name)
        qty_value = self._to_number(quantity)
        sale_value = self._to_number(sale_price)
        total_value = self._to_number(total_sale)
        if "共计" in text and sale_value > 0 and total_value <= 0:
            total_sale = self._format_number(sale_value)
            total_value = sale_value
            sale_price = ""
            sale_value = 0
        if unit in {"公斤", "斤"} and not re.search(r"(订|下单|买|各订|代付)\s*\d+(?:\.\d+)?\s*(公斤|斤)", text):
            quantity = ""
            unit = ""
            qty_value = 0
        if qty_value > 0 and total_value > 0:
            sale_price = self._format_number(total_value / qty_value)
        elif qty_value > 0 and sale_value > 0:
            total_sale = self._format_number(qty_value * sale_value)
        elif total_value > 0:
            sale_price = self._format_number(total_value)
        if not quantity and not unit and (total_value > 0 or sale_value > 0):
            if re.search(rf"(订|买|各订)\s*[一1]\s*({ORDER_UNIT_PATTERN})", text):
                quantity = "1"
                one_unit = re.search(rf"(订|买|各订)\s*[一1]\s*(?P<unit>{ORDER_UNIT_PATTERN})", text)
                unit = str(one_unit.group("unit") or "") if one_unit else ""
            elif re.search(r"(订|买|各订)\s*个", text):
                quantity = "1"
                unit = "个"
        if not product_name and not quantity and total_value <= 0 and sale_value <= 0:
            return {}
        return {
            "receiver": receiver,
            "order_unit": unit,
            "product_name": product_name,
            "brand": brand_name,
            "spec": spec_text,
            "quantity": self._format_number(self._to_number(quantity)) if quantity else "",
            "unit": unit,
            "sale_price": self._format_number(self._to_number(sale_price)) if sale_price else "",
            "total_sale": total_sale,
            "remark": text,
        }

    def _normalize_spec_text(self, value: str) -> str:
        spec = self._normalize_order_text(value)
        if not spec:
            return ""
        # 规格字段只保留型号特征，中文说明(如“样”“盒装”)在这里剔除。
        spec = re.sub(r"[\u4e00-\u9fa5]+", " ", spec)
        spec = re.sub(r"\s*/\s*", "/", spec)
        spec = re.sub(r"\s*-\s*", "-", spec)
        spec = re.sub(r"\s*\(\s*", "(", spec)
        spec = re.sub(r"\s*\)\s*", ")", spec)
        spec = spec.strip(" ,，;；:：()（）[]")
        spec = re.sub(r"\s{2,}", " ", spec).strip()
        # 排除手机号等明显非规格噪声。
        spec = re.sub(r"(?<!\d)1[3-9]\d{9}(?!\d)", " ", spec)
        spec = re.sub(r"\s{2,}", " ", spec).strip()
        return spec

    def _extract_primary_spec_code(self, text: str) -> str:
        candidate = self._normalize_spec_text(text)
        if not candidate:
            return ""
        for match in SPEC_MARKER_RE.finditer(candidate):
            token = self._normalize_spec_text(str(match.group("spec") or ""))
            if token and self._looks_like_spec_candidate(token, allow_numeric_only=True):
                return token
        for match in SPEC_CODE_TOKEN_RE.finditer(candidate):
            token = self._normalize_spec_text(str(match.group(0) or ""))
            if not token:
                continue
            token_upper = token.upper().replace(" ", "")
            if token_upper in NON_BRAND_STOPWORDS:
                continue
            if self._looks_like_spec_candidate(token, allow_numeric_only=True):
                return token
        return ""

    def _extract_spec_key(self, spec: str) -> str:
        normalized = self._normalize_spec_text(spec)
        if not normalized:
            return ""
        primary_code = self._extract_primary_spec_code(normalized)
        if primary_code:
            return self._normalize_spec_key(primary_code)
        code_text, size_text = self._split_spec_code_size(normalized, allow_numeric_code=True)
        if code_text:
            return self._normalize_spec_key(code_text)
        if size_text:
            return self._normalize_spec_key(size_text)
        return self._normalize_spec_key(normalized)

    def _normalize_spec_key(self, value: str) -> str:
        spec = self._normalize_spec_text(value)
        if not spec:
            return ""
        return re.sub(r"\s+", "", spec).lower()

    def _looks_like_spec_candidate(self, token: str, *, allow_numeric_only: bool = False) -> bool:
        candidate = self._normalize_spec_text(token)
        if not candidate:
            return False
        has_formulation = bool(FORMULATION_SPEC_RE.search(candidate))
        if has_formulation and self._extract_primary_spec_code(candidate):
            return True
        if re.search(r"[\u4e00-\u9fa5]", candidate):
            return False
        if len(candidate) < 2 or len(candidate) > 56:
            return False
        if " " in candidate:
            if not has_formulation:
                if not re.fullmatch(
                    r"[A-Za-z0-9._/\-*]{2,48}\s+(?:(?:\d+\s*[Tt]\s*/\s*\d+)|(?:\d+\s*[*xX×]\s*)?\d+(?:\.\d+)?\s*(?:μl|ul|ml|l|g|kg|mg|cm|mm|kda|da|bp|t))",
                    candidate,
                    flags=re.IGNORECASE,
                ):
                    return False
        if re.search(r"(老师|收货|订|代付|下单|买|元|合计|共计|现货|赠品)", candidate):
            return False
        if re.search(r"(?<!\d)1[3-9]\d{9}(?!\d)", candidate):
            return False
        upper = candidate.upper().replace(" ", "")
        if upper in NON_BRAND_STOPWORDS:
            return False
        if re.fullmatch(r"\d+(?:\.\d+)?", candidate):
            return allow_numeric_only and len(candidate) >= 4
        if re.fullmatch(r"\d+(?:\.\d+)?\s*(?:元|人民币)", candidate, flags=re.IGNORECASE):
            return False
        has_letter = bool(re.search(r"[A-Za-z]", candidate))
        digit_count = len(re.findall(r"\d", candidate))
        has_digit = digit_count > 0
        if SPEC_SIZE_RE.fullmatch(candidate):
            return True
        if has_letter and has_digit:
            if any(ch in candidate for ch in "-_/*"):
                return True
            if digit_count >= 2:
                return True
            if re.fullmatch(r"[A-Za-z]{1,6}\d[A-Za-z]?", candidate):
                return True
            return False
        if "/" in candidate:
            parts = [part for part in candidate.split("/") if part]
            if len(parts) >= 2 and all(self._looks_like_spec_candidate(part, allow_numeric_only=allow_numeric_only) for part in parts):
                return True
        if re.fullmatch(r"\d{4,12}", candidate):
            return allow_numeric_only
        if re.fullmatch(r"[A-Za-z]{1,4}\d{2,8}[A-Za-z0-9]{0,8}", candidate):
            return True
        return False

    def _split_spec_code_size(self, value: str, *, allow_numeric_code: bool = False) -> tuple[str, str]:
        candidate = self._normalize_spec_text(value)
        if not candidate:
            return "", ""
        candidate = re.sub(r"^(?:货号|型号|cat\s*no\.?|catno|no\.)\s*[:：]?\s*", "", candidate, flags=re.IGNORECASE).strip()
        if not candidate:
            return "", ""
        size_matches = list(SPEC_SIZE_RE.finditer(candidate))
        if not size_matches:
            if self._looks_like_spec_candidate(candidate, allow_numeric_only=allow_numeric_code) and not SPEC_SIZE_RE.fullmatch(candidate):
                return candidate, ""
            return "", ""
        best_match = max(
            size_matches,
            key=lambda item: (
                self._score_size_quality(str(item.group(0) or "")),
                len(str(item.group(0) or "")),
                item.start(),
            ),
        )
        size_text = self._normalize_spec_text(str(best_match.group(0) or ""))
        left = candidate[: best_match.start()].strip(" -_/")
        right = candidate[best_match.end() :].strip(" -_/")
        code_text = self._normalize_spec_text(f"{left} {right}")
        if code_text and not self._looks_like_spec_candidate(code_text, allow_numeric_only=True):
            code_text = ""
        return code_text, size_text

    def _compose_spec_value(self, code: str, size: str) -> str:
        normalized_code = self._normalize_spec_text(code)
        normalized_size = self._normalize_spec_text(size)
        if normalized_code and normalized_size:
            code_key = self._normalize_spec_key(normalized_code)
            size_key = self._normalize_spec_key(normalized_size)
            if size_key and size_key in code_key:
                return normalized_code
            return self._normalize_spec_text(f"{normalized_code} {normalized_size}")
        return normalized_code or normalized_size

    def _score_size_quality(self, size: str) -> int:
        normalized_size = self._normalize_spec_text(size).lower()
        if not normalized_size:
            return 0
        score = 0
        if re.search(r"\d+\s*t\s*/\s*\d+", normalized_size):
            score += 18
        if re.search(r"(?:μl|ul|ml|l|mg|g|kg)\b", normalized_size):
            score += 22
        if re.search(r"(?:kda|da|bp)\b", normalized_size):
            score += 6
        if re.search(r"[*xX×]", normalized_size):
            score += 10
        return score

    def _augment_spec_with_formulation(self, spec: str, *, scope_text: str) -> str:
        normalized_spec = self._normalize_spec_text(spec)
        scope = self._normalize_order_text(scope_text)
        if not normalized_spec or not scope:
            return normalized_spec
        formulation_match = FORMULATION_SPEC_RE.search(scope)
        if not formulation_match:
            return normalized_spec
        formulation = self._normalize_spec_text(str(formulation_match.group(0) or ""))
        if not formulation:
            return normalized_spec
        spec_key = self._normalize_spec_key(normalized_spec)
        formulation_key = self._normalize_spec_key(formulation)
        if formulation_key and formulation_key in spec_key:
            return self._dedupe_spec_formulation_size(normalized_spec)
        code_text = self._extract_primary_spec_code(normalized_spec)
        if not code_text:
            return normalized_spec
        formulation_size_match = SPEC_SIZE_RE.search(formulation)
        if formulation_size_match:
            formulation_size = self._normalize_spec_text(str(formulation_size_match.group(0) or ""))
            spec_sizes = list(SPEC_SIZE_RE.finditer(normalized_spec))
            if spec_sizes:
                tail_match = spec_sizes[-1]
                tail_size = self._normalize_spec_text(str(tail_match.group(0) or ""))
                if tail_size and self._normalize_spec_key(tail_size) == self._normalize_spec_key(formulation_size):
                    base = self._normalize_spec_text(normalized_spec[: tail_match.start()].strip())
                    if base and self._extract_primary_spec_code(base):
                        normalized_spec = base
        return self._dedupe_spec_formulation_size(f"{normalized_spec} {formulation}")

    def _dedupe_spec_formulation_size(self, spec: str) -> str:
        normalized_spec = self._normalize_spec_text(spec)
        if not normalized_spec:
            return ""
        formulation_match = FORMULATION_SPEC_RE.search(normalized_spec)
        if not formulation_match:
            return normalized_spec
        formulation = self._normalize_spec_text(str(formulation_match.group(0) or ""))
        formulation_size_matches = list(SPEC_SIZE_RE.finditer(formulation))
        if not formulation_size_matches:
            return normalized_spec
        formulation_size = self._normalize_spec_text(str(formulation_size_matches[-1].group(0) or ""))
        before = normalized_spec[: formulation_match.start()].strip(" -_/")
        after = normalized_spec[formulation_match.end() :].strip(" -_/")
        before_size_matches = list(SPEC_SIZE_RE.finditer(before))
        if before_size_matches and formulation_size:
            tail_match = before_size_matches[-1]
            tail_size = self._normalize_spec_text(str(tail_match.group(0) or ""))
            if tail_size and self._normalize_spec_key(tail_size) == self._normalize_spec_key(formulation_size):
                before = before[: tail_match.start()].strip(" -_/")
        return self._normalize_spec_text(" ".join(item for item in (before, formulation, after) if item))

    def _extract_product_spec(self, text: str, *, product_name: str = "", brand_hint: str = "") -> str:
        normalized_text = self._normalize_order_text(text)
        normalized_product = self._normalize_order_text(product_name)
        if not normalized_text and not normalized_product:
            return ""
        brand_token = self._normalize_order_text(brand_hint)
        scope = self._normalize_order_text(f"{normalized_product} {normalized_text}")
        if not scope:
            return ""
        cleaned_scope = re.sub(r"(?:\d+(?:\.\d+)?(?:\s*[*xX×/]\s*\d+(?:\.\d+)?)+)\s*=\s*\d+(?:\.\d+)?\s*元?", " ", scope)
        cleaned_scope = re.sub(r"\d+(?:\.\d+)?\s*元", " ", cleaned_scope)
        code_scores: dict[str, tuple[int, str]] = {}
        size_scores: dict[str, tuple[int, str]] = {}
        pair_scores: dict[str, tuple[int, str, str, str]] = {}

        def upsert_code(value: str, score: int) -> None:
            code = self._normalize_spec_text(value)
            if not code:
                return
            key = self._normalize_spec_key(code)
            if not key:
                return
            existing = code_scores.get(key)
            if existing is None or score > existing[0]:
                code_scores[key] = (score, code)

        def upsert_size(value: str, score: int) -> None:
            size = self._normalize_spec_text(value)
            if not size:
                return
            key = self._normalize_spec_key(size)
            if not key:
                return
            existing = size_scores.get(key)
            if existing is None or score > existing[0]:
                size_scores[key] = (score, size)

        def upsert_pair(code: str, size: str, score: int, *, display: str = "") -> None:
            composed = self._compose_spec_value(code, size)
            if not composed:
                return
            composed_code, composed_size = self._split_spec_code_size(composed, allow_numeric_code=True)
            if not composed_code and not composed_size:
                return
            display_text = self._normalize_spec_text(display) or composed
            key = f"{self._normalize_spec_key(composed_code)}|{self._normalize_spec_key(composed_size)}"
            existing = pair_scores.get(key)
            if existing is None or score > existing[0]:
                pair_scores[key] = (score, composed_code, composed_size, display_text)

        def add_candidate(value: str, *, explicit: bool = False) -> None:
            candidate = self._normalize_spec_text(value)
            if not candidate:
                return
            if brand_token and self._normalize_spec_key(candidate) == self._normalize_spec_key(brand_token):
                return
            code_text, size_text = self._split_spec_code_size(candidate, allow_numeric_code=explicit)
            base_score = 0
            if explicit:
                base_score += 120
            if code_text and size_text:
                base_score += 80
            elif code_text:
                base_score += 50
            elif size_text:
                base_score += 20
            if any(ch in candidate for ch in "-_/*"):
                base_score += 25
            digit_count = len(re.findall(r"\d", candidate))
            if digit_count >= 2:
                base_score += 16
            if re.search(r"[A-Za-z]", candidate):
                base_score += 12
            if len(candidate) > 36:
                base_score -= 8
            if len(candidate) <= 5 and digit_count <= 1:
                base_score -= 20
            code_allow_numeric = explicit or bool(size_text)
            if code_text and not self._looks_like_spec_candidate(code_text, allow_numeric_only=code_allow_numeric):
                code_text = ""
            if size_text and not self._looks_like_spec_candidate(size_text, allow_numeric_only=False):
                size_text = ""
            if code_text:
                upsert_code(code_text, base_score + 10)
            if size_text:
                upsert_size(size_text, base_score + 6 + self._score_size_quality(size_text))
            if code_text and size_text:
                upsert_pair(
                    code_text,
                    size_text,
                    base_score + 28 + self._score_size_quality(size_text),
                    display=candidate,
                )
            if not code_text and not size_text and self._looks_like_spec_candidate(candidate, allow_numeric_only=explicit):
                upsert_code(candidate, base_score)

        for marker_match in SPEC_MARKER_RE.finditer(scope):
            add_candidate(str(marker_match.group("spec") or ""), explicit=True)

        for token_match in SPEC_TOKEN_RE.finditer(cleaned_scope):
            token = str(token_match.group(0) or "")
            if not token:
                continue
            if len(token) > 36 and "/" not in token:
                continue
            add_candidate(token, explicit=False)

        for numeric_match in re.finditer(r"(?<!\d)\d{5,12}(?!\d)", cleaned_scope):
            numeric_code = self._normalize_spec_text(str(numeric_match.group(0) or ""))
            if not numeric_code:
                continue
            start = numeric_match.start()
            end = numeric_match.end()
            short_window = cleaned_scope[max(0, start - 6) : min(len(cleaned_scope), end + 6)]
            if re.search(r"(?:\.\d|[*xX×/]|=|元)", short_window):
                continue
            around = cleaned_scope[max(0, start - 24) : min(len(cleaned_scope), end + 24)]
            size_matches = [str(item.group(0) or "") for item in SPEC_SIZE_RE.finditer(around)]
            near_size_text = ""
            if size_matches:
                near_size_text = max(size_matches, key=lambda value: (self._score_size_quality(value), len(value)))
            near_spec_marker = bool(re.search(r"(货号|型号|编号|cat\s*no\.?|catno|no\.)", around, flags=re.IGNORECASE))
            if not near_size_text and not near_spec_marker:
                continue
            upsert_code(numeric_code, 48 if near_size_text else 42)
            if near_size_text:
                upsert_pair(
                    numeric_code,
                    near_size_text,
                    64 + self._score_size_quality(near_size_text),
                    display=f"{numeric_code} {near_size_text}",
                )

        for size_match in SPEC_SIZE_RE.finditer(scope):
            upsert_size(str(size_match.group(0) or ""), 24)

        best_pair = ""
        if pair_scores:
            sorted_pairs = sorted(pair_scores.values(), key=lambda item: (item[0], len(item[3])), reverse=True)
            pair_score, pair_code, pair_size, pair_display = sorted_pairs[0]
            if pair_score > 0:
                best_pair = self._normalize_spec_text(pair_display) or self._compose_spec_value(pair_code, pair_size)
        best_code = ""
        if code_scores:
            sorted_codes = sorted(code_scores.values(), key=lambda item: (item[0], len(item[1])), reverse=True)
            best_code = self._normalize_spec_text(sorted_codes[0][1])
        best_size = ""
        if size_scores:
            sorted_sizes = sorted(size_scores.values(), key=lambda item: (item[0], len(item[1])), reverse=True)
            best_size = self._normalize_spec_text(sorted_sizes[0][1])
        combined = best_pair or self._compose_spec_value(best_code, best_size)
        if combined:
            combined = self._augment_spec_with_formulation(combined, scope_text=scope)
        if combined and (not brand_token or self._normalize_spec_key(combined) != self._normalize_spec_key(brand_token)):
            return combined
        if best_code and (not brand_token or self._normalize_spec_key(best_code) != self._normalize_spec_key(brand_token)):
            return self._augment_spec_with_formulation(best_code, scope_text=scope)
        if best_size:
            return best_size
        return ""

    def _score_spec_candidate(self, spec: str, *, product_name: str = "", context_text: str = "") -> int:
        candidate = self._normalize_spec_text(spec)
        if not candidate:
            return 0
        has_formulation = bool(FORMULATION_SPEC_RE.search(candidate))
        code_hint = self._extract_primary_spec_code(candidate)
        if not self._looks_like_spec_candidate(candidate, allow_numeric_only=True):
            if not (has_formulation and code_hint):
                return 0
        product = self._normalize_order_text(product_name)
        context = self._normalize_order_text(context_text)
        score = 0
        code_text, size_text = self._split_spec_code_size(candidate, allow_numeric_code=True)
        if not code_text and code_hint:
            code_text = self._normalize_spec_text(code_hint)
        has_letter = bool(re.search(r"[A-Za-z]", candidate))
        digit_count = len(re.findall(r"\d", candidate))
        if code_text and size_text:
            score += 85
        elif code_text:
            score += 55
        elif size_text:
            score += 18
        if has_formulation:
            score += 14
        if code_text and re.search(r"[A-Za-z]", code_text) and re.search(r"\d", code_text):
            score += 24
        if any(ch in candidate for ch in "-_/*"):
            score += 20
        if SPEC_SIZE_RE.fullmatch(candidate):
            score += 6
        if digit_count >= 3:
            score += 12
        if product and candidate in product:
            score += 40
        if product and code_text and code_text in product:
            score += 36
        if product and size_text and size_text in product:
            score += 22
        if context and candidate in context:
            score += 15
        if context and code_text and code_text in context:
            score += 14
        if context and size_text and size_text in context:
            score += 8
        if len(candidate) > 26:
            score -= 8
        if not has_letter and not code_text and size_text:
            score -= 5
        if size_text and not code_text:
            product_key = self._normalize_order_text(product).lower()
            size_key = self._normalize_spec_key(size_text)
            if product_key and size_key and size_key not in product_key:
                score -= 40
                if len(self._normalize_order_text(product)) <= 8:
                    score -= 40
        return score

    def _extract_product_name(self, text: str) -> str:
        candidate = self._normalize_order_text(text)
        if not candidate:
            return ""
        if any(term in candidate for term in ("不下单", "取消", "退掉", "暂时不用")):
            return ""
        special_buy = re.search(r"代购(?:买)?\s*\d+(?:\.\d+)?\s*(?:公斤|斤|个|包|箱|瓶|套|件|份|块|台|支|株|卷|提|板)?\s*(?P<name>[\u4e00-\u9fa5A-Za-z0-9\-()（）_+./]+)", candidate)
        if special_buy and str(special_buy.group("name") or "").strip():
            fallback_name = self._normalize_order_text(str(special_buy.group("name") or ""))
            if fallback_name:
                candidate = fallback_name
        if "[引用" in candidate and "]" in candidate:
            candidate = self._normalize_order_text(re.sub(r"\[引用[^\]]*\]", " ", candidate))

        marker_match = re.search(r"(各订|订|下单|代付|买)\s*[一二两俩三四五六七八九十百半\d]", candidate)
        cut_index = marker_match.start() if marker_match else len(candidate)
        formula_match = re.search(r"(?:\d+(?:\.\d+)?(?:\s*[*xX×/]\s*\d+(?:\.\d+)?)+)\s*=\s*\d+(?:\.\d+)?\s*元?", candidate)
        if formula_match:
            cut_index = min(cut_index, formula_match.start())
        total_price_match = re.search(r"\d+(?:\.\d+)?\s*元", candidate)
        if total_price_match:
            cut_index = min(cut_index, total_price_match.start())
        candidate = candidate[:cut_index].strip() if cut_index > 0 else candidate

        candidate = re.sub(rf"\b\d+(?:\.\d+)?\s*(?:{ORDER_UNIT_PATTERN})\b", "", candidate)
        candidate = re.sub(rf"\b[一二两俩三四五六七八九十百半]\s*(?:{ORDER_UNIT_PATTERN})\b", "", candidate)
        candidate = re.sub(r"(已订|（已订）|\(已订\)|（已）|\(已\)|已送|代付|下单|代购|代购买)", "", candidate)
        candidate = re.sub(r"(合计|共计|满\d+元.*赠品.*|赠品选?)", "", candidate)
        candidate = re.sub(r"^这\s*\d+(?:\.\d+)?\s*个?抗体", "", candidate)
        candidate = re.sub(r"(老师|收货[人者]?)", "", candidate)
        candidate = re.sub(r"^[\u4e00-\u9fa5A-Za-z0-9]{1,20}的", "", candidate)
        candidate = re.sub(r"货号[：:]\s*[A-Za-z0-9_-]+", "", candidate, flags=re.IGNORECASE)
        candidate = re.sub(r"^(淘宝|京东|阿里巴巴|拼多多|美团|饿了么)\s*", "", candidate)
        candidate = re.sub(rf"^这\s*\d+(?:\.\d+)?\s*(?:{ORDER_UNIT_PATTERN})", "", candidate)
        candidate = re.sub(rf"^这\s*[一二两俩三四五六七八九十百半]\s*(?:{ORDER_UNIT_PATTERN})", "", candidate)
        candidate = re.sub(r"\s+", " ", candidate).strip(" ,，。;；:-[]")

        if re.match(r"^[（(].+[）)]$", candidate):
            return ""
        if re.match(r"^[\u4e00-\u9fa5]{2,4}\]?$", candidate) and "老师" in str(text or ""):
            # Avoid dropping short real product names (e.g. 深孔板/滤膜/灯管).
            if not re.search(r"(管|板|膜|液|剂|盒|瓶|皿|枪|吸头|培养|试|酸|钠|胶|粉|纸|架|刷|离心|灯|机|仪|泵|炉|箱|柜|秤|天平)", candidate) and not SHORT_OFFICE_PRODUCT_HINT_RE.search(candidate):
                return ""
        if re.fullmatch(r"\d+(?:\.\d+)?(?:\s*[*xX×/+.-]\s*\d+(?:\.\d+)?)+\s*(?:=\s*\d+(?:\.\d+)?)?\s*元?", candidate):
            return ""
        if re.fullmatch(r"\d+(?:\.\d+)?\s*元?", candidate):
            return ""
        if re.search(r"(按表格订货|照表格订货|\[文件\]|\.xlsx?$|\.xls$)", candidate, flags=re.IGNORECASE):
            return ""
        if candidate in {"淘宝", "京东", "阿里巴巴", "拼多多", "代付"}:
            return ""
        return candidate

    @staticmethod
    def _normalize_order_text(text: str) -> str:
        normalized = str(text or "").replace("\xa0", " ")
        normalized = VALIDATION_MARKER_RE.sub(" ", normalized)
        normalized = re.sub(r"\s+", " ", normalized)
        return normalized.strip()

    @staticmethod
    def _cn_to_number(value: str) -> float:
        text = str(value or "").strip()
        mapping = {
            "半": 0.5,
            "一": 1.0,
            "二": 2.0,
            "两": 2.0,
            "俩": 2.0,
            "三": 3.0,
            "四": 4.0,
            "五": 5.0,
            "六": 6.0,
            "七": 7.0,
            "八": 8.0,
            "九": 9.0,
            "十": 10.0,
        }
        if text in mapping:
            return mapping[text]
        if text == "十二":
            return 12.0
        if text == "十一":
            return 11.0
        if text == "十三":
            return 13.0
        return 0.0

    def _extract_quantity_unit(self, text: str) -> tuple[str, str]:
        qty_match = QTY_UNIT_RE.search(text)
        if qty_match:
            return str(qty_match.group("qty") or ""), str(qty_match.group("unit") or "")
        qty_cn = QTY_CN_UNIT_RE.search(text)
        if qty_cn:
            cn_value = self._cn_to_number(str(qty_cn.group("qty_cn") or ""))
            if cn_value > 0:
                return self._format_number(cn_value), str(qty_cn.group("unit") or "")
        return "", ""

    @staticmethod
    def _extract_formula_total(text: str) -> str:
        formula = re.search(r"(?:\d+(?:\.\d+)?(?:\s*[*xX×/]\s*\d+(?:\.\d+)?)+)\s*=\s*(?P<total>\d+(?:\.\d+)?)\s*元?", text)
        if formula:
            return str(formula.group("total") or "")
        simple = re.search(r"(?P<a>\d+(?:\.\d+)?)\s*[*xX×]\s*(?P<b>\d+(?:\.\d+)?)\s*=\s*(?P<c>\d+(?:\.\d+)?)\s*元?", text)
        if simple:
            return str(simple.group("c") or "")
        divide = re.search(r"(?P<a>\d+(?:\.\d+)?)\s*/\s*(?P<b>\d+(?:\.\d+)?)\s*=\s*(?P<c>\d+(?:\.\d+)?)\s*元?", text)
        if divide:
            return str(divide.group("c") or "")
        return ""

    def _parse_multiline_message_item(self, content: str, *, brand_context: str = "") -> dict[str, Any]:
        lines = [self._normalize_order_text(line) for line in re.split(r"[\r\n]+", str(content or "")) if self._normalize_order_text(line)]
        if len(lines) < 2:
            return {}
        signal_indexes = [idx for idx, line in enumerate(lines) if self._line_has_order_signal(line)]
        if not signal_indexes:
            return {}
        best: dict[str, Any] = {}
        best_score = -1.0
        for idx in signal_indexes:
            composed = self._compose_multiline_order_text(lines, idx)
            if not composed:
                continue
            parsed = self._parse_line_item(composed, brand_context=brand_context)
            if not parsed:
                continue
            score = 0.0
            if parsed.get("product_name"):
                score += 1.0
            if parsed.get("quantity"):
                score += 0.6
            if parsed.get("sale_price") or parsed.get("total_sale"):
                score += 0.4
            if score > best_score:
                best_score = score
                best = parsed
        return best

    def _compose_multiline_order_text(self, lines: list[str], index: int) -> str:
        selected: list[str] = []
        for prev in range(index - 1, -1, -1):
            line = lines[prev]
            if self._looks_like_non_product_line(line):
                continue
            if line in selected:
                continue
            selected.append(line)
            break
        selected.append(lines[index])
        for nxt in range(index + 1, min(len(lines), index + 3)):
            line = lines[nxt]
            if self._extract_formula_total(line) or PRICE_RE.search(line):
                selected.append(line)
                break
        return self._normalize_order_text(" ".join(selected))

    @staticmethod
    def _looks_like_non_product_line(line: str) -> bool:
        text = str(line or "").strip()
        if not text:
            return True
        if re.match(r"^[\u4e00-\u9fa5A-Za-z0-9]{2,24}老师$", text):
            return True
        if re.match(r"^(货号|批号|型号)[：:]", text):
            return True
        if re.match(r"^[A-Za-z0-9._-]{2,20}$", text):
            return True
        if re.match(r"^\d+(?:\.\d+)?\s*元$", text):
            return True
        return False

    def _extract_name_owner(self, text: str) -> tuple[str, str]:
        if self._message_contains_at_mention(text):
            return "", ""
        sanitized_text = self._strip_mentions_for_name(text)
        match = NAME_OWNER_RE.search(sanitized_text)
        if not match:
            single = re.search(r"(?P<name>(?:[\u4e00-\u9fa5A-Za-z0-9]\s*){1,24})老师", sanitized_text)
            if single:
                name = self._sanitize_person_name(re.sub(r"\s+", "", str(single.group("name") or "")))
                return name, name
            # Fallback: many orders end with a plain contact name line like "冯世浩".
            tail_name = re.search(r"(?:^|[\r\n\s])(?P<name>[\u4e00-\u9fa5]{2,4})\s*$", str(sanitized_text or ""))
            if tail_name and self._looks_like_order_message(str(sanitized_text or "")):
                name = self._sanitize_person_name(str(tail_name.group("name") or ""))
                if name and name not in {"老师", "同学", "客户"}:
                    return name, ""
            return "", ""
        name = self._sanitize_person_name(str(match.group("name") or ""))
        owner = self._sanitize_person_name(str(match.group("owner") or ""))
        return name, owner

    def _sanitize_person_name(self, value: str) -> str:
        text = self._normalize_order_text(str(value or ""))
        if not text:
            return ""
        text = re.sub(r"@[^\s，,。；;：:\r\n]+", " ", text)
        text = re.sub(r"\d+(?:\.\d+)?\s*元", " ", text)
        text = re.sub(r"\d+(?:\.\d+)?\s*(?:折|折后|\\/|\\*)", " ", text)
        text = re.sub(rf"\d+(?:\.\d+)?\s*(?:{ORDER_UNIT_PATTERN})", " ", text)
        text = re.sub(r"(?:订|代付|下单|补订|再订|买|各订|已订|货号|型号|规格|老师)$", " ", text)
        text = self._normalize_order_text(text)
        chinese_candidates = re.findall(r"[\u4e00-\u9fa5]{2,4}", text)
        stopwords = {
            "老师",
            "同学",
            "客户",
            "试剂",
            "溶液",
            "细胞",
            "培养",
            "离心",
            "蛋白",
            "抗体",
            "普通",
            "麦克林",
            "阿拉丁",
            "索莱宝",
            "碧云天",
            "康宁",
            "赛宁",
            "国药",
            "探索",
        }
        for candidate in reversed(chinese_candidates):
            if candidate and candidate not in stopwords and not any(marker in candidate for marker in ("试剂", "溶液", "细胞", "培养", "离心")):
                return candidate
        ascii_candidate = re.sub(r"[^A-Za-z]", "", text).strip()
        if 2 <= len(ascii_candidate) <= 16 and not re.search(r"(?:订|元|ml|mg|kg|g|sku|id)", ascii_candidate, flags=re.IGNORECASE):
            return ascii_candidate
        return ""

    @staticmethod
    def _message_contains_at_mention(text: str) -> bool:
        normalized = str(text or "")
        return "@" in normalized and bool(re.search(r"@[^\s，,。；;：:\r\n]+", normalized))

    def _strip_mentions_for_name(self, text: str) -> str:
        normalized = str(text or "")
        if "@" not in normalized:
            return normalized
        stripped = re.sub(r"@[^\s，,。；;：:\r\n]+", " ", normalized)
        stripped = stripped.replace("@所有人", " ")
        return re.sub(r"[ \t]+", " ", stripped)

    def _parse_message_datetime(self, message: dict[str, Any]) -> datetime | None:
        if not isinstance(message, dict):
            return None
        value = str(message.get("message_time") or message.get("observed_at") or "").strip()
        return self._parse_datetime_text(value)

    @staticmethod
    def _parse_datetime_text(value: str) -> datetime | None:
        text = str(value or "").strip()
        if not text:
            return None
        normalized = text.replace("/", "-").replace("T", " ")
        if normalized.endswith("Z"):
            normalized = normalized[:-1] + "+00:00"
        try:
            parsed = datetime.fromisoformat(normalized)
            if parsed.tzinfo is not None:
                parsed = parsed.astimezone(UTC8).replace(tzinfo=None)
            return parsed
        except ValueError:
            pass
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
            try:
                return datetime.strptime(normalized, fmt)
            except ValueError:
                continue
        return None

    def _extract_time_code(self, value: str) -> str:
        text = str(value or "").strip()
        if not re.search(r"\d{1,2}:\d{2}", text):
            return ""
        parsed = self._parse_datetime_text(text)
        if not parsed:
            return ""
        return parsed.strftime("%H:%M:%S")

    def _infer_name_owner_from_recent_messages(
        self,
        *,
        ordered_messages: list[dict[str, Any]],
        message_index: int,
        max_messages: int = 3,
        max_minutes: int = 5,
    ) -> tuple[str, str, bool]:
        if not ordered_messages or message_index <= 1:
            return "", "", False
        capped_max_messages = max(1, min(int(max_messages or 3), 3))
        capped_max_minutes = max(1, int(max_minutes or 5))
        current_pos = max(0, min(message_index - 1, len(ordered_messages) - 1))
        current_message = ordered_messages[current_pos] if current_pos < len(ordered_messages) else {}
        current_time = self._parse_message_datetime(current_message)
        candidates: list[dict[str, Any]] = []
        checked_count = 0
        for prior_pos in range(current_pos - 1, -1, -1):
            if checked_count >= capped_max_messages:
                break
            checked_count += 1
            prior_message = ordered_messages[prior_pos]
            prior_content = str(prior_message.get("content") or "")
            if not prior_content.strip():
                continue
            if self._message_contains_at_mention(prior_content):
                continue
            prior_time = self._parse_message_datetime(prior_message)
            if current_time and prior_time:
                delta_seconds = (current_time - prior_time).total_seconds()
                if delta_seconds < 0 or delta_seconds > capped_max_minutes * 60:
                    continue
            prior_name, prior_owner = self._extract_name_owner(prior_content)
            resolved_name = str(prior_name or "").strip()
            resolved_owner = str(prior_owner or resolved_name).strip()
            if not resolved_name and not resolved_owner:
                continue
            sanitized_prior = self._strip_mentions_for_name(prior_content)
            has_teacher_suffix = bool(re.search(r"[\u4e00-\u9fa5A-Za-z0-9]{1,24}老师", sanitized_prior))
            candidates.append(
                {
                    "name": resolved_name,
                    "owner": resolved_owner,
                    "teacher": has_teacher_suffix,
                }
            )
        if not candidates:
            return "", "", False
        selected = next((item for item in candidates if bool(item.get("teacher"))), candidates[0])
        name = str(selected.get("name") or "").strip()
        owner = str(selected.get("owner") or name).strip()
        if not name and not owner:
            return "", "", False
        if not name:
            name = owner
        if not owner:
            owner = name
        return name, owner, True

    def _apply_name_context_fallback_to_row(
        self,
        row: dict[str, Any],
        *,
        ordered_messages: list[dict[str, Any]],
        message_index: int,
        max_messages: int = 3,
        max_minutes: int = 5,
        confidence_penalty: float = 0.06,
    ) -> dict[str, Any]:
        output = dict(row)
        has_name = bool(str(output.get("name") or "").strip())
        has_owner = bool(str(output.get("owner") or "").strip())
        if has_name and has_owner:
            return output
        inferred_name, inferred_owner, inferred = self._infer_name_owner_from_recent_messages(
            ordered_messages=ordered_messages,
            message_index=message_index,
            max_messages=max_messages,
            max_minutes=max_minutes,
        )
        if not inferred:
            return output
        if not has_name and inferred_name:
            output["name"] = inferred_name
            has_name = True
        if not has_owner and (inferred_owner or inferred_name):
            output["owner"] = inferred_owner or inferred_name
            has_owner = True
        if not has_name and not has_owner:
            return output
        note = "姓名来自近3句上下文(<=5分钟)"
        remark = str(output.get("remark") or "")
        if note not in remark:
            output["remark"] = (remark + f" | {note}").strip(" |")
        penalty = max(0.0, min(float(confidence_penalty or 0.0), 0.3))
        output["confidence"] = self._coerce_confidence(self._coerce_confidence(output.get("confidence")) - penalty)
        output["needs_review"] = bool(
            output.get("needs_review")
            or float(output.get("confidence") or 0) < 0.75
            or not str(output.get("product_name") or "").strip()
            or not str(output.get("quantity") or "").strip()
        )
        return output

    def _extract_date_code(self, value: str) -> str:
        text = str(value or "")
        match = DATE_RE.search(text)
        if match:
            year = int(match.group("year") or 0)
            month = int(match.group("month") or 0)
            day = int(match.group("day") or 0)
            if year and month and day:
                return f"{year:04d}-{month:02d}-{day:02d}"
        match = MONTH_DAY_RE.search(text)
        if match:
            month = int(match.group("month") or 0)
            day = int(match.group("day") or 0)
            if month and day:
                inferred_year = datetime.now().year
                year_match = re.search(r"(?P<year>20\d{2})", text)
                if year_match:
                    inferred_year = int(year_match.group("year") or inferred_year)
                return f"{inferred_year:04d}-{month:02d}-{day:02d}"
        match = YEAR_MONTH_DAY_CN_RE.search(text)
        if match:
            year = int(match.group("year") or 0)
            month = int(match.group("month") or 0)
            day = int(match.group("day") or 0)
            if year and month and day:
                return f"{year:04d}-{month:02d}-{day:02d}"
        mmdd_match = MMDD_RE.match(text.strip())
        if mmdd_match:
            month = int(mmdd_match.group("month") or 0)
            day = int(mmdd_match.group("day") or 0)
            if month and day:
                inferred_year = datetime.now().year
                return f"{inferred_year:04d}-{month:02d}-{day:02d}"
        short_md = MONTH_DAY_SLASH_RE.search(text)
        if short_md:
            month = int(short_md.group("month") or 0)
            day = int(short_md.group("day") or 0)
            if month and day:
                inferred_year = datetime.now().year
                year_match = re.search(r"(?P<year>20\d{2})", text)
                if year_match:
                    inferred_year = int(year_match.group("year") or inferred_year)
                return f"{inferred_year:04d}-{month:02d}-{day:02d}"
        return ""

    def _looks_like_order_message(self, text: str) -> bool:
        normalized = str(text or "")
        if len(normalized) < 3:
            return False
        if self._looks_like_cancel_or_status_line(normalized):
            return False
        has_product_hint = bool(re.search(r"(试剂|离心管|细胞|抗体|蛋白|货号|型号|架|试纸|耗材|培养基|移液器|灯管|孔板|滤膜|培养皿)", normalized))
        if any(term in normalized for term in ("预存", "欠款", "还欠", "扣除", "补过来", "平台预存")) and not has_product_hint:
            return False
        lines = [line.strip() for line in re.split(r"[\r\n]+", normalized) if line.strip()]
        for line in lines:
            if self._line_has_order_signal(line):
                return True
        has_price = bool(PRICE_RE.search(normalized) or self._extract_formula_total(normalized))
        if not has_price:
            return False
        if QTY_UNIT_RE.search(normalized):
            return True
        hits = sum(1 for term in ORDER_HINT_TERMS if term in normalized)
        return hits >= 2

    def _looks_like_llm_order_candidate(self, text: str) -> bool:
        normalized = self._normalize_order_text(text)
        if len(normalized) < 3:
            return False
        if self._looks_like_cancel_or_status_line(normalized):
            return False
        if any(term in normalized for term in ("象牙宝平台", "走平台", "给登记", "登记下", "登记一下")):
            return False
        has_product_hint = bool(re.search(r"(试剂|离心管|细胞|抗体|蛋白|货号|型号|架|试纸|耗材|培养基|移液器|灯管|孔板|滤膜|培养皿)", normalized))
        has_quantity_hint = bool(QTY_UNIT_RE.search(normalized) or QTY_CN_UNIT_RE.search(normalized))
        has_money = bool(PRICE_RE.search(normalized) or self._extract_formula_total(normalized))
        if any(term in normalized for term in ("预存", "欠款", "还欠", "扣除", "补过来", "平台预存")) and not has_product_hint:
            return False
        if re.match(r"^(是的|好的|嗯|行|对|收到|ok|OK|嗯嗯)[，,\s]*\d+(?:\.\d+)?", normalized) and has_quantity_hint and not has_product_hint and not has_money:
            return False
        followup_match = FOLLOWUP_CONFIRM_RE.match(normalized)
        if followup_match and has_quantity_hint and not has_money:
            body = self._normalize_order_text(str(followup_match.group("body") or ""))
            if re.match(r"^(?:这|那)\s*\d+(?:\.\d+)?\s*(?:个|株|盒|瓶|套|包|支|件|板|提|组)", body):
                return False
            if re.match(r"^(?:都|全部|都按|按上面|按之前)", body):
                return False
        if any(token in normalized for token in LLM_ORDER_INTENT_TERMS):
            if has_product_hint or has_money or has_quantity_hint:
                return True
            return False
        return has_money and (has_product_hint or has_quantity_hint)

    def _looks_like_followup_confirmation(self, text: str) -> bool:
        normalized = self._normalize_order_text(text)
        if len(normalized) < 2:
            return False
        if any(term in normalized for term in FOLLOWUP_SETTLEMENT_HINTS):
            return False
        match = FOLLOWUP_CONFIRM_RE.match(normalized)
        if match:
            body = str(match.group("body") or normalized)
        else:
            # No explicit confirmation prefix: only consider very short numeric replies like "2瓶".
            if len(normalized) > 18:
                return False
            if PRICE_RE.search(normalized) or self._extract_formula_total(normalized):
                return False
            if any(term in normalized for term in ORDER_ACTION_TERMS):
                return False
            if re.search(r"(试剂|离心管|细胞|抗体|蛋白|货号|型号|培养|滤膜|孔板|手套)", normalized):
                return False
            body = normalized
        qty, unit = self._extract_quantity_unit(body)
        if qty and unit:
            return True
        if re.search(r"(?:补|改|改成|改为)\s*\d+(?:\.\d+)?\s*(?:个|盒|瓶|套|包|支|株|件|板|提|组)", body):
            return True
        if re.match(r"^\d+(?:\.\d+)?\s*(?:个|盒|瓶|套|包|支|株|件|板|提|组)$", body):
            return True
        return False

    def _apply_followup_confirmation(self, text: str, *, recent_rows: list[dict[str, Any]]) -> None:
        if not recent_rows:
            return
        normalized = self._normalize_order_text(text)
        match = FOLLOWUP_CONFIRM_RE.match(normalized)
        body = str(match.group("body") or normalized) if match else normalized
        qty, unit = self._extract_quantity_unit(body)
        qty_value = self._to_number(qty)
        if qty_value <= 0:
            if re.match(r"^\d+(?:\.\d+)?$", body):
                qty_value = self._to_number(body)
                qty = self._format_number(qty_value) if qty_value > 0 else ""
            else:
                return
        if qty_value <= 0:
            return
        target = None
        for row in reversed(recent_rows):
            record_type = str(row.get("record_type") or "order_item").strip().lower()
            if record_type not in {"order_item", "gift_item"}:
                continue
            if row.get("quantity"):
                continue
            target = row
            break
        if target is None:
            for row in reversed(recent_rows):
                record_type = str(row.get("record_type") or "order_item").strip().lower()
                if record_type in {"order_item", "gift_item"}:
                    target = row
                    break
        if target is None:
            return
        target["quantity"] = self._format_number(qty_value)
        if unit:
            target["unit"] = unit
            if not str(target.get("order_unit") or "").strip():
                target["order_unit"] = unit
        sale_price = self._to_number(str(target.get("sale_price") or ""))
        if sale_price > 0 and not str(target.get("total_sale") or "").strip():
            target["total_sale"] = self._format_number(sale_price * qty_value)
        target["needs_review"] = bool(
            float(target.get("confidence") or 0) < 0.75
            or not str(target.get("product_name") or "").strip()
            or not str(target.get("quantity") or "").strip()
        )
        target["confidence"] = min(float(target.get("confidence") or 0), 0.72)
        target["needs_review"] = True
        note = "上下文确认补全数量"
        remark = str(target.get("remark") or "")
        if note not in remark:
            target["remark"] = (remark + f" | {note}").strip(" |")

    def _should_force_multi_sku_split(self, content: str, *, min_skus: int = 2, action_threshold: int = 2) -> bool:
        text = self._normalize_order_text(content)
        if not text:
            return False
        action_count = len(re.findall(r"(?:各订|订|下单|代付|买|补订|再订|重订)", text))
        if action_count >= action_threshold:
            return True
        sku_like_hits = 0
        sku_like_hits += len(re.findall(r"(?:货号|型号)\s*[:：]?\s*[A-Za-z0-9._-]+", text, flags=re.IGNORECASE))
        sku_like_hits += len(re.findall(r"(?:、|,|，)\s*[A-Za-z0-9\u4e00-\u9fa5][^、,，]{1,28}\s*(?:\d+(?:\.\d+)?\s*(?:个|盒|瓶|套|包|支|株|件|板|提|组))", text))
        if sku_like_hits >= min_skus:
            return True
        if re.search(r"这\s*\d+\s*(?:个|株|盒|瓶|支|包|件|板|提|组)", text):
            return True
        if "分别" in text:
            return True
        return False

    def _llm_segment_order_message(self, content: str, *, max_segments: int = 6) -> list[str]:
        text = self._normalize_order_text(content)
        if not text:
            return []
        prompt = {
            "task": "将source_text按下单语义切分为多个候选订货片段。仅返回包含订货语义的片段。",
            "source_text": text,
            "rules": [
                "一个片段只表达一条订货信息。",
                "如果文本不含订货语义，返回空数组。",
                "不要改写原文含义，可做最小清洗。",
            ],
            "max_segments": int(max_segments),
            "response_contract": {
                "segments": ["string"],
                "confidence": "0-1",
                "reason": "string",
            },
        }
        payload = self._call_deepseek_json_cached(prompt, namespace="segment_message")
        segments_raw = payload.get("segments")
        if not isinstance(segments_raw, list):
            return []
        segments: list[str] = []
        for item in segments_raw:
            snippet = self._normalize_order_text(str(item or ""))
            if not snippet:
                continue
            if self._looks_like_cancel_or_status_line(snippet):
                continue
            if snippet not in segments:
                segments.append(snippet)
            if len(segments) >= max_segments:
                break
        return segments

    def _split_message_by_price_terminator(self, content: str, *, max_segments: int = 6) -> list[str]:
        lines = [self._normalize_order_text(line) for line in re.split(r"[\r\n]+", str(content or "")) if self._normalize_order_text(line)]
        if not lines:
            return []
        brand_context = self._extract_message_brand_context(content)
        segments: list[str] = []
        buffer: list[str] = []

        def flush_buffer() -> None:
            if not buffer:
                return
            raw = self._normalize_order_text(" ".join(buffer))
            if not raw:
                buffer.clear()
                return
            split_items = self._split_text_by_yuan_terminator(raw, max_segments=max_segments)
            for item in split_items:
                normalized_item = self._normalize_order_text(item)
                if not normalized_item:
                    continue
                if brand_context and not self._extract_brand_name(normalized_item):
                    normalized_item = self._normalize_order_text(f"{brand_context} {normalized_item}")
                if normalized_item not in segments:
                    segments.append(normalized_item)
            buffer.clear()

        pending_prefix = ""
        for idx, line in enumerate(lines):
            if pending_prefix:
                line = self._normalize_order_text(f"{pending_prefix}{line}")
                pending_prefix = ""
            if re.match(r"^[\u4e00-\u9fa5A-Za-z0-9]{2,24}老师$", line):
                flush_buffer()
                continue
            if self._looks_like_non_product_line(line) and not (PRICE_RE.search(line) or self._extract_formula_total(line)):
                continue
            yuan_matches = list(re.finditer(r"元", line))
            if yuan_matches and idx + 1 < len(lines):
                tail = self._normalize_order_text(line[yuan_matches[-1].end() :])
                next_line = lines[idx + 1]
                if (
                    tail
                    and len(tail) <= 12
                    and not re.search(r"(老师|收货|合计|共计|运费|优惠)", tail)
                    and (PRICE_RE.search(next_line) or self._extract_formula_total(next_line))
                ):
                    line = self._normalize_order_text(line[: yuan_matches[-1].end()])
                    pending_prefix = tail
            buffer.append(line)
            if PRICE_RE.search(line) or self._extract_formula_total(line):
                flush_buffer()
            if len(segments) >= max_segments:
                break
        flush_buffer()
        return segments[:max_segments]

    def _message_complexity_score(self, content: str) -> int:
        text = self._normalize_order_text(content)
        if not text:
            return 0
        score = 0
        lines = [line.strip() for line in re.split(r"[\r\n]+", text) if line.strip()]
        if len(lines) >= 2:
            score += 1
        order_actions = len(re.findall(r"(?:各订|订|下单|代付|买|补订|再订|重订)", text))
        if order_actions >= 2:
            score += 1
        qty_hits = len(QTY_UNIT_RE.findall(text)) + len(QTY_CN_UNIT_RE.findall(text))
        if qty_hits >= 2:
            score += 1
        money_hits = len(re.findall(r"\d+(?:\.\d+)?\s*元", text))
        formula_hits = len(re.findall(r"=\s*\d+(?:\.\d+)?\s*元?", text))
        if money_hits + formula_hits >= 2:
            score += 1
        if any(token in text for token in ("[引用", "货号", "型号", "规格", "发活细胞")):
            score += 1
        if re.search(r"这\s*\d+\s*(?:株|个|盒|瓶|套|包|支|件)", text) or "分别" in text:
            score += 1
        if re.search(r"[()（）]", text) and any(term in text for term in ORDER_ACTION_TERMS):
            score += 1
        if len(text) >= 140:
            score += 1
        return min(score, 6)

    def _rows_quality_score(self, rows: list[dict[str, Any]], *, source_text: str = "") -> float:
        if not rows:
            return 0.0
        source = self._normalize_order_text(source_text)
        scores: list[float] = []
        for row in rows:
            if not isinstance(row, dict):
                continue
            merged_row = {**self._empty_order_row(), **row}
            product = self._normalize_order_text(str(merged_row.get("product_name") or ""))
            name = self._normalize_order_text(str(merged_row.get("name") or ""))
            owner = self._normalize_order_text(str(merged_row.get("owner") or ""))
            unit = self._normalize_order_text(str(merged_row.get("unit") or ""))
            quantity = self._to_number(str(merged_row.get("quantity") or ""))
            sale_price = self._to_number(str(merged_row.get("sale_price") or ""))
            total_sale = self._to_number(str(merged_row.get("total_sale") or ""))
            confidence = self._coerce_confidence(merged_row.get("confidence"))
            score = 0.0
            if product:
                score += 1.8
                if re.fullmatch(r"[\u4e00-\u9fa5]{2,4}", product) and "老师" in source:
                    score -= 0.8
            if quantity > 0:
                score += 0.9
            if unit:
                score += 0.4
            if sale_price > 0:
                score += 0.5
            if total_sale > 0:
                score += 0.6
            if name:
                score += 0.2
            if owner:
                score += 0.2
            if merged_row.get("date"):
                score += 0.2
            if merged_row.get("needs_review") is False:
                score += 0.5
            score += confidence * 0.3
            if product and source and product in source:
                score += 0.2
            if self._should_drop_order_row(merged_row):
                score -= 1.2
            if not product and sale_price <= 0 and total_sale <= 0:
                score -= 1.0
            scores.append(score)
        if not scores:
            return 0.0
        return sum(scores) / float(len(scores))

    def _guess_scope_text_for_row(self, row: dict[str, Any], *, candidate_segments: list[str], fallback_text: str = "") -> str:
        if not isinstance(row, dict):
            return self._normalize_order_text(fallback_text)
        explicit = self._normalize_order_text(str(row.get("source_scope_text") or ""))
        if explicit:
            return explicit
        evidence = self._normalize_order_text(str(row.get("evidence_text") or ""))
        product = self._normalize_order_text(str(row.get("product_name") or ""))
        best_segment = ""
        best_score = -1
        for segment in candidate_segments:
            normalized_segment = self._normalize_order_text(segment)
            if not normalized_segment:
                continue
            score = 0
            if evidence and evidence in normalized_segment:
                score += 4
            if product and product in normalized_segment:
                score += 3
            if evidence and normalized_segment in evidence:
                score += 1
            if product and normalized_segment in product:
                score += 1
            if score > best_score:
                best_score = score
                best_segment = normalized_segment
        if best_score > 0 and best_segment:
            return best_segment
        return self._normalize_order_text(fallback_text)

    def _backfill_row_from_scope_text(
        self,
        row: dict[str, Any],
        *,
        scope_text: str,
        source_text: str = "",
        strict_scope: bool = False,
    ) -> dict[str, Any]:
        output = dict(row)
        scope_parts = [scope_text]
        if strict_scope:
            scope_parts.append(str(output.get("remark") or ""))
        else:
            scope_parts.extend([str(output.get("remark") or ""), str(output.get("evidence_text") or "")])
        scope = self._normalize_order_text(
            " ".join(item for item in scope_parts if str(item or "").strip())
        )
        if not scope:
            return output
        brand = self._normalize_order_text(str(output.get("brand") or ""))
        product = self._normalize_order_text(str(output.get("product_name") or ""))
        if not brand:
            inferred_brand = self._extract_brand_name(scope)
            if inferred_brand and not self._brand_token_looks_like_person_in_source(inferred_brand, scope):
                brand = inferred_brand
                output["brand"] = brand
        if not product:
            inferred_product = self._extract_product_name(scope)
            if inferred_product:
                product = inferred_product
                output["product_name"] = inferred_product
        if brand and product:
            output["product_name"] = self._prefix_brand_to_product_name(output.get("product_name") or "", brand, source_text=scope)
        qty_value = self._to_number(str(output.get("quantity") or ""))
        unit_text = self._normalize_order_text(str(output.get("unit") or ""))
        if qty_value <= 0:
            qty, unit = self._extract_quantity_unit(scope)
            qty_num = self._to_number(qty)
            if qty_num > 0:
                output["quantity"] = self._format_number(qty_num)
                qty_value = qty_num
            if unit and not unit_text:
                output["unit"] = unit
                unit_text = unit
        if unit_text and not str(output.get("order_unit") or "").strip():
            output["order_unit"] = unit_text
        formula_total = self._to_number(self._extract_formula_total(scope))
        total_sale_value = self._to_number(str(output.get("total_sale") or ""))
        sale_price_value = self._to_number(str(output.get("sale_price") or ""))
        total_cost_value = self._to_number(str(output.get("total_cost") or ""))
        if total_sale_value <= 0 and formula_total > 0:
            total_sale_value = formula_total
            output["total_sale"] = self._format_number(formula_total)
        if total_sale_value <= 0 and total_cost_value > 0:
            total_sale_value = total_cost_value
            output["total_sale"] = self._format_number(total_cost_value)
            note = "LLM字段纠偏: total_cost->total_sale"
            remark = str(output.get("remark") or "")
            if note not in remark:
                output["remark"] = (remark + f" | {note}").strip(" |")
        if sale_price_value <= 0:
            price_match = PRICE_RE.search(scope)
            if price_match:
                candidate_price = self._to_number(str(price_match.group("price") or ""))
                if candidate_price > 0:
                    sale_price_value = candidate_price
                    output["sale_price"] = self._format_number(candidate_price)
        if sale_price_value <= 0 and total_sale_value > 0:
            if qty_value > 0:
                sale_price_value = total_sale_value / qty_value
                output["sale_price"] = self._format_number(sale_price_value)
            elif qty_value == 0:
                output["sale_price"] = self._format_number(total_sale_value)
        if total_sale_value <= 0 and sale_price_value > 0 and qty_value > 0:
            total_sale_value = sale_price_value * qty_value
            output["total_sale"] = self._format_number(total_sale_value)
        if qty_value <= 0 and (sale_price_value > 0 or total_sale_value > 0):
            if any(term in scope for term in ORDER_ACTION_TERMS) or NAME_OWNER_RE.search(scope):
                output["quantity"] = "1"
                qty_value = 1.0
                if not str(output.get("unit") or "").strip():
                    output["unit"] = str(output.get("order_unit") or "") or "个"
                if not str(output.get("order_unit") or "").strip():
                    output["order_unit"] = str(output.get("unit") or "") or "个"
                if total_sale_value <= 0 and sale_price_value > 0:
                    total_sale_value = sale_price_value
                    output["total_sale"] = self._format_number(total_sale_value)
                note = "数量策略: 默认1"
                remark = str(output.get("remark") or "")
                if note not in remark:
                    output["remark"] = (remark + f" | {note}").strip(" |")
                output["needs_review"] = True
        current_spec = self._normalize_spec_text(str(output.get("spec") or ""))
        if not current_spec:
            spec_candidate = self._extract_product_spec(
                scope,
                product_name=str(output.get("product_name") or ""),
                brand_hint=str(output.get("brand") or ""),
            )
            if spec_candidate:
                output["spec"] = spec_candidate
        return output

    def _sanitize_product_name(self, product_name: str, *, scope_text: str = "") -> str:
        product = self._normalize_order_text(product_name)
        if not product:
            return ""
        product = re.sub(
            rf"\s*(?:订|已订|代付|下单|补订|再订|重订|买)\s*[-一二两俩三四五六七八九十百\d]*(?:\.\d+)?\s*(?:{ORDER_UNIT_PATTERN})?.*$",
            "",
            product,
        )
        if "元" in product or re.match(r"^\d+(?:\.\d+)?\s*[*xX×/+.-]\s*\d+(?:\.\d+)?", product):
            product = re.sub(
                r"\d+(?:\.\d+)?\s*[*xX×/+.-]\s*\d+(?:\.\d+)?(?:\s*=\s*\d+(?:\.\d+)?)?\s*元?.*$",
                "",
                product,
            )
        product = self._normalize_order_text(product).strip("：:，,；;。 ")
        if not product:
            return ""
        if re.fullmatch(r"\d+(?:\.\d+)?\s*元?", product):
            return ""
        if re.fullmatch(r"\d+(?:\.\d+)?(?:\s*[*xX×/+.-]\s*\d+(?:\.\d+)?)+\s*(?:=\s*\d+(?:\.\d+)?)?\s*元?", product):
            return ""
        if not re.search(r"[\u4e00-\u9fa5A-Za-z]", product):
            return ""
        scope = self._normalize_order_text(scope_text)
        if "元" in product and not any(token in product for token in ("试剂", "管", "盒", "瓶", "水", "膜", "板", "酶", "酸")):
            return ""
        product = re.sub(r"(?<=[\u4e00-\u9fa5])\s+(?=[\u4e00-\u9fa5])", "", product)
        return product

    def _remove_spec_tokens_from_product_name(self, product_name: str, spec: str) -> str:
        product = self._normalize_order_text(product_name)
        normalized_spec = self._normalize_spec_text(spec)
        if not product or not normalized_spec:
            return product
        cleaned = product
        primary_code = self._extract_primary_spec_code(normalized_spec)
        if primary_code:
            cleaned = re.sub(re.escape(primary_code), " ", cleaned, flags=re.IGNORECASE)
        for marker_match in SPEC_MARKER_RE.finditer(cleaned):
            token = self._normalize_spec_text(str(marker_match.group("spec") or ""))
            if token and self._normalize_spec_key(token) in self._normalize_spec_key(normalized_spec):
                cleaned = cleaned.replace(str(marker_match.group(0) or ""), " ")
        for formulation_match in FORMULATION_SPEC_RE.finditer(cleaned):
            formulation = self._normalize_spec_text(str(formulation_match.group(0) or ""))
            if formulation and self._normalize_spec_key(formulation) in self._normalize_spec_key(normalized_spec):
                cleaned = cleaned.replace(str(formulation_match.group(0) or ""), " ")
        cleaned = self._normalize_order_text(cleaned).strip("：:，,；;。 -")
        cleaned = re.sub(r"(?<=[\u4e00-\u9fa5])\s+(?=[\u4e00-\u9fa5])", "", cleaned)
        return cleaned or product

    def _repair_split_spec_prefix_from_product(
        self,
        *,
        product_name: str,
        spec: str,
        context_text: str,
    ) -> tuple[str, str]:
        product = self._normalize_order_text(product_name)
        normalized_spec = self._normalize_spec_text(spec)
        if not product or not normalized_spec:
            return product, normalized_spec
        first_token_match = re.match(r"(?P<token>[A-Za-z0-9][A-Za-z0-9._/\-*]{1,32})", normalized_spec)
        if not first_token_match:
            return product, normalized_spec
        first_token = self._normalize_spec_text(str(first_token_match.group("token") or ""))
        if not first_token or not re.search(r"[A-Za-z]", first_token) or not re.search(r"\d", first_token):
            return product, normalized_spec
        inline_prefix_match = re.search(
            rf"(?P<prefix>[A-Za-z]{{2,6}})\s*[-－–—]?\s*{re.escape(first_token)}(?=$|[\s，,；;:：])",
            product,
        )
        prefix_match = inline_prefix_match or re.search(r"(?P<prefix>[A-Za-z]{2,6})$", product)
        if not prefix_match:
            return product, normalized_spec
        prefix = self._normalize_spec_text(str(prefix_match.group("prefix") or ""))
        if prefix.upper() in NON_BRAND_STOPWORDS or prefix.lower() in {"mg", "kg", "ml", "ul", "mm", "cm"}:
            return product, normalized_spec
        joined_dash = self._normalize_spec_text(f"{prefix}-{first_token}")
        joined_plain = self._normalize_spec_text(f"{prefix}{first_token}")
        context_key = self._normalize_spec_key(context_text)
        joined = ""
        if self._normalize_spec_key(joined_dash) in context_key:
            joined = joined_dash
        elif self._normalize_spec_key(joined_plain) in context_key:
            joined = joined_dash if self._looks_like_spec_candidate(joined_dash, allow_numeric_only=True) else joined_plain
        elif len(prefix) >= 2 and self._looks_like_spec_candidate(joined_dash, allow_numeric_only=True):
            joined = joined_dash
        if not joined:
            return product, normalized_spec
        repaired_product = self._normalize_order_text(product[: prefix_match.start()].strip(" -_/，,；;:："))
        repaired_spec = self._normalize_spec_text(re.sub(re.escape(first_token), joined, normalized_spec, count=1))
        return repaired_product or product, repaired_spec or normalized_spec

    def _scope_supports_quantity(self, quantity: str, unit: str, *, scope_text: str) -> bool:
        qty_value = self._to_number(quantity)
        if qty_value <= 0:
            return False
        scope = self._normalize_order_text(scope_text)
        if not scope:
            return False
        qty = self._format_number(qty_value)
        unit_text = self._normalize_order_text(unit)
        qty_pattern = re.escape(qty).replace(r"\.0", r"(?:\.0)?")
        if unit_text and re.search(rf"(?<!\d){qty_pattern}\s*{re.escape(unit_text)}", scope):
            return True
        if qty_value == 1 and unit_text and re.search(rf"(?:一|1|一个)\s*{re.escape(unit_text)}", scope):
            return True
        if qty_value == 1 and unit_text and re.search(rf"(?:订|已订|买|代付|下单)\s*(?:一|1|一个|[-—－–ー])?\s*{re.escape(unit_text)}", scope):
            return True
        if re.search(rf"[*xX×]\s*{qty_pattern}\s*=", scope):
            return True
        if re.search(rf"(?<!\d){qty_pattern}\s*[*xX×]\s*\d+(?:\.\d+)?\s*=", scope):
            return True
        if qty_value == 1 and re.search(r"(?:订|已订|买|代付|下单)\s*(?:一|1|一个)", scope):
            return True
        return False

    def _clear_unsupported_quantity_from_scope(self, row: dict[str, Any], *, scope_text: str) -> dict[str, Any]:
        output = dict(row)
        quantity = str(output.get("quantity") or "").strip()
        if not quantity:
            return output
        unit = self._normalize_order_text(str(output.get("unit") or ""))
        if self._scope_supports_quantity(quantity, unit, scope_text=scope_text):
            return output
        if (
            self._to_number(quantity) == 1
            and "数量策略: 默认1" in str(output.get("remark") or "")
            and str(output.get("product_name") or "").strip()
            and (
                PRICE_RE.search(self._normalize_order_text(scope_text))
                or self._extract_formula_total(self._normalize_order_text(scope_text))
            )
        ):
            if not unit:
                output["unit"] = str(output.get("order_unit") or "") or "个"
            if not str(output.get("order_unit") or "").strip():
                output["order_unit"] = str(output.get("unit") or "") or "个"
            output["needs_review"] = True
            return output
        output["quantity"] = ""
        if unit and unit not in self._normalize_order_text(scope_text):
            output["unit"] = ""
            if str(output.get("order_unit") or "").strip() == unit:
                output["order_unit"] = ""
        output["needs_review"] = True
        output["confidence"] = min(float(output.get("confidence") or 0), 0.62)
        note = "数量超出当前产品片段，已清空"
        remark = str(output.get("remark") or "")
        if note not in remark:
            output["remark"] = (remark + f" | {note}").strip(" |")
        return output

    def _enforce_single_use_spec_for_row(
        self,
        row: dict[str, Any],
        *,
        used_spec_owners: dict[str, str],
        source_text: str,
    ) -> dict[str, Any]:
        output = dict(row)
        spec_text = self._normalize_spec_text(str(output.get("spec") or ""))
        if not spec_text:
            return output
        product_key = self._normalize_order_text(str(output.get("product_name") or "")).lower()
        spec_key = self._extract_spec_key(spec_text)
        if not spec_key:
            return output
        owner = used_spec_owners.get(spec_key, "")
        if not owner:
            used_spec_owners[spec_key] = product_key or "__unknown_product__"
            return output
        if owner == "__unknown_product__" and product_key:
            used_spec_owners[spec_key] = product_key
            return output
        if owner == (product_key or "__unknown_product__"):
            return output
        row_scope = self._normalize_order_text(
            f"{str(output.get('source_scope_text') or '')} {str(output.get('remark') or '')} {str(output.get('product_name') or '')}"
        )
        fallback_scope = row_scope or self._normalize_order_text(f"{str(output.get('evidence_text') or '')} {source_text}")
        alt_spec = self._extract_product_spec(
            fallback_scope,
            product_name=str(output.get("product_name") or ""),
            brand_hint=str(output.get("brand") or ""),
        )
        alt_key = self._extract_spec_key(alt_spec)
        if alt_spec and alt_key and (alt_key not in used_spec_owners or used_spec_owners.get(alt_key) == (product_key or "__unknown_product__")):
            output["spec"] = self._normalize_spec_text(alt_spec)
            used_spec_owners[alt_key] = product_key or "__unknown_product__"
            return output
        output["spec"] = ""
        output["needs_review"] = True
        conflict_note = "规格冲突: 同型号跨产品复用已清空"
        remark = str(output.get("remark") or "")
        if conflict_note not in remark:
            output["remark"] = (remark + f" | {conflict_note}").strip(" |")
        return output

    def _llm_extract_rows(
        self,
        message: dict[str, Any],
        *,
        date_output_mode: str = "YYYY-MM-DD",
        use_segmentation: bool = True,
        max_segments: int = 6,
        force_multi_split: bool = False,
    ) -> list[dict[str, Any]]:
        content = str(message.get("content") or "").strip()
        if not content:
            return []
        normalized_content = self._normalize_order_text(content)
        segments: list[str] = []
        complexity = self._message_complexity_score(normalized_content)
        deterministic_segments = self._split_message_by_price_terminator(normalized_content, max_segments=max_segments)
        if use_segmentation and (complexity >= 2 or "\n" in content or len(normalized_content) >= 120):
            segments = self._llm_segment_order_message(normalized_content, max_segments=max_segments)
        if deterministic_segments:
            merged_segments: list[str] = []
            for candidate in [*deterministic_segments, *segments]:
                normalized_candidate = self._normalize_order_text(candidate)
                if not normalized_candidate or normalized_candidate in merged_segments:
                    continue
                merged_segments.append(normalized_candidate)
            segments = merged_segments[:max_segments]
        extraction_scope = segments if segments else [normalized_content]
        prompt = {
            "task": "从微信聊天 source_text 中提取 0-N 条订货结构化记录，返回 rows 数组。必须仅依据原文，不允许虚构。",
            "source_text": normalized_content,
            "candidate_segments": extraction_scope,
            "rules": [
                "允许拆分成多条 rows。",
                "群成员名、发送人姓名、speaker_name 不是订单正文，不能当作品牌或产品。",
                "引用/回复/历史预览内容不是当前订单，不能从引用内容提取产品、品牌、规格、价格或姓名。",
                "一个气泡是一条独立订单语义范围，不同气泡之间不能串用品牌、规格、产品或价格。",
                "人名、老师名不能作为 product_name。",
                "括号内容优先写入 remark，不要替代主产品名。",
                "当出现多个产品名（尤其带多个货号/多行产品）时，必须拆成多条 rows，而不是合并为一条。",
                "若一段消息中出现多个以“元”结尾的价格片段，默认每个价格片段对应一条产品记录（运费/优惠/合计除外）。",
                "“元”可视为当前产品结束边界；若边界后出现新品牌/新产品，则两者之间的货号/型号/规格应归属上一条产品。",
                "若开头给出品牌名，后续产品没有出现其他品牌名时，品牌字段沿用该品牌。",
                "若原文出现品牌+产品组合，product_name 保留品牌前缀（便于业务识别），同时 brand 字段也要填品牌。",
                "出现“这3个/这4个/各订/分别”等表达时，优先按产品粒度拆分；若只有总价，允许按等分估算单价并在 remark 标注“总价等分”。",
                f"日期优先输出 {date_output_mode}；若无日期可留空。",
                "无法确定字段允许留空，并将 needs_review=true。",
            ],
            "force_multi_split": bool(force_multi_split),
            "row_schema": {
                "date": "string",
                "name": "string",
                "owner": "string",
                "receiver": "string",
                "campus": "string",
                "order_unit": "string",
                "product_name": "string",
                "quantity": "string",
                "unit": "string",
                "spec": "string",
                "brand": "string",
                "cost_price": "string",
                "sale_price": "string",
                "total_cost": "string",
                "total_sale": "string",
                "remark": "string",
                "evidence_text": "string",
                "record_type": "order_item|gift_item|summary|settlement|confirmation|other",
                "confidence": "0-1",
                "needs_review": "bool",
                "reason": "string",
            },
            "response_contract": {
                "rows": "array<object>",
                "confidence": "0-1",
                "reason": "string",
            },
        }
        payload: dict[str, Any] = {}
        try:
            raw_payload = self._call_deepseek_json_cached(prompt, namespace="extract_rows")
            if isinstance(raw_payload, dict):
                payload = raw_payload
        except Exception:
            payload = {}

        rows_raw: list[dict[str, Any]] = []
        if isinstance(payload.get("rows"), list):
            rows_raw = [item for item in payload.get("rows", []) if isinstance(item, dict)]
        elif isinstance(payload.get("row"), dict):
            rows_raw = [payload.get("row")]

        # Fallback: segment-level extraction when payload is empty or likely under-split.
        need_segment_fallback = not rows_raw
        if rows_raw and segments and len(segments) > 1 and len(rows_raw) == 1:
            first_product = self._normalize_order_text(str(rows_raw[0].get("product_name") or ""))
            if re.search(r"(这\s*\d+\s*(?:个|株|盒|瓶|支|包)|各订|分别|买二送一|赠品选择)", normalized_content):
                need_segment_fallback = True
            elif sum(1 for sep in ("、", ",", "，", "货号：", "货号:") if sep in first_product) >= 1:
                need_segment_fallback = True
        if force_multi_split and len(rows_raw) <= 1:
            need_segment_fallback = True

        if need_segment_fallback and segments:
            segment_rows: list[dict[str, Any]] = []
            for segment in segments[:max_segments]:
                segment_message = dict(message)
                segment_message["content"] = segment
                segment_row = self._llm_extract_single_row(segment_message, date_output_mode=date_output_mode)
                if isinstance(segment_row, dict):
                    segment_row["source_scope_text"] = self._normalize_order_text(segment)
                    segment_rows.append(segment_row)
            if segment_rows and (not rows_raw or len(segment_rows) >= len(rows_raw)):
                rows_raw = segment_rows
        if force_multi_split and len(rows_raw) <= 1:
            refined_rows = self._llm_refine_multi_rows(
                source_text=normalized_content,
                candidate_segments=segments if segments else [normalized_content],
                current_rows=rows_raw,
                date_output_mode=date_output_mode,
            )
            if len(refined_rows) > len(rows_raw):
                rows_raw = refined_rows

        # Deterministic safeguard: when price terminator already yields multi-segments,
        # extract each segment independently to avoid cross-segment spec/brand bleeding.
        if len(deterministic_segments) >= 2:
            deterministic_rows: list[dict[str, Any]] = []
            for segment in deterministic_segments[:max_segments]:
                normalized_segment = self._normalize_order_text(segment)
                if not normalized_segment or not self._line_has_order_signal(normalized_segment):
                    continue
                segment_message = dict(message)
                segment_message["content"] = normalized_segment
                segment_row = self._llm_extract_single_row(segment_message, date_output_mode=date_output_mode)
                if isinstance(segment_row, dict):
                    segment_row["source_scope_text"] = normalized_segment
                    deterministic_rows.append(segment_row)
            if deterministic_rows:
                current_quality = self._rows_quality_score(
                    [item for item in rows_raw if isinstance(item, dict)],
                    source_text=normalized_content,
                )
                deterministic_quality = self._rows_quality_score(deterministic_rows, source_text=normalized_content)
                if len(deterministic_rows) >= max(1, len(rows_raw)) or deterministic_quality + 0.2 >= current_quality:
                    rows_raw = deterministic_rows

        results: list[dict[str, Any]] = []
        seen: set[str] = set()
        used_spec_owners: dict[str, str] = {}
        payload_reason = str(payload.get("reason") or "")
        payload_confidence = self._coerce_confidence(payload.get("confidence"))
        message_id = str(message.get("raw_message_id") or "")
        for raw_row in rows_raw:
            if not isinstance(raw_row, dict):
                continue
            merged = self._empty_order_row()
            for key in merged.keys():
                if key in {"needs_review", "evidence_message_ids"}:
                    continue
                if key in raw_row:
                    merged[key] = raw_row.get(key)
            scope_text = self._normalize_order_text(str(raw_row.get("source_scope_text") or ""))
            if not scope_text:
                scope_text = self._guess_scope_text_for_row(
                    raw_row,
                    candidate_segments=extraction_scope,
                    fallback_text=normalized_content,
                )
            if scope_text:
                merged["source_scope_text"] = scope_text
            merged = self._backfill_row_from_scope_text(
                merged,
                scope_text=scope_text or normalized_content,
                source_text=normalized_content,
            )
            row_confidence = self._coerce_confidence(raw_row.get("confidence"))
            merged["confidence"] = max(row_confidence, payload_confidence)
            merged["needs_review"] = bool(raw_row.get("needs_review", merged["confidence"] < 0.75))
            row_reason = str(raw_row.get("reason") or payload_reason or "")
            if row_reason:
                existing_remark = str(merged.get("remark") or "")
                if row_reason not in existing_remark:
                    merged["remark"] = (existing_remark + f" | {row_reason}").strip(" |")
            if not str(merged.get("evidence_text") or "").strip():
                merged["evidence_text"] = str(raw_row.get("evidence_text") or normalized_content[:200])
            merged["evidence_message_ids"] = [message_id] if message_id else []
            finalized = self._finalize_order_row(merged, message, date_output_mode=date_output_mode)
            finalized = self._enforce_single_use_spec_for_row(
                finalized,
                used_spec_owners=used_spec_owners,
                source_text=normalized_content,
            )
            if self._should_drop_order_row(finalized):
                continue
            row_key = "|".join(
                [
                    str(finalized.get("date") or ""),
                    self._normalize_order_text(str(finalized.get("product_name") or "")).lower(),
                    self._format_number(self._to_number(str(finalized.get("quantity") or ""))) if finalized.get("quantity") else "",
                    self._normalize_order_text(str(finalized.get("unit") or "")).lower(),
                    self._format_number(self._to_number(str(finalized.get("sale_price") or ""))) if finalized.get("sale_price") else "",
                    self._format_number(self._to_number(str(finalized.get("total_sale") or ""))) if finalized.get("total_sale") else "",
                    self._normalize_order_text(str(finalized.get("name") or "")).lower(),
                ]
            )
            if row_key in seen:
                continue
            seen.add(row_key)
            results.append(finalized)
        return self._apply_brand_context_to_rows(results, source_text=normalized_content)

    def _llm_refine_multi_rows(
        self,
        *,
        source_text: str,
        candidate_segments: list[str],
        current_rows: list[dict[str, Any]],
        date_output_mode: str,
    ) -> list[dict[str, Any]]:
        text = self._normalize_order_text(source_text)
        if not text:
            return []
        prompt = {
            "task": "你是订单结构化复核器。请把 current_rows 修正为按SKU拆分后的 rows，尤其处理多货号/多产品合并成一条的问题。",
            "source_text": text,
            "candidate_segments": [self._normalize_order_text(item) for item in candidate_segments if self._normalize_order_text(item)],
            "current_rows": current_rows,
            "rules": [
                "必须仅依据 source_text。",
                "一个SKU一条row，禁止把多个产品并在同一条。",
                "如果 source_text 中存在多个以“元”结束的价格片段，默认拆成多条 row。",
                "“元”后若出现新品牌/新产品，则这段之间出现的货号/型号/规格归到上一条产品。",
                "若首行给出品牌名，且后续没有新品牌，则所有 row 的 brand 继承该品牌。",
                "人名/老师名不能成为 product_name。",
                "无法确定可留空并 needs_review=true。",
                f"日期优先输出 {date_output_mode}。",
            ],
            "response_contract": {
                "rows": "array<object>",
                "reason": "string",
                "confidence": "0-1",
            },
        }
        payload: dict[str, Any] = {}
        try:
            raw_payload = self._call_deepseek_json_cached(prompt, namespace="refine_multi_rows")
            if isinstance(raw_payload, dict):
                payload = raw_payload
        except Exception:
            payload = {}
        rows_raw = payload.get("rows")
        if not isinstance(rows_raw, list):
            return []
        refined: list[dict[str, Any]] = []
        for item in rows_raw:
            if isinstance(item, dict):
                refined.append(item)
        return refined

    def _line_has_order_signal(self, line: str) -> bool:
        normalized = self._normalize_order_text(line)
        if self._looks_like_cancel_or_status_line(normalized):
            return False
        has_price = bool(PRICE_RE.search(normalized) or self._extract_formula_total(normalized))
        qty, unit = self._extract_quantity_unit(normalized)
        has_order_action = any(term in normalized for term in ORDER_ACTION_TERMS)
        if qty and unit and has_price:
            return True
        if has_price and has_order_action:
            return True
        if re.search(r"(各订|订|下单|买)\s*[一二两俩三四五六七八九十百半\d]", normalized) and has_price:
            return True
        if self._extract_formula_total(normalized) and any(term in normalized for term in ("订", "代付", "各订", "发货", "已订")):
            return True
        if has_price and NAME_OWNER_RE.search(normalized):
            product = self._extract_product_name(normalized)
            if product:
                return True
        if has_price:
            product = self._extract_product_name(normalized)
            if product and not self._looks_like_non_product_line(product):
                if not re.search(r"(运费|快递费|合计|共计|满减|优惠|抹零)", product):
                    return True
        return False

    @staticmethod
    def _looks_like_cancel_or_status_line(text: str) -> bool:
        normalized = str(text or "").strip()
        if not normalized:
            return True
        if any(term in normalized for term in CANCEL_OR_STATUS_TERMS):
            return True
        if INVENTORY_STATUS_RE.search(normalized):
            return True
        if re.search(r"(?:现货|库存).{0,20}(?:差|缺)\s*\d+(?:\.\d+)?\s*(?:瓶|盒|箱|包|支|套|个|件|板|提|组)", normalized):
            return True
        if normalized in {"已订", "现货", "库存有现货"}:
            return True
        return False

    def _should_drop_order_row(self, row: dict[str, Any]) -> bool:
        record_type = str(row.get("record_type") or "order_item").strip().lower()
        product = self._normalize_order_text(str(row.get("product_name") or ""))
        remark = self._normalize_order_text(str(row.get("remark") or ""))
        name = self._normalize_order_text(str(row.get("name") or ""))
        brand = self._normalize_order_text(str(row.get("brand") or ""))
        quantity = self._to_number(str(row.get("quantity") or ""))
        total = self._to_number(str(row.get("total_sale") or ""))
        sale_price = self._to_number(str(row.get("sale_price") or ""))
        if record_type in {"summary", "settlement", "confirmation", "other"}:
            return True
        if record_type == "gift_item":
            if not product:
                return True
            if quantity <= 0 and total <= 0 and sale_price <= 0 and "赠" not in f"{product}{remark}":
                return True
        if not product:
            return True
        if product and not name and quantity <= 0 and total <= 0 and sale_price <= 0:
            return True
        if len(product) <= 2 and total <= 0 and sale_price <= 0:
            return True
        if product in {"都", "这个", "那个", "再加干冰费", "急用", "麻烦给订上", "麻烦给订上[抱拳"}:
            return True
        if product in {"书", "打印纸", "a4打印纸", "a4纸"} or any(term in product.lower() for term in ("打印纸", "a4打印")):
            return True
        if product.lower() in {"μ", "ul", "μl", "μl。", "ul。"}:
            return True
        if re.fullmatch(r"\d+\s*(?:μ|μl|ul|ml)", product.lower()):
            return True
        if re.search(r"(?:\bmM\b|DMSO)", product, flags=re.IGNORECASE) and not re.search(r"[\u4e00-\u9fa5]", product):
            if not re.search(r"\b(?:acid|antibody|protein|kit|buffer|medium|reagent|tube|membrane)\b", product, flags=re.IGNORECASE):
                return True
        if brand.lower() in {"none", "null", "unknown"} and re.search(r"(?:\bmM\b|DMSO|/)", product, flags=re.IGNORECASE):
            return True
        if any(term in product for term in ("象牙宝平台", "走平台", "给登记", "登记下", "登记一下")):
            return True
        if any(term in product for term in ("确定一下", "是不是", "是否")) and not name and total <= 0 and sale_price <= 0:
            return True
        if "价格" in product and ("运费" in product or "产品" in product) and quantity <= 0:
            return True
        if re.search(r"(现货|不供货|先带走|已带走|加干冰费|麻烦给订上|急用)", product):
            if total <= 0 and sale_price <= 0 and ("订" not in remark or quantity <= 0):
                return True
        if product in NON_ORDER_PRODUCT_HINTS:
            return True
        if any(term in product for term in SUMMARY_OR_PROMO_TERMS) and "订" not in remark and quantity <= 0:
            return True
        if re.search(r"^(合计|共计|满)\s*\d*", product):
            return True
        if re.fullmatch(r"\d+(?:\.\d+)?(?:\s*[*xX×/+.-]\s*\d+(?:\.\d+)?)+\s*(?:=\s*\d+(?:\.\d+)?)?", product):
            return True
        if re.fullmatch(r"(?:cat\s*no\.?|货号)\s*[:：]?\s*[a-z0-9._-]+", product, flags=re.IGNORECASE):
            return True
        if product in {"这", "这个", "那个", "那就", "给", "满"}:
            return True
        if quantity > 1 and row.get("unit") in {"公斤", "斤"} and "订1" in remark:
            return True
        return False

    def _empty_order_row(self) -> dict[str, Any]:
        return {
            "date": "",
            "time": "",
            "name": "",
            "owner": "",
            "receiver": "",
            "campus": "",
            "order_unit": "",
            "product_name": "",
            "quantity": "",
            "unit": "",
            "spec": "",
            "brand": "",
            "cost_price": "",
            "sale_price": "",
            "total_cost": "",
            "total_sale": "",
            "remark": "",
            "evidence_text": "",
            "record_type": "order_item",
            "confidence": 0.0,
            "needs_review": True,
            "evidence_message_ids": [],
            "risk_flags": [],
        }

    def _estimate_confidence(self, row: dict[str, Any]) -> float:
        confidence = 0.35
        if row.get("product_name"):
            confidence += 0.2
        if row.get("brand"):
            confidence += 0.05
        if row.get("quantity"):
            confidence += 0.15
        if row.get("unit"):
            confidence += 0.1
        if row.get("sale_price"):
            confidence += 0.1
        if row.get("name"):
            confidence += 0.05
        if row.get("owner"):
            confidence += 0.05
        return max(0.0, min(confidence, 0.95))

    def _confidence_fill_for_row(self, row: dict[str, Any]) -> PatternFill | None:
        risk_flags = {str(flag) for flag in row.get("risk_flags", []) if str(flag)}
        if risk_flags & EXPORT_RED_RISK_FLAGS:
            return CONFIDENCE_COLOR_RED
        if risk_flags & EXPORT_YELLOW_RISK_FLAGS:
            return CONFIDENCE_COLOR_YELLOW
        confidence = self._coerce_confidence(row.get("confidence"))
        if confidence >= 0.95:
            return None
        if confidence >= 0.80:
            return CONFIDENCE_COLOR_GREEN
        if confidence >= 0.60:
            return CONFIDENCE_COLOR_YELLOW
        return CONFIDENCE_COLOR_RED

    def _classify_record_type(self, row: dict[str, Any], *, source_text: str = "") -> str:
        product = self._normalize_order_text(str(row.get("product_name") or ""))
        remark = self._normalize_order_text(str(row.get("remark") or ""))
        text = self._normalize_order_text(f"{product} {remark} {source_text}")
        quantity = self._to_number(str(row.get("quantity") or ""))
        total = self._to_number(str(row.get("total_sale") or ""))
        sale_price = self._to_number(str(row.get("sale_price") or ""))
        if any(term in text for term in ("预存", "欠款", "返款", "抵扣", "补差", "结算")) and quantity <= 0:
            return "settlement"
        if self._looks_like_followup_confirmation(text) and not product and quantity <= 0:
            return "confirmation"
        if any(term in text for term in ("赠品", "赠送", "买二送一", "送")):
            return "gift_item"
        if any(term in text for term in ("合计", "共计", "满减", "优惠", "抹零")) and quantity <= 0 and total <= 0:
            return "summary"
        if not product and quantity <= 0 and total <= 0 and sale_price <= 0:
            return "other"
        return "order_item"

    def _infer_default_unit_for_product(self, product_name: str) -> str:
        product = self._normalize_order_text(product_name)
        if not product:
            return "个"
        if re.search(r"(试剂盒|ELISA|kit|Kit)", product, flags=re.IGNORECASE):
            return "盒"
        if "一双" in product or "手套" in product:
            return "双"
        if any(term in product for term in ("滤纸", "打印纸")):
            return "包"
        if any(term in product for term in ("乳液", "液", "乙醇", "乙腈", "甲醇", "正己烷")):
            return "瓶"
        if any(term in product for term in ("管", "抗体", "蛋白", "烧杯", "针头")):
            return "个"
        return "个"

    def _default_single_quantity_for_priced_row(self, row: dict[str, Any]) -> dict[str, Any]:
        output = dict(row)
        product = str(output.get("product_name") or "").strip()
        if not product:
            return output
        sale_price = self._to_number(str(output.get("sale_price") or ""))
        total_sale = self._to_number(str(output.get("total_sale") or ""))
        if sale_price <= 0 and total_sale <= 0:
            return output
        if not str(output.get("quantity") or "").strip():
            output["quantity"] = "1"
            output["needs_review"] = True
            note = "数量策略: 默认1"
            remark = str(output.get("remark") or "")
            if note not in remark:
                output["remark"] = (remark + f" | {note}").strip(" |")
        if not str(output.get("unit") or "").strip():
            output["unit"] = str(output.get("order_unit") or "") or self._infer_default_unit_for_product(product)
            output["needs_review"] = True
        if not str(output.get("order_unit") or "").strip():
            output["order_unit"] = str(output.get("unit") or "") or "个"
        if not str(output.get("total_sale") or "").strip() and sale_price > 0 and self._to_number(str(output.get("quantity") or "")) > 0:
            output["total_sale"] = self._format_number(sale_price * self._to_number(str(output.get("quantity") or "")))
        return output

    def _apply_missing_quantity_strategy(
        self,
        row: dict[str, Any],
        *,
        source_text: str,
        strategy: str = "strict",
    ) -> dict[str, Any]:
        output = dict(row)
        quantity_text = str(output.get("quantity") or "").strip()
        if quantity_text:
            return output
        strategy_key = str(strategy or "strict").strip().lower()
        if strategy_key != "assume_one":
            return output
        text = self._normalize_order_text(source_text)
        if not text:
            return output
        if not (self._to_number(str(output.get("sale_price") or "")) > 0 or self._to_number(str(output.get("total_sale") or "")) > 0):
            return output
        qty, unit = self._extract_quantity_unit(text)
        if self._to_number(qty) > 0:
            output["quantity"] = self._format_number(self._to_number(qty))
            if unit and not str(output.get("unit") or "").strip():
                output["unit"] = unit
            if unit and not str(output.get("order_unit") or "").strip():
                output["order_unit"] = unit
        else:
            output["quantity"] = "1"
            if not str(output.get("unit") or "").strip():
                output["unit"] = str(output.get("order_unit") or "") or "个"
            if not str(output.get("order_unit") or "").strip():
                output["order_unit"] = str(output.get("unit") or "") or "个"
            note = "数量策略: 默认1"
            remark = str(output.get("remark") or "")
            if note not in remark:
                output["remark"] = (remark + f" | {note}").strip(" |")
        sale = self._to_number(str(output.get("sale_price") or ""))
        total = self._to_number(str(output.get("total_sale") or ""))
        qty_value = self._to_number(str(output.get("quantity") or ""))
        if qty_value > 0 and sale > 0 and total <= 0:
            output["total_sale"] = self._format_number(qty_value * sale)
        output["needs_review"] = bool(
            float(output.get("confidence") or 0) < 0.75
            or not str(output.get("product_name") or "").strip()
            or not str(output.get("quantity") or "").strip()
        )
        return output

    def _finalize_order_row(
        self,
        row: dict[str, Any],
        message: dict[str, Any],
        *,
        date_output_mode: str = "YYYY-MM-DD",
    ) -> dict[str, Any]:
        output = {**self._empty_order_row(), **row}
        for key, value in list(output.items()):
            if value is None:
                output[key] = ""
        output["confidence"] = self._coerce_confidence(output.get("confidence"))
        source_time = recorder_time_text(message)
        fallback_date = self._extract_date_code(source_time)
        fallback_time = self._extract_time_code(source_time)
        risk_flags = {
            str(flag)
            for flag in (output.get("risk_flags") or [])
            if str(flag)
        }
        risk_flags.update(str(flag) for flag in (message.get("risk_flags") or []) if str(flag))
        risk_flags.update(str(flag) for flag in recorder_message_view(message).get("risk_flags", []) if str(flag))
        output["date"] = self._normalize_order_date(
            output.get("date"),
            fallback_date=fallback_date,
            date_output_mode=date_output_mode,
        )
        if not str(output.get("time") or "").strip() and fallback_time:
            output["time"] = fallback_time
        if not str(output.get("name") or "").strip() or not str(output.get("owner") or "").strip():
            fallback_name, fallback_owner = self._extract_name_owner(str(message.get("content") or ""))
            if not str(output.get("name") or "").strip() and fallback_name:
                output["name"] = fallback_name
            if not str(output.get("owner") or "").strip():
                output["owner"] = fallback_owner or str(output.get("name") or "")
        raw_name = str(output.get("name") or "").strip()
        raw_owner = str(output.get("owner") or "").strip()
        clean_name = self._sanitize_person_name(raw_name)
        clean_owner = self._sanitize_person_name(raw_owner)
        if clean_name and clean_name != raw_name:
            output["name"] = clean_name
            output["needs_review"] = True
        elif raw_name and not clean_name and (PRICE_RE.search(raw_name) or re.search(r"\d", raw_name) or any(marker in raw_name for marker in ("订", "代付", "货号", "型号"))):
            output["name"] = ""
            output["needs_review"] = True
        if clean_owner and clean_owner != raw_owner:
            output["owner"] = clean_owner
            output["needs_review"] = True
        elif raw_owner and not clean_owner and (PRICE_RE.search(raw_owner) or re.search(r"\d", raw_owner) or any(marker in raw_owner for marker in ("订", "代付", "货号", "型号"))):
            output["owner"] = ""
            output["needs_review"] = True
        if not str(output.get("owner") or "").strip() and str(output.get("name") or "").strip():
            output["owner"] = str(output.get("name") or "")
        explicit_scope = self._normalize_order_text(str(output.get("source_scope_text") or ""))
        backfill_scope = explicit_scope or self._normalize_order_text(
            str(output.get("evidence_text") or "")
            or str(message.get("content") or "")
        )
        output = self._backfill_row_from_scope_text(
            output,
            scope_text=backfill_scope,
            source_text=str(message.get("content") or ""),
            strict_scope=bool(explicit_scope),
        )
        if output.get("quantity") and output.get("sale_price") and not output.get("total_sale"):
            total = self._to_number(str(output.get("quantity") or "")) * self._to_number(str(output.get("sale_price") or ""))
            if total:
                output["total_sale"] = self._format_number(total)
        product_name = str(output.get("product_name") or "")
        sanitized_product = self._sanitize_product_name(product_name, scope_text=backfill_scope)
        if sanitized_product != self._normalize_order_text(product_name):
            output["product_name"] = sanitized_product
            output["needs_review"] = True
            output["confidence"] = min(float(output.get("confidence") or 0), 0.78 if sanitized_product else 0.62)
            product_name = sanitized_product
        brand_name = str(output.get("brand") or "")
        remark_text = str(output.get("remark") or "")
        evidence_text = str(output.get("evidence_text") or "")
        scope_text = self._normalize_order_text(str(output.get("source_scope_text") or ""))
        scoped_context_text = scope_text or self._normalize_order_text(evidence_text)
        local_context_text = self._normalize_order_text(f"{scoped_context_text} {remark_text} {product_name}")
        message_text = scope_text or local_context_text or str(message.get("content") or "")
        current_spec = self._normalize_spec_text(str(output.get("spec") or ""))
        candidate_primary = self._extract_product_spec(
            self._normalize_order_text(f"{product_name} {remark_text}"),
            product_name=product_name,
            brand_hint=brand_name,
        )
        candidate_secondary = self._extract_product_spec(
            self._normalize_order_text(f"{product_name} {scoped_context_text} {message_text}"),
            product_name=product_name,
            brand_hint=brand_name,
        )
        context_text = self._normalize_order_text(f"{product_name} {remark_text} {scoped_context_text}")
        options = [current_spec, candidate_primary, candidate_secondary]
        best_spec = ""
        best_score = -1
        for option in options:
            score = self._score_spec_candidate(option, product_name=product_name, context_text=context_text)
            if score > best_score:
                best_score = score
                best_spec = self._normalize_spec_text(option)
        rich_candidates = [
            self._normalize_spec_text(option)
            for option in options
            if self._normalize_spec_text(option)
            and FORMULATION_SPEC_RE.search(self._normalize_spec_text(option))
            and self._extract_primary_spec_code(option)
        ]
        if rich_candidates:
            rich_best = max(
                rich_candidates,
                key=lambda item: (
                    len(self._extract_primary_spec_code(item)),
                    len(item),
                ),
            )
            chosen_code = self._extract_primary_spec_code(best_spec)
            rich_code = self._extract_primary_spec_code(rich_best)
            if rich_code and (not chosen_code or len(rich_code) >= len(chosen_code) + 1):
                best_spec = rich_best
        output["spec"] = best_spec if (best_score > 0 or (best_spec and rich_candidates)) else ""
        if output["spec"]:
            repaired_product, repaired_spec = self._repair_split_spec_prefix_from_product(
                product_name=str(output.get("product_name") or ""),
                spec=str(output.get("spec") or ""),
                context_text=context_text,
            )
            output["product_name"] = repaired_product
            output["spec"] = repaired_spec
            trimmed_product = self._remove_spec_tokens_from_product_name(str(output.get("product_name") or ""), str(output.get("spec") or ""))
            if trimmed_product and trimmed_product != self._normalize_order_text(str(output.get("product_name") or "")):
                output["product_name"] = trimmed_product
        if explicit_scope:
            output = self._clear_unsupported_quantity_from_scope(output, scope_text=explicit_scope)
        output = self._default_single_quantity_for_priced_row(output)
        output["needs_review"] = bool(
            output.get("needs_review")
            or float(output.get("confidence") or 0) < 0.75
            or not output.get("product_name")
            or not output.get("quantity")
        )
        if not output.get("product_name"):
            risk_flags.add("missing_product")
        if not output.get("quantity"):
            risk_flags.add("missing_quantity")
        if not output.get("sale_price") and not output.get("total_sale"):
            risk_flags.add("missing_price")
        if risk_flags:
            output["risk_flags"] = sorted(risk_flags)
            output["needs_review"] = True
            if risk_flags & EXPORT_RED_RISK_FLAGS:
                output["confidence"] = min(float(output.get("confidence") or 0), 0.59)
            elif risk_flags & EXPORT_YELLOW_RISK_FLAGS:
                output["confidence"] = min(float(output.get("confidence") or 0), 0.79)
        ids = [str(item) for item in output.get("evidence_message_ids", []) if str(item)]
        if not ids:
            raw_message_id = str(message.get("raw_message_id") or "")
            if raw_message_id:
                ids = [raw_message_id]
        output["evidence_message_ids"] = ids
        if not str(output.get("evidence_text") or "").strip():
            output["evidence_text"] = self._normalize_order_text(recorder_content_text(message))[:300]
        numeric_fields = ("quantity", "sale_price", "total_sale", "cost_price", "total_cost")
        for field in numeric_fields:
            raw_value = str(output.get(field) or "").strip()
            if not raw_value:
                output[field] = ""
                continue
            numeric_value = self._to_number(raw_value)
            if numeric_value > 0:
                output[field] = self._format_number(numeric_value)
            else:
                output[field] = raw_value
        # 当前版本进价相关字段暂不可靠，统一留空，避免误导。
        output["cost_price"] = ""
        output["total_cost"] = ""
        if str(output.get("unit") or "").strip() and not str(output.get("order_unit") or "").strip():
            output["order_unit"] = str(output.get("unit") or "")
        return output

    def _normalize_order_date(self, value: Any, *, fallback_date: str = "", date_output_mode: str = "YYYY-MM-DD") -> str:
        raw = str(value or "").strip()
        resolved = ""
        if raw:
            mmdd_match = MMDD_RE.match(raw)
            if mmdd_match:
                month = int(mmdd_match.group("month") or 0)
                day = int(mmdd_match.group("day") or 0)
                if month and day:
                    inferred_year = datetime.now().year
                    fallback_match = DATE_RE.search(str(fallback_date or ""))
                    if fallback_match:
                        inferred_year = int(fallback_match.group("year") or inferred_year)
                    resolved = f"{inferred_year:04d}-{month:02d}-{day:02d}"
            else:
                resolved = self._extract_date_code(raw)
        if not resolved:
            resolved = str(fallback_date or "")
        if not resolved:
            return ""
        mode = str(date_output_mode or "YYYY-MM-DD").strip().upper()
        if mode in {"MMDD", "MMDD_TEXT", "MMDD-TEXT"}:
            match = DATE_RE.search(resolved)
            if match:
                return f"{int(match.group('month') or 0):02d}{int(match.group('day') or 0):02d}"
        return resolved

    def _llm_extract_single_row(self, message: dict[str, Any], *, date_output_mode: str = "YYYY-MM-DD") -> dict[str, Any] | None:
        content = str(message.get("content") or "").strip()
        if not content:
            return None
        prompt = {
            "task": "从微信订货聊天中提取一行结构化订货数据。只允许依据source_text，不允许虚构；群成员名和引用历史不是当前订单事实。",
            "source_text": content,
            "rules": [
                "群成员名、发送人姓名不是品牌或产品。",
                "引用/回复/历史预览内容不得作为产品、品牌、规格、价格或姓名证据。",
                "一个气泡是一条独立订单语义范围，不同气泡不能串用品牌、规格、产品或价格。",
            ],
            "required_schema": {
                "date": "日期文本，优先YYYY-MM-DD，可为空",
                "name": "姓名，可为空",
                "owner": "责任人，可为空",
                "receiver": "收货人，可为空",
                "campus": "校区，可为空",
                "order_unit": "订货单位，可为空",
                "product_name": "货品名称，尽量提取",
                "quantity": "数量，可为空",
                "unit": "单位，可为空",
                "spec": "规格，可为空",
                "brand": "品牌，可为空",
                "cost_price": "进价单价，允许空",
                "sale_price": "售价单价，可为空",
                "total_cost": "总进价，允许空",
                "total_sale": "总售价，可为空",
                "remark": "备注，可为空",
                "evidence_text": "证据片段，建议引用原文关键短句",
                "record_type": "order_item|gift_item|summary|settlement|confirmation|other",
                "confidence": "0-1",
            },
            "response_contract": {
                "row": "object",
                "needs_review": "bool",
                "reason": "string",
            },
        }
        payload = self._call_deepseek_json_cached(prompt, namespace="extract_single_row")
        row = payload.get("row") if isinstance(payload.get("row"), dict) else {}
        if not row:
            return None
        result = self._empty_order_row()
        for key in result.keys():
            if key in row and key not in {"needs_review", "evidence_message_ids"}:
                result[key] = row.get(key)
        result["confidence"] = max(self._coerce_confidence(result.get("confidence")), self._coerce_confidence(row.get("confidence")))
        result["needs_review"] = bool(payload.get("needs_review", True))
        result["remark"] = str(result.get("remark") or payload.get("reason") or content[:120])
        if not str(result.get("evidence_text") or "").strip():
            result["evidence_text"] = str(row.get("evidence_text") or content[:180])
        result["evidence_message_ids"] = [str(message.get("raw_message_id") or "")]
        result["source_scope_text"] = self._normalize_order_text(content)
        return self._finalize_order_row(result, message, date_output_mode=date_output_mode)

    def _llm_supplement_row(
        self,
        message: dict[str, Any],
        row: dict[str, Any],
        *,
        date_output_mode: str = "YYYY-MM-DD",
        mode: str = "missing_core_fields_only",
    ) -> dict[str, Any] | None:
        content = str(message.get("content") or "").strip()
        if not content:
            return None
        mode_key = str(mode or "missing_core_fields_only").strip().lower()
        if mode_key == "missing_core_fields_only":
            field_scope = ("product_name", "brand", "quantity", "unit", "sale_price", "total_sale", "name", "owner")
        else:
            field_scope = ("name", "owner", "receiver", "order_unit", "product_name", "quantity", "unit", "sale_price", "total_sale", "spec", "brand")
        missing_fields = [key for key in field_scope if not str(row.get(key) or "").strip()]
        if not missing_fields:
            return None
        prompt = {
            "task": "请基于 source_text 和 rule_row，补全缺失字段。禁止覆盖已有非空字段，禁止虚构。",
            "source_text": content,
            "rule_row": row,
            "missing_fields": missing_fields,
            "response_contract": {
                "fill": {"field": "value"},
                "confidence": "0-1",
                "reason": "string",
            },
        }
        payload = self._call_deepseek_json_cached(prompt, namespace="supplement_row")
        fill = payload.get("fill") if isinstance(payload.get("fill"), dict) else {}
        if not fill:
            return None
        if "quantity" in fill and "unit" in fill and not re.search(r"(订|下单|买|各订|代付)\s*[一二两俩三四五六七八九十百半\d]+", content):
            # 防止把规格(如4kg/500ml)误当采购数量
            fill.pop("quantity", None)
            fill.pop("unit", None)
        updated = dict(row)
        for key, value in fill.items():
            if key in updated and not str(updated.get(key) or "").strip():
                updated[key] = value
        llm_conf = self._coerce_confidence(payload.get("confidence"))
        updated["confidence"] = max(float(updated.get("confidence") or 0), min(0.95, llm_conf * 0.9))
        reason = str(payload.get("reason") or "")
        if reason:
            updated["remark"] = str(updated.get("remark") or "")
            if reason not in updated["remark"]:
                updated["remark"] = (updated["remark"] + f" | LLM补充: {reason}").strip(" |")
        if not str(updated.get("evidence_text") or "").strip():
            updated["evidence_text"] = content[:180]
        return self._finalize_order_row(updated, message, date_output_mode=date_output_mode)

    @staticmethod
    def _to_number(value: str) -> float:
        text = str(value or "").strip()
        if not text:
            return 0.0
        try:
            return float(text)
        except ValueError:
            try:
                cleaned = re.sub(r"[^\d.]+", "", text)
                return float(cleaned) if cleaned else 0.0
            except ValueError:
                return 0.0

    @staticmethod
    def _format_number(value: float) -> str:
        if abs(value - int(value)) < 1e-9:
            return str(int(value))
        return f"{value:.2f}".rstrip("0").rstrip(".")

    @staticmethod
    def _coerce_confidence(value: Any) -> float:
        if isinstance(value, bool):
            return 1.0 if value else 0.0
        if isinstance(value, (int, float)):
            return max(0.0, min(float(value), 1.0))
        if isinstance(value, dict):
            for key in ("confidence", "score", "value"):
                if key in value:
                    return RecorderExportRunService._coerce_confidence(value.get(key))
            return 0.0
        text = str(value or "").strip()
        if not text:
            return 0.0
        try:
            parsed = float(text)
        except ValueError:
            return 0.0
        return max(0.0, min(parsed, 1.0))

    def _upsert_run(self, run: dict[str, Any]) -> dict[str, Any]:
        records = [item for item in self._read_json(self.runs_path, default=[]) if isinstance(item, dict)]
        by_id = {str(item.get("run_id") or ""): item for item in records}
        existing = by_id.get(str(run.get("run_id") or ""), {})
        merged = {**existing, **run}
        merged["updated_at"] = str(run.get("updated_at") or now_iso())
        by_id[str(merged.get("run_id") or "")] = merged
        saved = sorted(by_id.values(), key=lambda item: str(item.get("created_at") or ""), reverse=True)[:MAX_RUN_RECORDS]
        self._write_json(self.runs_path, saved)
        return merged

    def _read_json(self, path: Path, *, default: Any) -> Any:
        if not path.exists():
            return default
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return default
        return payload

    def _write_json(self, path: Path, payload: Any) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        temp = path.with_suffix(path.suffix + ".tmp")
        temp.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        os.replace(temp, path)


def normalize_filters(payload: dict[str, Any]) -> dict[str, Any]:
    now = datetime.now()
    start_time = str(payload.get("start_time") or "").strip()
    end_time = str(payload.get("end_time") or "").strip()
    date_from = str(payload.get("date_from") or "").strip()
    date_to = str(payload.get("date_to") or "").strip()
    quick_range = str(payload.get("quick_range") or "").strip().lower()

    def parse_date_text(value: str) -> datetime | None:
        text = str(value or "").strip()
        if not text:
            return None
        match = re.match(r"^(?P<y>\d{4})-(?P<m>\d{2})-(?P<d>\d{2})$", text)
        if not match:
            return None
        try:
            return datetime(int(match.group("y")), int(match.group("m")), int(match.group("d")))
        except ValueError:
            return None

    def to_range_start(value: datetime) -> str:
        return value.strftime("%Y-%m-%d 00:00:00")

    def to_range_end(value: datetime) -> str:
        return value.strftime("%Y-%m-%d 23:59:59")

    if quick_range and not (start_time or end_time or date_from or date_to):
        if quick_range == "day":
            today = datetime(now.year, now.month, now.day)
            date_from = today.strftime("%Y-%m-%d")
            date_to = date_from
        elif quick_range == "week":
            today = datetime(now.year, now.month, now.day)
            weekday = today.weekday()
            week_start = today - timedelta(days=weekday)
            week_end = week_start + timedelta(days=6)
            date_from = week_start.strftime("%Y-%m-%d")
            date_to = week_end.strftime("%Y-%m-%d")
        elif quick_range == "month":
            month_start = datetime(now.year, now.month, 1)
            if now.month == 12:
                next_month = datetime(now.year + 1, 1, 1)
            else:
                next_month = datetime(now.year, now.month + 1, 1)
            month_end = next_month - timedelta(days=1)
            date_from = month_start.strftime("%Y-%m-%d")
            date_to = month_end.strftime("%Y-%m-%d")

    if date_from and not start_time:
        parsed = parse_date_text(date_from)
        if parsed:
            start_time = to_range_start(parsed)
    if date_to and not end_time:
        parsed = parse_date_text(date_to)
        if parsed:
            end_time = to_range_end(parsed)

    return {
        "conversation_ids": [str(item) for item in payload.get("conversation_ids", []) if str(item)],
        "target_names": [str(item) for item in payload.get("target_names", []) if str(item)],
        "query": str(payload.get("query") or ""),
        "start_time": start_time,
        "end_time": end_time,
        "date_from": date_from,
        "date_to": date_to,
        "quick_range": quick_range,
        "sender": str(payload.get("sender") or ""),
        "content_type": str(payload.get("content_type") or ""),
        "conversation_type": str(payload.get("conversation_type") or ""),
        "keywords": [str(item) for item in payload.get("keywords", []) if str(item)],
        "offset": max(0, int(payload.get("offset", 0) or 0)),
        "limit": max(1, min(int(payload.get("limit", 10000) or 10000), 10000)),
    }
